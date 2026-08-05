# -*- coding: utf-8 -*-
"""tools/parse_facility_beds.py のテスト。

3層構成:
  1. 合成フィクスチャ(openpyxlで組み立てた最小xlsx)によるヘッダー解決・
     センチネル分岐・連番検証・終端検出の単体テスト(実データに依存しない)
  2. 実データ(R7/001723127.xlsx)の前提確認(行数・シート数・record_idの
     整合等)。339シートの再パースはmodule scopeのfixtureで1回だけ行う
  3. 再現性(`build_and_write()` の出力がコミット済みの
     data/processed/facility_*.csv とバイト一致すること)
"""
import json

import openpyxl
import pytest

from tools.lib.layout import expect
from tools.lib.provenance import REPO_ROOT, recorded_hash
from tools.parse_facility_beds import (
    EXPECTED_COLUMNS,
    KNOWN_ISSUES,
    LayoutMismatchError,
    NUM_SHEETS,
    VALUE_STATUS_BLANK,
    VALUE_STATUS_NOT_CALCULATED,
    VALUE_STATUS_NOT_DISCLOSED,
    VALUE_STATUS_NOT_REPORTED,
    VALUE_STATUS_OBSERVED,
    VALUE_STATUS_SOURCE_DASH,
    _classify_metric_cell,
    _parse_sheet_name,
    _raw_header_signature,
    build_and_write,
    load_workbook,
    parse_sheet,
    parse_workbook,
    resolve_facility_columns,
)

PROCESSED_DIR = REPO_ROOT / "data" / "processed"


# =====================================================================
# 1. 合成フィクスチャによる単体テスト
# =====================================================================

# 11〜13行目のヘッダー(結合セルあり)。R7/001723127.xlsx の実測値をそのまま
# 再現する(全339シートで完全一致していることを事前に確認済み)。
HEADER_MERGES = [
    "B11:G13", "H11:H13", "I11:N11", "O11:Q11", "R11:V11", "W11:AH11",
    "I12:I13", "J12:J13", "K12:K13", "L12:L13", "M12:M13", "N12:N13",
    "O12:O13", "P12:P13", "Q12:Q13", "R12:R13", "S12:S13", "T12:T13",
    "U12:U13", "V12:V13", "W12:W13", "X12:X13", "Y12:Y13", "Z12:Z13",
    "AA12:AD12", "AE12:AH12",
]
HEADER_ROW11 = {
    2: "①医療機関名",
    8: "②所在地",
    9: "③一般・療養病床",
    15: "④医師数",
    18: "⑤医療機関機能",
    23: "⑥診療実績（オープンデータ）",
}
HEADER_ROW12 = {
    9: "休棟中等含む計", 10: "高度急性期", 11: "急性期", 12: "回復期", 13: "慢性期", 14: "休棟中等",
    15: "常勤", 16: "非常勤", 17: "100床\n当たり",
    18: "特定機能", 19: "地域支援", 20: "三次救急", 21: "二次救急", 22: "在宅療養支援",
    23: "救急車の\n受入件数", 24: "全身麻酔\n手術件数", 25: "分娩\n件数", 26: "手術総数",
    27: "平均在棟日数", 31: "新規入棟患者",
}
HEADER_ROW13 = {
    27: "高度急性期", 28: "急性期", 29: "回復期", 30: "慢性期",
    31: "高度急性期", 32: "急性期", 33: "回復期", 34: "慢性期",
}

# 医療機関表(14行目〜)の列番号。実データと同じ位置(B,H,I〜AH)。
COL_NAME = 2
COL_MUNICIPALITY = 8
METRIC_COLS = list(range(9, 18)) + list(range(23, 35))  # I〜Q, W〜AH(R〜Vの機能列は除く)
FUNCTION_COLS = list(range(18, 23))  # R〜V


def _build_header_sheet(wb=None, *, sheet_name="101テスト区域"):
    """11〜13行目のヘッダー(結合セル含む)だけを持つワークシートを作る。"""
    if wb is None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(sheet_name)
    for col, text in HEADER_ROW11.items():
        ws.cell(row=11, column=col, value=text)
    for col, text in HEADER_ROW12.items():
        ws.cell(row=12, column=col, value=text)
    for col, text in HEADER_ROW13.items():
        ws.cell(row=13, column=col, value=text)
    for rng in HEADER_MERGES:
        ws.merge_cells(rng)
    return ws


def _write_facility_row(ws, row, seq, *, name="テスト病院", municipality="テスト市", overrides=None):
    """医療機関1行分を書き込む。数値指標は既定で0、機能列は既定でNone。

    `overrides` は {列番号: 値} で個別の列だけ上書きする。
    """
    overrides = overrides or {}
    ws.cell(row=row, column=1, value=seq)
    ws.cell(row=row, column=COL_NAME, value=name)
    ws.cell(row=row, column=COL_MUNICIPALITY, value=municipality)
    for col in METRIC_COLS:
        ws.cell(row=row, column=col, value=overrides.get(col, 0))
    for col in FUNCTION_COLS:
        ws.cell(row=row, column=col, value=overrides.get(col, None))


# --- resolve_facility_columns() ---------------------------------------


def test_resolve_facility_columns_matches_known_layout():
    ws = _build_header_sheet()
    columns = resolve_facility_columns(ws, sheet_name=ws.title)

    assert columns[("name",)] == 2
    assert columns[("municipality",)] == 8
    assert columns[("bed", "休棟中等含む計")] == 9
    assert columns[("bed", "高度急性期")] == 10
    assert columns[("bed", "休棟中等")] == 14
    assert columns[("doctor", "医師数（常勤）")] == 15
    assert columns[("doctor", "医師数（100床当たり）")] == 17
    assert columns[("function", "特定機能")] == 18
    assert columns[("function", "在宅療養支援")] == 22
    assert columns[("clinical_simple", "救急車の受入件数")] == 23
    assert columns[("clinical_simple", "手術総数")] == 26
    assert columns[("clinical_grouped", "平均在棟日数", "高度急性期")] == 27
    assert columns[("clinical_grouped", "平均在棟日数", "慢性期")] == 30
    assert columns[("clinical_grouped", "新規入棟患者", "高度急性期")] == 31
    assert columns[("clinical_grouped", "新規入棟患者", "慢性期")] == 34
    assert set(columns.keys()) == set(EXPECTED_COLUMNS.values())


def test_resolve_facility_columns_detects_unknown_header():
    """未知の見出し(部分一致で誤爆させない)が中断されるか。"""
    ws = _build_header_sheet()
    ws.cell(row=12, column=10, value="改変された見出し")  # J12(高度急性期)
    with pytest.raises(LayoutMismatchError):
        resolve_facility_columns(ws, sheet_name=ws.title)


def test_resolve_facility_columns_rejects_substring_collision():
    """「急性期」が「高度急性期」に部分一致して誤って解決されないか。

    J12(高度急性期)を「急性期」に書き換えると、K12(元々「急性期」)と三つ組が
    完全一致してしまい、同一項目が離れた2列に出現する形になる。三つ組の
    完全一致で判定しているため、これは「未知の見出し」ではなく「同一項目の
    重複出現」として検知されるべき。
    """
    ws = _build_header_sheet()
    ws.cell(row=12, column=10, value="急性期")  # J12を「高度急性期」から書き換え
    with pytest.raises(LayoutMismatchError):
        resolve_facility_columns(ws, sheet_name=ws.title)


def test_resolve_facility_columns_detects_missing_item():
    """結合セルのアンカーが消えて項目が解決できなくなった場合に検知できるか。

    AE12(結合AE12:AH12のアンカー、「新規入棟患者」)を消すと、AE〜AH列4列とも
    12行目のテキストがNoneになり(結合の非アンカー列はもともとNoneのため
    前方補完される値そのものが失われる)、「新規入棟患者」の4項目すべてが
    未知のヘッダーとして検知されるはず。
    """
    ws = _build_header_sheet()
    # `ws.cell(row=..., column=..., value=None)` は openpyxl の仕様上
    # 「valueキーワード省略」と区別されず無視される(no-op)ため、既存の値を
    # 消すには `.value = None` の代入を使う。
    ws.cell(row=12, column=31).value = None  # AE12(新規入棟患者、結合アンカー)を消す
    with pytest.raises(LayoutMismatchError):
        resolve_facility_columns(ws, sheet_name=ws.title)


# --- _classify_metric_cell() -------------------------------------------


def test_classify_metric_cell_observed():
    assert _classify_metric_cell(123, context="x") == (VALUE_STATUS_OBSERVED, 123, None)
    assert _classify_metric_cell(1.5, context="x") == (VALUE_STATUS_OBSERVED, 1.5, None)


def test_classify_metric_cell_blank():
    assert _classify_metric_cell(None, context="x") == (VALUE_STATUS_BLANK, None, None)


def test_classify_metric_cell_source_dash():
    assert _classify_metric_cell("-", context="x") == (VALUE_STATUS_SOURCE_DASH, None, "-")


def test_classify_metric_cell_not_disclosed():
    assert _classify_metric_cell("*", context="x") == (VALUE_STATUS_NOT_DISCLOSED, None, "*")


def test_classify_metric_cell_not_reported():
    assert _classify_metric_cell("未報告", context="x") == (VALUE_STATUS_NOT_REPORTED, None, "未報告")


def test_classify_metric_cell_not_calculated():
    assert _classify_metric_cell("XXX", context="x") == (VALUE_STATUS_NOT_CALCULATED, None, "XXX")


def test_classify_metric_cell_rejects_unknown_sentinel():
    with pytest.raises(LayoutMismatchError):
        _classify_metric_cell("N/A", context="x")


def test_classify_metric_cell_rejects_bool():
    with pytest.raises(LayoutMismatchError):
        _classify_metric_cell(True, context="x")


# --- parse_sheet(): 連番・終端・センチネルの分岐 --------------------------


def _parse_kwargs(columns):
    return dict(
        sheet_name="101テスト区域",
        area_code="0101",
        pref_code="01",
        pref_name="テスト県",
        area_name="テスト区域",
        columns=columns,
    )


def test_parse_sheet_basic_rows_and_observations():
    ws = _build_header_sheet()
    columns = resolve_facility_columns(ws, sheet_name=ws.title)
    _write_facility_row(ws, 14, 1, name="病院A", overrides={9: 100, 23: "-", 18: "特"})
    _write_facility_row(ws, 15, 2, name="病院B", municipality=None, overrides={9: "未報告"})

    result = parse_sheet(ws, **_parse_kwargs(columns))

    assert len(result.basic_rows) == 2
    assert result.facility_count == 2
    assert result.basic_rows[0]["record_id"] == "R7-0101-14"
    assert result.basic_rows[0]["facility_name"] == "病院A"
    assert result.basic_rows[1]["municipality"] is None

    # 21指標 × 2施設 = 42行
    assert len(result.observation_rows) == 21 * 2

    obs_a = {(r["metric"], r["bed_function"]): r for r in result.observation_rows if r["record_id"] == "R7-0101-14"}
    assert obs_a[("病床数", "休棟中等含む計")]["value"] == 100
    assert obs_a[("病床数", "休棟中等含む計")]["value_status"] == VALUE_STATUS_OBSERVED
    assert obs_a[("救急車の受入件数", "")]["value_status"] == VALUE_STATUS_SOURCE_DASH

    obs_b = {(r["metric"], r["bed_function"]): r for r in result.observation_rows if r["record_id"] == "R7-0101-15"}
    assert obs_b[("病床数", "休棟中等含む計")]["value_status"] == VALUE_STATUS_NOT_REPORTED

    # 医療機関機能: 病院Aのみ「特定機能」に該当(値'特')、病院Bは機能なし
    assert len(result.function_rows) == 1
    assert result.function_rows[0]["record_id"] == "R7-0101-14"
    assert result.function_rows[0]["function_name"] == "特定機能"
    assert result.function_rows[0]["source_literal"] == "特"


def test_parse_sheet_detects_sequence_gap():
    ws = _build_header_sheet()
    columns = resolve_facility_columns(ws, sheet_name=ws.title)
    _write_facility_row(ws, 14, 1)
    _write_facility_row(ws, 15, 3)  # 2が欠番
    with pytest.raises(LayoutMismatchError):
        parse_sheet(ws, **_parse_kwargs(columns))


def test_parse_sheet_detects_duplicate_sequence():
    ws = _build_header_sheet()
    columns = resolve_facility_columns(ws, sheet_name=ws.title)
    _write_facility_row(ws, 14, 1)
    _write_facility_row(ws, 15, 1)  # 重複
    with pytest.raises(LayoutMismatchError):
        parse_sheet(ws, **_parse_kwargs(columns))


def test_parse_sheet_detects_stray_integer_after_end():
    """終了行より後にA列へ整数が再出現した場合に検知できるか(書式だけの
    空行と本物の行の取り違え防止)。"""
    ws = _build_header_sheet()
    columns = resolve_facility_columns(ws, sheet_name=ws.title)
    _write_facility_row(ws, 14, 1)
    # 15行目は空行(A列None)のまま飛ばし、16行目に整数A列が再出現
    ws.cell(row=16, column=1, value=99)
    with pytest.raises(LayoutMismatchError):
        parse_sheet(ws, **_parse_kwargs(columns))


def test_parse_sheet_rejects_blank_facility_name():
    ws = _build_header_sheet()
    columns = resolve_facility_columns(ws, sheet_name=ws.title)
    _write_facility_row(ws, 14, 1, name=None)
    with pytest.raises(LayoutMismatchError):
        parse_sheet(ws, **_parse_kwargs(columns))


def test_parse_sheet_rejects_unknown_sentinel_in_metric_cell():
    ws = _build_header_sheet()
    columns = resolve_facility_columns(ws, sheet_name=ws.title)
    _write_facility_row(ws, 14, 1, overrides={9: "N/A"})
    with pytest.raises(LayoutMismatchError):
        parse_sheet(ws, **_parse_kwargs(columns))


def test_parse_sheet_rejects_unknown_function_cell_value():
    ws = _build_header_sheet()
    columns = resolve_facility_columns(ws, sheet_name=ws.title)
    _write_facility_row(ws, 14, 1, overrides={18: 123})  # 機能列に数値(不正)
    with pytest.raises(LayoutMismatchError):
        parse_sheet(ws, **_parse_kwargs(columns))


# --- _parse_sheet_name() -------------------------------------------------


def test_parse_sheet_name_splits_code_and_rest():
    assert _parse_sheet_name("101北海道南渡島") == ("0101", "北海道南渡島")
    assert _parse_sheet_name("4705沖縄県八重山") == ("4705", "沖縄県八重山")


def test_parse_sheet_name_rejects_no_leading_digits():
    with pytest.raises(LayoutMismatchError):
        _parse_sheet_name("北海道南渡島")


# --- KNOWN_ISSUES(既知のデータ品質問題)の構造 ---------------------------


def test_known_issues_have_required_shape():
    """KNOWN_ISSUESの各エントリが `tools/parse_area_beds.py` と同じ規約
    (id/scope/summary/evidence/action)に沿っているか(合成データ不要、実データの
    ロードもしない軽量チェック)。
    """
    assert KNOWN_ISSUES  # 空では意味がない
    ids = [issue["id"] for issue in KNOWN_ISSUES]
    assert len(ids) == len(set(ids))  # idが重複しない
    for issue in KNOWN_ISSUES:
        assert set(issue.keys()) >= {"id", "scope", "summary", "evidence", "action"}
        assert isinstance(issue["evidence"], list) and issue["evidence"]
        assert isinstance(issue["summary"], str) and issue["summary"]
        assert isinstance(issue["action"], str) and issue["action"]


# =====================================================================
# 2. 実データ(R7/001723127.xlsx)の前提確認
# =====================================================================


@pytest.fixture(scope="module")
def workbook():
    wb, source_sha256 = load_workbook()
    return {"wb": wb, "source_sha256": source_sha256}


@pytest.fixture(scope="module")
def parsed(workbook):
    """339シート全体のパース結果。module scopeで1回だけ実行する
    (このファイル内の実データテストが共有する)。
    """
    return parse_workbook(workbook["wb"])


def test_sheet_count_is_339(workbook):
    assert len(workbook["wb"].sheetnames) == NUM_SHEETS == 339


def test_facility_basic_row_count(parsed):
    assert len(parsed.basic_rows) == 11760


def test_facility_observations_row_count(parsed):
    # 11,760施設 × 21指標(病床6+医師3+診療実績単発4+平均在棟日数4+新規入棟患者4)
    assert len(parsed.observation_rows) == 11760 * 21 == 246960


def test_facility_functions_row_count(parsed):
    assert len(parsed.function_rows) == 7574


def test_value_status_breakdown(parsed):
    from collections import Counter

    counts = Counter(r["value_status"] for r in parsed.observation_rows)
    assert counts == {
        VALUE_STATUS_OBSERVED: 169975,
        VALUE_STATUS_SOURCE_DASH: 70227,
        VALUE_STATUS_NOT_DISCLOSED: 3312,
        VALUE_STATUS_BLANK: 3284,
        VALUE_STATUS_NOT_REPORTED: 162,
    }
    # 'XXX'(not_calculated)は本ファイルの実測データには存在しない
    assert VALUE_STATUS_NOT_CALCULATED not in counts


def test_function_name_breakdown(parsed):
    from collections import Counter

    counts = Counter(r["function_name"] for r in parsed.function_rows)
    assert counts == {
        "二次救急": 3406,
        "在宅療養支援": 3027,
        "地域支援": 716,
        "三次救急": 337,
        "特定機能": 88,
    }


def test_not_reported_only_appears_on_bed_total_metric(parsed):
    """'未報告'(value_status='not_reported')は③一般・療養病床「休棟中等含む計」
    (metric='病床数', bed_function='休棟中等含む計')列にのみ出現する
    (実測結果。ドキュメント化した前提の回帰テスト)。
    """
    rows = [r for r in parsed.observation_rows if r["value_status"] == VALUE_STATUS_NOT_REPORTED]
    assert len(rows) == 162
    assert {(r["metric"], r["bed_function"]) for r in rows} == {("病床数", "休棟中等含む計")}


def test_not_disclosed_only_appears_on_clinical_simple_metrics(parsed):
    """'*'(value_status='not_disclosed')は診療実績4指標(救急車の受入件数・
    全身麻酔手術件数・分娩件数・手術総数)にのみ出現する(実測結果)。
    """
    rows = [r for r in parsed.observation_rows if r["value_status"] == VALUE_STATUS_NOT_DISCLOSED]
    assert len(rows) == 3312
    assert {r["metric"] for r in rows} <= {"救急車の受入件数", "全身麻酔手術件数", "分娩件数", "手術総数"}


def test_record_id_referential_integrity(parsed):
    basic_ids = {r["record_id"] for r in parsed.basic_rows}
    assert len(basic_ids) == len(parsed.basic_rows)  # 重複がない
    assert all(r["record_id"] in basic_ids for r in parsed.observation_rows)
    assert all(r["record_id"] in basic_ids for r in parsed.function_rows)


def test_spot_first_facility(parsed):
    row = next(r for r in parsed.basic_rows if r["record_id"] == "R7-0101-14")
    assert row["facility_name"] == "市立函館病院"
    assert row["municipality"] == "函館市"
    assert row["area_name"] == "南渡島"
    assert row["pref_name"] == "北海道"
    assert row["source_sheet"] == "101北海道南渡島"
    assert row["facility_seq"] == 1

    obs = {
        (r["metric"], r["bed_function"]): r
        for r in parsed.observation_rows
        if r["record_id"] == "R7-0101-14"
    }
    assert obs[("病床数", "休棟中等含む計")]["value"] == 582
    assert obs[("病床数", "高度急性期")]["value"] == 292
    assert obs[("医師数（常勤）", "")]["value"] == 130

    funcs = {r["function_name"] for r in parsed.function_rows if r["record_id"] == "R7-0101-14"}
    assert funcs == {"地域支援", "三次救急", "二次救急"}


def test_summary_mismatch_matches_known_issues(parsed):
    """区域サマリ(医療機関数F+G)との非fatal突合の実測値がKNOWN_ISSUESの記述
    (76シートは未報告件数で説明可能、2シートのみ原因不明)と一致するか。
    `parsed`(module scopeのfixture)を再利用し、追加のワークブックロードはしない。
    """
    from collections import Counter

    assert parsed.summary_matches == 261
    assert len(parsed.summary_mismatches) == 78
    assert parsed.summary_unavailable == 0

    not_reported_by_area = Counter(
        r["area_code"]
        for r in parsed.observation_rows
        if r["value_status"] == VALUE_STATUS_NOT_REPORTED
    )
    area_by_sheet = {r["source_sheet"]: r["area_code"] for r in parsed.basic_rows}

    unexplained = []
    for sheet_name, expected, actual in parsed.summary_mismatches:
        diff = actual - expected
        not_reported_count = not_reported_by_area.get(area_by_sheet[sheet_name], 0)
        if diff != not_reported_count:
            unexplained.append(sheet_name)

    assert unexplained == ["1503新潟県県央", "2007長野県松本"]


def test_min_and_max_facility_count_areas(parsed):
    from collections import Counter

    counts = Counter(r["area_code"] for r in parsed.basic_rows)
    assert counts["2006"] == 1  # 長野県木曽(最小)
    assert counts["0104"] == 333  # 北海道札幌(最大)
    assert len(counts) == 339


def test_area_code_matches_area_basic(parsed):
    import csv

    # area_basic.csvはR6/R7がpublished_fyで並存する(M9)ため、facility_basic.csv
    # (R7のみのファイル001723127.xlsx由来)と突き合わせるにはR7行だけに絞り込む
    # (pref_name/area_nameの値自体はR6/R7で同一だが、辞書の後勝ちでR6行に依存
    # する状態を避ける)。
    reference = {}
    with open(PROCESSED_DIR / "area_basic.csv", encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            if row["published_fy"] != "R7":
                continue
            reference[row["area_code"]] = (row["pref_name"], row["area_name"])

    area_codes = {r["area_code"] for r in parsed.basic_rows}
    assert area_codes == set(reference.keys())
    by_area = {r["area_code"]: r for r in parsed.basic_rows}
    for area_code, (pref_name, area_name) in reference.items():
        row = by_area[area_code]
        assert row["pref_name"] == pref_name
        assert row["area_name"] == area_name


# --- レイアウト崩れ検知(合成フィクスチャ。実データの追加ロードはしない) -------
#
# `parse_workbook()` を実データ(339シート、1ロード20〜35秒)でもう一度開き直す
# のではなく、検知ロジックそのものを直接テストする(レビュー指摘1-2)。
# シート数チェックは `parse_workbook()` がワークブック取得直後に行う最初の処理
# (ヘッダーが有効である必要すらない)なので合成ワークブックで直接再現でき、
# ヘッダー差異検知は `_raw_header_signature()` の比較結果を直接検証すれば足りる
# (本番コードは `expect()` でこの比較結果を使って例外を送出するだけであり、
# `expect()` 自体の送出動作はここでも実際に使って確認する)。


def test_sheet_count_mismatch_is_detected():
    """シート数が339でない場合に検知できるか(合成ワークブック、2シートのみ)。

    `parse_workbook()` は `wb.sheetnames` を取得した直後に `NUM_SHEETS`(339)との
    一致を検証するため、ヘッダー等の中身が有効である必要はない。
    """
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:  # openpyxl.Workbook()既定の空シートを除去
        del wb["Sheet"]
    _build_header_sheet(wb, sheet_name="101テスト区域A")
    _build_header_sheet(wb, sheet_name="102テスト区域B")
    assert len(wb.sheetnames) == 2 != NUM_SHEETS

    with pytest.raises(LayoutMismatchError):
        parse_workbook(wb)


def test_header_drift_is_detected():
    """あるシートのヘッダーが先頭シートと異なる場合に、`_raw_header_signature()`の
    比較で検知できるか(合成シート2枚、片方だけ改変)。

    `parse_workbook()` は全シートの `_raw_header_signature()` を先頭シートと
    `expect()` で比較しており(不一致ならLayoutMismatchError)、ここではその
    比較経路を合成データで直接再現する。
    """
    ws_a = _build_header_sheet(sheet_name="101テスト区域A")
    wb = ws_a.parent
    ws_b = _build_header_sheet(wb, sheet_name="102テスト区域B")

    # 改変前は両シートとも `_build_header_sheet()` で同一に組み立てているため一致する。
    assert _raw_header_signature(ws_a) == _raw_header_signature(ws_b)

    ws_b.cell(row=12, column=10, value="改変されたヘッダー")  # J12(高度急性期)
    sig_a = _raw_header_signature(ws_a)
    sig_b_modified = _raw_header_signature(ws_b)
    assert sig_a != sig_b_modified

    with pytest.raises(LayoutMismatchError):
        expect(sig_a, sig_b_modified, "テスト: ヘッダー(1行目のタイトル・2,3,10,11〜13行目)が先頭シートと不一致")


# =====================================================================
# 3. 再現性(バイト一致)
# =====================================================================

CSV_NAMES = [
    "facility_basic.csv",
    "facility_observations.csv",
    "facility_functions.csv",
]


def test_reproducibility_byte_identical(tmp_path):
    paths = build_and_write(tmp_path)
    assert paths.keys() == {"basic", "observations", "functions"}
    expected_sha256 = recorded_hash("R7/001723127.xlsx")

    for name in CSV_NAMES:
        committed_path = PROCESSED_DIR / name
        assert committed_path.exists(), (
            f"{committed_path} が存在しません"
            "(先に `python tools/parse_facility_beds.py` を実行してください)"
        )
        new_bytes = (tmp_path / name).read_bytes()
        old_bytes = committed_path.read_bytes()
        assert new_bytes == old_bytes, f"{name} がコミット済みデータとバイト一致しません"

        new_meta = json.loads((tmp_path / f"{name}.meta.json").read_text(encoding="utf-8"))
        old_meta = json.loads((PROCESSED_DIR / f"{name}.meta.json").read_text(encoding="utf-8"))

        assert new_meta["source"]["source_sha256"] == expected_sha256
        assert old_meta["source"]["source_sha256"] == expected_sha256

        # processing.date は実行日ごとに変わるため、比較対象から除外する
        new_meta["processing"]["date"] = None
        old_meta["processing"]["date"] = None
        assert new_meta == old_meta, f"{name}.meta.json の内容(processing.dateを除く)が一致しません"
