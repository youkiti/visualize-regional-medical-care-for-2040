# -*- coding: utf-8 -*-
"""厚労省「構想区域別の医療機関の病床機能報告上の病床数、診療実績、医師数等」
(R7: 001723127.xlsx、別添５②)を tidy CSV へ変換する。

他のパーサ(`tools/parse_area_beds.py`・`tools/parse_prefecture_beds.py`)が
「1シート内で区域ごとのブロックが繰り返される」帳票を扱うのに対し、この
xlsxは逆に「339構想区域それぞれが専用の1シート」という構造(339シート、
シート名は「101北海道南渡島」=区域コード+都道府県名+区域名の形式)。その
ためブロック走査基盤(`tools/lib/block_report.py`)は使わず、シート単位の
走査を本ファイルに実装する。ヘッダーの結合セル解決・非数値センチネルの
分類・レイアウト崩れ検知は `tools/lib/layout.py` の共通ユーティリティ
(`LayoutMismatchError`/`expect`/`normalize_header_text`)を利用する。

原典の構造(339シート共通、実測で確認済み):
  - 1行目=タイトル、2行目=注記(4本)
  - 3行目「(1)構想区域の状況」、4〜6行=そのヘッダー、7〜9行=その値
    (区域サマリ。本パーサの出力対象外。7行目の医療機関数・病床数計は
    非fatalな整合性チェックにのみ使う。下記「区域サマリとの非fatal突合」参照)
  - 10行目「(2)区域内の医療機関（病床数の多い順）」
  - 11〜13行目=医療機関表の3段見出し(結合セルあり)
  - 14行目〜シート末尾=医療機関1件=1行。A列が1始まりの連番(件数はシートに
    より1〜300件超まで様々)

医療機関表の列構成(ヘッダー文字列から解決。列位置はハードコードしない):
  B(結合B:G)=①医療機関名 / H=②所在地(市区町村。政令市は「札幌市北区」、
  特別区は「文京区」形式) / I〜N=③一般・療養病床(休棟中等含む計・高度
  急性期・急性期・回復期・慢性期・休棟中等) / O〜Q=④医師数(常勤・非常勤・
  100床当たり) / R〜V=⑤医療機関機能(特定機能・地域支援・三次救急・二次
  救急・在宅療養支援。セルの原文は'特'/'地'/'三次'/'二次'/'在支'等の略記) /
  W〜Z=⑥診療実績(救急車の受入件数・全身麻酔手術件数・分娩件数・手術総数) /
  AA〜AD=平均在棟日数(高度急性期・急性期・回復期・慢性期) / AE〜AH=新規
  入棟患者(同4区分)。

処理内容:
  1. `verify_source()` で元データのSHA-256を `SHA256SUMS` と照合(改変検知)
  2. openpyxl(`data_only=True`)でワークブックを開き、シート数がちょうど339
     であることを検証
  3. シート名を「先頭の数値(区域コード)+残り」に分解し、区域コードを
     ゼロ埋め4桁へ正規化した上で `data/processed/area_basic.csv` と突合
     (都道府県名+区域名の連結がシート名の残り部分と一致することを検証。
     可変長の都道府県名を切り出して比較するのではなく、area_basic.csv側
     から期待表示名を組み立てて比較する)
  4. 11〜13行目の医療機関表ヘッダーを結合セル対応(前方補完)で解決し、
     339シート全てが先頭シートと完全一致することを検証する。列の識別は
     (11行目, 12行目, 13行目)の文字列の三つ組の完全一致で行うため、
     「急性期」が「高度急性期」に部分一致して誤った列を拾うことはない
  5. 各シートの医療機関行(14行目〜)をA列の連番(1始まり・欠番/重複なし)が
     途切れるまで走査し、それ以降にA列へ整数が再出現しないことを確認する
     (書式だけの空行と本物の行を取り違えないため)
  6. 数値以外のセルは '-'(source_dash)・空欄(blank)・'*'(not_disclosed、
     NDB非公表)・'未報告'(not_reported、未報告の医療機関)・'XXX'
     (not_calculated、他帳票の前例に備えた分岐)のいずれかであることを
     確認し、それ以外はレイアウト変更とみなして中断する(静かに欠測へ
     落とさない)
  7. 3つの tidy CSV + 由来メタデータ(`<csv名>.meta.json`)を `data/processed/`
     に出力する
       - facility_basic.csv: 1行=1医療機関の識別情報
       - facility_observations.csv: 病床数・医師数・診療実績等の数値指標(long)
       - facility_functions.csv: 医療機関機能(該当した行のみ、long)

⚠ 非数値センチネルの実測結果(本ファイル固有。CLAUDE.md「センチネル値の罠」
に類例があるが、本ファイルでは'XXX'ではなく'*'と'未報告'が実際に出現する):
  - '*'(not_disclosed、3,312セル): 診療実績4列(救急車の受入件数・全身麻酔
    手術件数・分娩件数・手術総数)にのみ出現。原典2行目注記3「NDBの利用に
    関するガイドラインを踏まえ、病床機能報告の報告結果のうち一部非公表と
    しているものがあり」に対応する
  - '未報告'(not_reported、162セル): ③一般・療養病床「休棟中等含む計」列
    (I列)にのみ出現。原典2行目注記4「未報告の医療機関があり得ることに
    留意」に対応する。この場合、所在地・病床機能別内訳・医師数・機能・
    診療実績等ほぼ全ての付随項目が原典側で空欄になる
  - 'XXX'(not_calculated)は本ファイルの実測データでは一度も出現しないが、
    `001723349.xlsx`(構想区域の病床数等)の推計流出入患者割合での前例が
    あるため、分岐だけ用意している(CLAUDE.md「センチネル値の罠」参照)

⚠ 区域サマリ(7行目)との非fatal突合(実測結果。fatal検証には**しない**。
理由は下記の通り、定義の違い・未報告医療機関の扱いの違いで説明できる
差異であり、本パーサの出力(facility_*.csv)自体には影響しないため):
  - 医療機関数(F列「一般病院」+G列「有床診療所」) と 医療機関行数:
    339シート中78シートで不一致。76シートは「未報告」医療機関の件数と
    完全に一致する差分(未報告医療機関は一覧には載るがF+G集計には含まれて
    いないとみられる)。残り2シート(1503新潟県県央, 2007長野県松本)のみ
    +1件の差異が未報告件数で説明できず、原因不明(値は原典どおり出力する)
  - 病床数計(I列)と 医療機関行I列の合計: 339シート中304シートで不一致
    だが、個別機能別(J〜N列=高度急性期・急性期・回復期・慢性期・休棟中等)
    は339シート全てで完全一致する。差分は常に医療機関側N列(休棟中等)の
    合計と一致しており、区域サマリのI列(④一般・療養病床計「休棟中等
    "除く"」)が休棟中等を含まないのに対し、医療機関表のI列(③一般・療養
    病床「休棟中等"含む"計」)は休棟中等を含む、という定義の違いで完全に
    説明できる(バグではない)

必要環境: Python 3.11+, openpyxl

使い方:
    python tools/parse_facility_beds.py

⚠ R6について: R6にも対応する帳票(`R6/別添５②（構想区域の詳細状況）.xlsx`)
が存在するが、本パーサはR7のみに対応する(R6のレイアウト調査・対応は
本チャンクの対象外)。
"""
import csv
import datetime
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

if __package__ in (None, ""):
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

from tools.lib.codes import normalize_area_code, normalize_pref_code
from tools.lib.layout import LayoutMismatchError, expect, normalize_header_text
from tools.lib.provenance import REPO_ROOT, verify_source, write_csv_with_meta

SOURCE_NAME = "③構想区域の医療機関の病床数、診療実績等（別添５）"
SOURCE_PATH_IN_REPO = "R7/001723127.xlsx"
SOURCE_DOWNLOAD_URL = "https://www.mhlw.go.jp/content/10800000/001723127.xlsx"
SOURCE_PAGE_URL = "https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/0000080850_00014.html"
SOURCE_FISCAL_YEAR = "令和7年度（2025年度）"
SOURCE_ACQUIRED_DATE = "2026-08-04"

LICENSE_NOTE = (
    "厚生労働省ホームページ利用規約（政府標準利用規約準拠） "
    "https://www.mhlw.go.jp/chosakuken/index.html"
)

PUBLISHED_FY = "R7"
NUM_SHEETS = 339

TITLE_ROW = 1
NOTES_ROW = 2
SECTION1_ROW = 3   # "(1)構想区域の状況"
SECTION2_ROW = 10  # "(2)区域内の医療機関（病床数の多い順）"
HEADER_ROWS = (11, 12, 13)
DATA_START_ROW = 14
MAX_COL = 34  # AH列(新規入棟患者・慢性期)

# 区域サマリ(7行目)の医療機関数。非fatal突合(`_summary_soft_check`)専用。
SUMMARY_ROW = 7
SUMMARY_COL_HOSPITAL = 6  # F列: 一般病院
SUMMARY_COL_CLINIC = 7    # G列: 有床診療所

LEVEL1_NAME = "①医療機関名"
LEVEL1_MUNICIPALITY = "②所在地"
LEVEL1_BED = "③一般・療養病床"
LEVEL1_DOCTOR = "④医師数"
LEVEL1_FUNCTION = "⑤医療機関機能"
LEVEL1_CLINICAL = "⑥診療実績（オープンデータ）"

BED_FUNCTIONS_FACILITY = ["休棟中等含む計", "高度急性期", "急性期", "回復期", "慢性期", "休棟中等"]
DOCTOR_LABELS = ["常勤", "非常勤", "100床当たり"]
DOCTOR_METRIC_NAMES = {
    "常勤": "医師数（常勤）",
    "非常勤": "医師数（非常勤）",
    "100床当たり": "医師数（100床当たり）",
}
FUNCTION_NAMES = ["特定機能", "地域支援", "三次救急", "二次救急", "在宅療養支援"]
CLINICAL_SIMPLE_METRICS = ["救急車の受入件数", "全身麻酔手術件数", "分娩件数", "手術総数"]
CLINICAL_GROUPED_METRICS = ["平均在棟日数", "新規入棟患者"]
CLINICAL_GROUPED_BED_FUNCTIONS = ["高度急性期", "急性期", "回復期", "慢性期"]

METRIC_BED = "病床数"

# value_status の分類(facility_observations.csv)。
VALUE_STATUS_OBSERVED = "observed"
VALUE_STATUS_SOURCE_DASH = "source_dash"
VALUE_STATUS_BLANK = "blank"
VALUE_STATUS_NOT_CALCULATED = "not_calculated"  # 'XXX'。本ファイルの実測では未検出
VALUE_STATUS_NOT_DISCLOSED = "not_disclosed"    # '*'。NDB非公表
VALUE_STATUS_NOT_REPORTED = "not_reported"      # '未報告'。病床機能報告未報告

SHEET_NAME_RE = re.compile(r"^(\d+)(.+)$")

PREF_CODE_DESC = "都道府県コード(ゼロ埋め2桁の文字列、01=北海道…47=沖縄県)。構想区域コードの上2桁から算出"
PREF_NAME_DESC = "都道府県名(data/processed/area_basic.csvより)"
AREA_CODE_DESC = "構想区域(二次医療圏)コード(ゼロ埋め4桁の文字列)。シート名先頭の数値から正規化"
AREA_NAME_DESC = "構想区域名(data/processed/area_basic.csvより)"
PUBLISHED_FY_DESC = "公表年度を表す識別子。'R7'=令和7年度公表分"
RECORD_ID_DESC = (
    "この医療機関個票行を一意に識別するID。"
    "\"{published_fy}-{area_code}-{source_row}\" の形式(例 'R7-0101-14')。"
    "**恒久的な施設IDではない**: 原典の行位置(区域内で病床数の多い順)に"
    "由来するため、医療機関の増減・順位変動があれば同じ行位置でも別の"
    "施設を指すようになり、また公表年度が変われば同じIDが別施設を指し"
    "うる。名称・所在地からのハッシュ生成でもない(改称でIDが変わる、"
    "同名施設で衝突する、を避けるため)。年度をまたいだ施設の同定には"
    "使えないことに留意"
)

FIELDS_BASIC = {
    "published_fy": PUBLISHED_FY_DESC,
    "pref_code": PREF_CODE_DESC,
    "pref_name": PREF_NAME_DESC,
    "area_code": AREA_CODE_DESC,
    "area_name": AREA_NAME_DESC,
    "record_id": RECORD_ID_DESC,
    "source_sheet": "原典のシート名(例 '101北海道南渡島' = 区域コード+都道府県名+区域名)",
    "source_row": "原典シート内の行番号(1始まり)。record_idの構成要素",
    "facility_seq": "原典A列の連番(1始まり、区域内で病床数の多い順)。シート内でのみ意味を持つ",
    "facility_name": "医療機関名",
    "municipality": (
        "所在地(市区町村名)。政令指定都市は区名まで('札幌市北区'等)、"
        "特別区は区名のみ('文京区'等)。未報告の医療機関"
        "(facility_observations.csvのvalue_status='not_reported'参照)は"
        "原典側で所在地欄も空欄のため空になる"
    ),
}

FIELDS_OBSERVATIONS = {
    "published_fy": PUBLISHED_FY_DESC,
    "pref_code": PREF_CODE_DESC,
    "pref_name": PREF_NAME_DESC,
    "area_code": AREA_CODE_DESC,
    "area_name": AREA_NAME_DESC,
    "record_id": "facility_basic.csvのrecord_idと対応する外部キー(制約はfacility_basic.csvのfields.record_id参照)",
    "metric": (
        "指標名(原典の見出し語をそのまま使用)。'病床数'/'医師数（常勤）'/"
        "'医師数（非常勤）'/'医師数（100床当たり）'/'救急車の受入件数'/"
        "'全身麻酔手術件数'/'分娩件数'/'手術総数'/'平均在棟日数'/'新規入棟患者'"
    ),
    "bed_function": (
        "病床機能区分。metric='病床数'では'休棟中等含む計'/'高度急性期'/"
        "'急性期'/'回復期'/'慢性期'/'休棟中等'の6区分、metric='平均在棟日数'/"
        "'新規入棟患者'では'高度急性期'/'急性期'/'回復期'/'慢性期'の4区分。"
        "それ以外の指標では空文字"
    ),
    "value": (
        "指標の値(原典セルの値をそのまま。丸め・整形はしていない)。"
        "value_status='observed'(数値)のときのみ入る。それ以外は空"
        "(欠測ではない。原因はvalue_status/source_literalを参照)"
    ),
    "value_status": (
        "セルの値の分類。'observed'=数値、'source_dash'=原典が'-'、"
        "'blank'=原典が空セル、'not_calculated'=原典が'XXX'(未算出。他帳票"
        "001723349の推計流出入患者割合の前例に備えた分岐で、本ファイルの"
        "実測データでは未検出)、'not_disclosed'=原典が'*'(NDBの利用に関する"
        "ガイドラインにより一部非公表。原典2行目の注記3に対応)、"
        "'not_reported'=原典が'未報告'(病床機能報告を未報告の医療機関。"
        "原典2行目の注記4に対応)。これ以外の非数値が出た場合はレイアウト"
        "変更とみなしパーサが中断する(静かに欠測へ落とさない)"
    ),
    "source_literal": "value_status='observed'以外のときの原典セルの文字列そのまま(観測値の場合は空)",
}

FIELDS_FUNCTIONS = {
    "published_fy": PUBLISHED_FY_DESC,
    "pref_code": PREF_CODE_DESC,
    "pref_name": PREF_NAME_DESC,
    "area_code": AREA_CODE_DESC,
    "area_name": AREA_NAME_DESC,
    "record_id": "facility_basic.csvのrecord_idと対応する外部キー(制約はfacility_basic.csvのfields.record_id参照)",
    "function_name": (
        "医療機関機能区分(ヘッダー文字列から解決した名前)。'特定機能'/"
        "'地域支援'/'三次救急'/'二次救急'/'在宅療養支援'"
    ),
    "source_literal": "原典セルの略記文字列そのまま(例 '特'/'地'/'三次'/'二次'/'在支')。参考情報",
}

CAVEAT_BASIC = (
    "医療機関の個票一覧(識別情報のみ)。数値指標はfacility_observations.csv、"
    "医療機関機能はfacility_functions.csvを参照。record_idは恒久的な施設IDでは"
    "ない点に特に留意(fields.record_id参照)。"
)
CAVEAT_OBSERVATIONS = (
    "病床数・医師数・診療実績等は原典(2025年度)時点の病床機能報告等に基づく。"
    "診療実績4指標(救急車の受入件数・全身麻酔手術件数・分娩件数・手術総数)は"
    "NDBの利用に関するガイドラインにより一部非公表('*'、value_status="
    "'not_disclosed')となっている場合がある。病床機能報告そのものを行って"
    "いない医療機関('未報告'、value_status='not_reported')も一覧に含まれるが、"
    "その場合はほぼ全ての数値指標が原典側で空欄になる(facility_basic.csvの"
    "所在地も同様に空)。"
)
CAVEAT_FUNCTIONS = (
    "該当が確認できた行のみを出力する(空欄の行は出力しない)。原典の空欄が"
    "「非該当」か「未回答」かを機械的に判別できないため、0/falseに解釈して"
    "出力することはしない。特定機能病院・地域医療支援病院・三次救急は厚生"
    "労働省医政局地域医療計画課調べ、二次救急・在宅療養支援病院・診療所は"
    "2025年度の病床機能報告による(原典の出典説明書001723348.pdfより)。"
)

KNOWN_ISSUES = [
    {
        "id": "facility_basic_summary_hospital_count_mismatch",
        "scope": {
            "csv": "facility_basic.csv",
            "note": (
                "区域サマリ(7行目、原典側の集計値)と個票行数(facility_basic.csvの"
                "行数)の不一致。facility_basic.csv自体の値は個票をそのまま出力した"
                "ものであり、この既知事項によって値が変わるわけではない"
            ),
        },
        "summary": (
            "339シート中78シートで、区域サマリの医療機関数(F列「一般病院」+G列"
            "「有床診療所」)と、その区域の医療機関個票の行数が一致しない"
        ),
        "evidence": [
            "78シート中76シートは、差分が「未報告」医療機関"
            "(facility_observations.csvのmetric='病床数', bed_function='休棟中等含む計', "
            "value_status='not_reported')の件数と完全に一致する"
            "(例: 104北海道札幌は期待値324・実際333・差9・未報告9件で一致)。"
            "未報告の医療機関は個票一覧には載るが、区域サマリのF+G集計には含まれて"
            "いないとみられる",
            "残る2シート(1503新潟県県央: 期待値11・実際12、2007長野県松本: 期待値33・"
            "実際34)のみ、+1件の差異が未報告医療機関数(いずれも0件)では説明できず、"
            "原因不明",
        ],
        "action": (
            "値は原典どおり出力している(勝手に補正しない)。この2シート"
            "(1503新潟県県央・2007長野県松本)の医療機関数は、区域サマリとの整合が"
            "取れないため集計(例:区域別の医療機関数の表示)に使わないこと"
        ),
    },
]

STEPS_COMMON = [
    "verify_source()でR7/001723127.xlsxのSHA-256をSHA256SUMSと照合",
    "openpyxl(data_only=True)でワークブックを開き、シート数がちょうど339であることを検証",
    "シート名を「先頭の数値(区域コード)+残り」に分解し、区域コードをゼロ埋め4桁へ正規化した上で"
    "data/processed/area_basic.csvと突合(都道府県名+区域名の連結がシート名の残り部分と一致することを検証)",
    "11〜13行目の医療機関表ヘッダーを結合セル対応(前方補完)で解決し、全339シートが先頭シートと"
    "(1行目のタイトル・2,3,10,11〜13行目について、1行目最終列の区域名注記を除き)完全一致することを"
    "検証。列の識別は3行分の文字列の三つ組の"
    "完全一致で行うため、「急性期」が「高度急性期」に部分一致して誤った列を拾うことはない",
    "各シートの医療機関行(14行目〜)をA列の連番(1始まり・欠番/重複なし)が途切れるまで走査し、"
    "それ以降にA列へ整数値が再出現しないことを確認(書式だけの空行と本物の行を取り違えないため)",
    "数値以外のセルは'-'(source_dash)/空欄(blank)/'*'(not_disclosed)/'未報告'(not_reported)/"
    "'XXX'(not_calculated)のいずれかであることを確認し、それ以外はレイアウト変更とみなして中断",
    "facility_observations.csv・facility_functions.csvのrecord_idが全てfacility_basic.csvに"
    "存在することを検証",
]


def _expected_columns():
    """(11行目, 12行目, 13行目)の文字列の三つ組と、列の意味の対応表を作る。

    列は位置ではなくこの三つ組で解決する(CLAUDE.mdの規律)。①医療機関名は
    B〜G列にまたがる結合セルのため三つ組が6列連続で同一になるが、
    `resolve_facility_columns()` 側で「同一三つ組が連続するrun」として扱う
    ことで単一列(先頭=B列)に正しく解決される(他の項目はもともと1列だけの
    runになるため、同じ仕組みで自然に解決される)。
    """
    items = []
    items.append(((LEVEL1_NAME, LEVEL1_NAME, LEVEL1_NAME), ("name",)))
    items.append(((LEVEL1_MUNICIPALITY, LEVEL1_MUNICIPALITY, LEVEL1_MUNICIPALITY), ("municipality",)))
    for bf in BED_FUNCTIONS_FACILITY:
        items.append(((LEVEL1_BED, bf, bf), ("bed", bf)))
    for label in DOCTOR_LABELS:
        items.append(((LEVEL1_DOCTOR, label, label), ("doctor", DOCTOR_METRIC_NAMES[label])))
    for fn in FUNCTION_NAMES:
        items.append(((LEVEL1_FUNCTION, fn, fn), ("function", fn)))
    for metric in CLINICAL_SIMPLE_METRICS:
        items.append(((LEVEL1_CLINICAL, metric, metric), ("clinical_simple", metric)))
    for metric in CLINICAL_GROUPED_METRICS:
        for bf in CLINICAL_GROUPED_BED_FUNCTIONS:
            items.append(((LEVEL1_CLINICAL, metric, bf), ("clinical_grouped", metric, bf)))
    triples = [t for t, _ in items]
    assert len(triples) == len(set(triples)), "EXPECTED_COLUMNSの三つ組定義に重複があります"
    return dict(items)


EXPECTED_COLUMNS = _expected_columns()


def _filled_header_grid(ws, rows, col_start, col_end):
    """`rows`(複数行)×col_start〜col_endの範囲で、結合セルを前方補完した
    正規化テキストの辞書({行番号: [列ごとの文字列, ...]})を返す。

    結合セルは左上(アンカー)以外のセルの値がNoneになる(read_onlyかどうかに
    関わらず)ため、`ws.merged_cells.ranges` を使ってアンカーの値を範囲全体へ
    複写してから正規化する。
    """
    grid = {
        r: [ws.cell(row=r, column=c).value for c in range(col_start, col_end + 1)]
        for r in rows
    }
    for rng in ws.merged_cells.ranges:
        if rng.max_row < min(rows) or rng.min_row > max(rows):
            continue
        anchor_value = ws.cell(row=rng.min_row, column=rng.min_col).value
        for r in rows:
            if not (rng.min_row <= r <= rng.max_row):
                continue
            lo = max(rng.min_col, col_start)
            hi = min(rng.max_col, col_end)
            for c in range(lo, hi + 1):
                grid[r][c - col_start] = anchor_value
    return {r: [normalize_header_text(v) for v in vals] for r, vals in grid.items()}


def resolve_facility_columns(ws, *, sheet_name, col_start=2, col_end=MAX_COL):
    """医療機関表ヘッダー(11〜13行目)を解決し、{記述子: 列番号} を返す。

    記述子は `("name",)` / `("municipality",)` / `("bed", 病床機能)` /
    `("doctor", 指標名)` / `("function", 機能名)` /
    `("clinical_simple", 指標名)` / `("clinical_grouped", 指標名, 病床機能)`
    のいずれか。期待する項目がそれぞれちょうど1回(1つのrun)だけ解決される
    ことを検証し、未知の見出しや解決漏れがあれば `LayoutMismatchError` で
    中断する。
    """
    grid = _filled_header_grid(ws, HEADER_ROWS, col_start, col_end)
    n = col_end - col_start + 1
    triples = [tuple(grid[r][idx] for r in HEADER_ROWS) for idx in range(n)]

    runs = []
    run_start_idx = 0
    for idx in range(1, n + 1):
        if idx == n or triples[idx] != triples[idx - 1]:
            runs.append((triples[idx - 1], col_start + run_start_idx, col_start + idx - 1))
            run_start_idx = idx

    run_triples = [t for t, _, _ in runs]
    if len(run_triples) != len(set(run_triples)):
        raise LayoutMismatchError(
            f"{sheet_name}: ヘッダーの同一項目が離れた位置に複数回出現しています: {runs}"
        )

    resolved = {}
    for triple, start, end in runs:
        if triple not in EXPECTED_COLUMNS:
            raise LayoutMismatchError(
                f"{sheet_name}: 未知のヘッダーです(列{start}〜{end}): {triple}"
            )
        resolved[EXPECTED_COLUMNS[triple]] = start

    missing = set(EXPECTED_COLUMNS.values()) - set(resolved.keys())
    if missing:
        raise LayoutMismatchError(f"{sheet_name}: 解決できなかった項目があります: {missing}")

    return resolved


def _raw_header_signature(ws, max_col=MAX_COL):
    """1,2,3,10,11〜13行目の生セル値をシート間比較用に読む。

    結合セルの前方補完はしない(非アンカーセルのNoneも含めてそのまま比較する
    ことで、結合構造自体がシート間で異なっていても検知できる)。

    ⚠ 1行目(タイトル行)の最終列(AH列)だけは例外: 「（北海道・南渡島）」の
    ように「（都道府県名・区域名）」が区域ごとに埋め込まれており、意図的に
    シートごとに異なる(実測で確認済み。それ以外の1行目の列は全シートで
    Noneかタイトル文字列(A1列)のみ)。この列だけ比較対象から除外する。
    """
    rows = (NOTES_ROW, SECTION1_ROW, SECTION2_ROW) + HEADER_ROWS
    title_row_signature = tuple(
        ws.cell(row=TITLE_ROW, column=c).value for c in range(1, max_col)
    )
    return (title_row_signature,) + tuple(
        tuple(ws.cell(row=r, column=c).value for c in range(1, max_col + 1)) for r in rows
    )


def _parse_sheet_name(name: str):
    """シート名を (区域コード(ゼロ埋め4桁文字列), 残り文字列) に分解する。"""
    m = SHEET_NAME_RE.match(name)
    if not m:
        raise LayoutMismatchError(f"シート名から区域コードを抽出できません: {name!r}")
    area_code = normalize_area_code(int(m.group(1)))
    return area_code, m.group(2)


def _is_int_like(value) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return True
    if isinstance(value, float):
        return value.is_integer()
    return False


def _classify_metric_cell(raw, *, context: str):
    """医療機関表の数値指標セルを (value_status, value, source_literal) に分類する。

    数値ならそのまま`observed`として値を採用する。'-'・'*'・'未報告'・'XXX'は
    既知の非数値センチネルとして分類し、値は空・原文をsource_literalへ保持
    する。それ以外の非数値はレイアウト変更とみなし `LayoutMismatchError` で
    中断する(静かに欠測へ落とさない)。
    """
    if raw is None:
        return VALUE_STATUS_BLANK, None, None
    if isinstance(raw, bool):
        raise LayoutMismatchError(f"{context}: セルの値がbool型です: {raw!r}")
    if isinstance(raw, (int, float)):
        return VALUE_STATUS_OBSERVED, raw, None
    if raw == "-":
        return VALUE_STATUS_SOURCE_DASH, None, raw
    if raw == "*":
        return VALUE_STATUS_NOT_DISCLOSED, None, raw
    if raw == "未報告":
        return VALUE_STATUS_NOT_REPORTED, None, raw
    if raw == "XXX":
        return VALUE_STATUS_NOT_CALCULATED, None, raw
    raise LayoutMismatchError(f"{context}: 未知の非数値センチネルです(レイアウト変更の可能性): {raw!r}")


@dataclass
class SheetParseResult:
    basic_rows: list = field(default_factory=list)
    observation_rows: list = field(default_factory=list)
    function_rows: list = field(default_factory=list)
    facility_count: int = 0


def parse_sheet(ws, *, sheet_name, area_code, pref_code, pref_name, area_name, columns) -> SheetParseResult:
    """1シート分(1構想区域)の医療機関表(14行目〜)をtidy行へ変換する。

    `columns` は `resolve_facility_columns()` の戻り値({記述子: 列番号})。
    """
    result = SheetParseResult()
    col_name = columns[("name",)]
    col_municipality = columns[("municipality",)]

    row = DATA_START_ROW
    expected_seq = 1
    while _is_int_like(ws.cell(row=row, column=1).value):
        seq_val = ws.cell(row=row, column=1).value
        seq = int(seq_val)
        expect(seq, expected_seq, f"{sheet_name} 行{row}: 医療機関の連番(A列)")

        name = ws.cell(row=row, column=col_name).value
        if not isinstance(name, str) or not name.strip():
            raise LayoutMismatchError(f"{sheet_name} 行{row}: 医療機関名が空です: {name!r}")

        municipality = ws.cell(row=row, column=col_municipality).value
        if municipality is not None and not isinstance(municipality, str):
            raise LayoutMismatchError(
                f"{sheet_name} 行{row}: 所在地が文字列ではありません: {municipality!r}"
            )

        record_id = f"{PUBLISHED_FY}-{area_code}-{row}"
        result.basic_rows.append(
            {
                "published_fy": PUBLISHED_FY,
                "pref_code": pref_code,
                "pref_name": pref_name,
                "area_code": area_code,
                "area_name": area_name,
                "record_id": record_id,
                "source_sheet": sheet_name,
                "source_row": row,
                "facility_seq": seq,
                "facility_name": name,
                "municipality": municipality,
            }
        )

        for key, col in columns.items():
            kind = key[0]
            if kind in ("name", "municipality"):
                continue
            context = f"{sheet_name} 行{row} 列{col}"
            raw = ws.cell(row=row, column=col).value

            if kind == "function":
                if raw is None:
                    continue
                if not isinstance(raw, str) or not raw.strip():
                    raise LayoutMismatchError(f"{context}: 医療機関機能セルの値が不正です: {raw!r}")
                result.function_rows.append(
                    {
                        "published_fy": PUBLISHED_FY,
                        "pref_code": pref_code,
                        "pref_name": pref_name,
                        "area_code": area_code,
                        "area_name": area_name,
                        "record_id": record_id,
                        "function_name": key[1],
                        "source_literal": raw,
                    }
                )
                continue

            if kind == "bed":
                metric, bed_function = METRIC_BED, key[1]
            elif kind == "doctor":
                metric, bed_function = key[1], ""
            elif kind == "clinical_simple":
                metric, bed_function = key[1], ""
            elif kind == "clinical_grouped":
                metric, bed_function = key[1], key[2]
            else:  # pragma: no cover - EXPECTED_COLUMNSの定義から到達しないはず
                raise AssertionError(f"未知の記述子kindです: {kind}")

            value_status, value, source_literal = _classify_metric_cell(raw, context=context)
            result.observation_rows.append(
                {
                    "published_fy": PUBLISHED_FY,
                    "pref_code": pref_code,
                    "pref_name": pref_name,
                    "area_code": area_code,
                    "area_name": area_name,
                    "record_id": record_id,
                    "metric": metric,
                    "bed_function": bed_function,
                    "value": value,
                    "value_status": value_status,
                    "source_literal": source_literal,
                }
            )

        expected_seq += 1
        row += 1

    result.facility_count = expected_seq - 1

    # 終端検出: それ以降にA列へ整数値が再出現したら中断する
    # (書式だけの空行と本物の行を取り違えないため)。
    for check_row in range(row, ws.max_row + 6):
        if _is_int_like(ws.cell(row=check_row, column=1).value):
            stray = ws.cell(row=check_row, column=1).value
            raise LayoutMismatchError(
                f"{sheet_name} 行{check_row}: 医療機関表終了後にA列へ整数値が再出現しました: {stray!r}"
            )

    return result


def _summary_soft_check(ws, *, facility_count: int):
    """区域サマリ(7行目)の医療機関数(F+G)と実際の医療機関行数を比較する
    (非fatal。docstring「区域サマリとの非fatal突合」参照)。

    戻り値: (期待値, 実際の行数, 一致したか) または、サマリ側が数値でない
    (レイアウト前提が崩れている)場合は `None`。
    """
    hospital = ws.cell(row=SUMMARY_ROW, column=SUMMARY_COL_HOSPITAL).value
    clinic = ws.cell(row=SUMMARY_ROW, column=SUMMARY_COL_CLINIC).value
    if not isinstance(hospital, (int, float)) or not isinstance(clinic, (int, float)):
        return None
    if isinstance(hospital, bool) or isinstance(clinic, bool):
        return None
    expected = int(hospital) + int(clinic)
    return expected, facility_count, expected == facility_count


def load_workbook():
    """R7/001723127.xlsx を開く(SHA-256照合を済ませたうえで)。

    戻り値: (workbook, source_sha256)
    """
    source_sha256 = verify_source(SOURCE_PATH_IN_REPO)
    print(f"[ok] 生データ検証: {SOURCE_PATH_IN_REPO} = {source_sha256[:16]}...")
    xlsx_path = REPO_ROOT / SOURCE_PATH_IN_REPO
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    return wb, source_sha256


def _load_area_basic_reference():
    """`data/processed/area_basic.csv` から area_code -> (pref_name, area_name)
    の参照テーブルを読み込む(シート名の分解が正しいことの検証用)。

    001723127.xlsx（医療機関個票）はまだR7のみに対応するパーサ(「データ構成」
    参照)なので、参照先も published_fy=='R7' の行に絞り込む。area_basic.csvは
    R6/R7がpublished_fyで並存するようになった(M9)ため678行あり、絞り込まずに
    読むと同じarea_codeがR7行・R6行の順で2回出現し、辞書代入で後勝ち(=R6の値)
    になってしまう(値そのものはR6/R7で同じはずだが、この帳票の突合が本来意図
    しないR6行に依存する状態は避ける)。
    """
    path = REPO_ROOT / "data" / "processed" / "area_basic.csv"
    reference = {}
    with open(path, encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            if row["published_fy"] != "R7":
                continue
            reference[row["area_code"]] = (row["pref_name"], row["area_name"])
    return reference


@dataclass
class WorkbookParseResult:
    title: str
    notes: list
    basic_rows: list
    observation_rows: list
    function_rows: list
    summary_matches: int
    summary_mismatches: list
    summary_unavailable: int


def parse_workbook(wb) -> WorkbookParseResult:
    """ワークブック全体(339シート)をパースし、3種のtidy行と区域サマリとの
    非fatal突合の集計結果を返す。
    """
    sheet_names = wb.sheetnames
    expect(len(sheet_names), NUM_SHEETS, "シート数")

    reference_ws = wb[sheet_names[0]]
    reference_signature = _raw_header_signature(reference_ws)
    title = reference_ws.cell(row=TITLE_ROW, column=1).value
    notes_raw = reference_ws.cell(row=NOTES_ROW, column=1).value
    notes = notes_raw.split("\n") if notes_raw else []
    columns = resolve_facility_columns(reference_ws, sheet_name=sheet_names[0])

    area_basic_ref = _load_area_basic_reference()
    seen_area_codes = []

    basic_rows, observation_rows, function_rows = [], [], []
    summary_matches = 0
    summary_mismatches = []
    summary_unavailable = 0

    for sheet_name in sheet_names:
        ws = wb[sheet_name]

        signature = _raw_header_signature(ws)
        expect(signature, reference_signature, f"{sheet_name}: ヘッダー(1行目のタイトル・2,3,10,11〜13行目)が先頭シートと不一致")

        area_code, rest = _parse_sheet_name(sheet_name)
        if area_code in seen_area_codes:
            raise LayoutMismatchError(f"{sheet_name}: 構想区域コード{area_code}が重複しています")
        seen_area_codes.append(area_code)

        if area_code not in area_basic_ref:
            raise LayoutMismatchError(f"{sheet_name}: 構想区域コード{area_code}がarea_basic.csvに存在しません")
        pref_name, area_name = area_basic_ref[area_code]
        expect(rest, pref_name + area_name, f"{sheet_name}: シート名の残り部分がarea_basic.csvと不一致")
        pref_code = normalize_pref_code(int(area_code[:2]))

        sheet_result = parse_sheet(
            ws,
            sheet_name=sheet_name,
            area_code=area_code,
            pref_code=pref_code,
            pref_name=pref_name,
            area_name=area_name,
            columns=columns,
        )
        basic_rows.extend(sheet_result.basic_rows)
        observation_rows.extend(sheet_result.observation_rows)
        function_rows.extend(sheet_result.function_rows)

        soft = _summary_soft_check(ws, facility_count=sheet_result.facility_count)
        if soft is None:
            summary_unavailable += 1
        else:
            expected, actual, ok = soft
            if ok:
                summary_matches += 1
            else:
                summary_mismatches.append((sheet_name, expected, actual))

    expect(set(seen_area_codes), set(area_basic_ref.keys()), "構想区域コードの集合がarea_basic.csvと不一致")
    expect(seen_area_codes, sorted(seen_area_codes), "シートの並び順(区域コード昇順であること)")

    # facility_observations.csv・facility_functions.csv の record_id が
    # 全て facility_basic.csv に存在することを検証する(要件#7)。
    basic_ids = {r["record_id"] for r in basic_rows}
    for r in observation_rows:
        if r["record_id"] not in basic_ids:
            raise LayoutMismatchError(f"record_id {r['record_id']!r} がfacility_basicに存在しません(observations)")
    for r in function_rows:
        if r["record_id"] not in basic_ids:
            raise LayoutMismatchError(f"record_id {r['record_id']!r} がfacility_basicに存在しません(functions)")

    return WorkbookParseResult(
        title=title,
        notes=notes,
        basic_rows=basic_rows,
        observation_rows=observation_rows,
        function_rows=function_rows,
        summary_matches=summary_matches,
        summary_mismatches=summary_mismatches,
        summary_unavailable=summary_unavailable,
    )


def _rows_to_tuples(rows, header):
    return [tuple(r[h] for h in header) for r in rows]


def build_and_write(out_dir: Path) -> dict:
    """R7/001723127.xlsxをパースし、3つのCSV+meta.jsonを `out_dir` へ出力する。

    書き出したCSVパスの辞書({"basic": ..., "observations": ..., "functions": ...})
    を返す(再現性テスト等での再利用を想定)。
    """
    out_dir = Path(out_dir)
    wb, source_sha256 = load_workbook()
    result = parse_workbook(wb)
    print(
        f"[ok] パース完了: facility_basic={len(result.basic_rows)}行 "
        f"facility_observations={len(result.observation_rows)}行 "
        f"facility_functions={len(result.function_rows)}行"
    )
    print(
        f"[info] 区域サマリの医療機関数(F+G)と個票行数の一致(非fatal参考値): "
        f"{result.summary_matches}/{NUM_SHEETS}シート "
        f"(不一致{len(result.summary_mismatches)}件, 判定不能{result.summary_unavailable}件)"
    )

    today = datetime.date.today().isoformat()
    base_source = {
        "name": SOURCE_NAME,
        "publisher": "厚生労働省",
        "url": SOURCE_DOWNLOAD_URL,
        "page_url": SOURCE_PAGE_URL,
        "fiscal_year": SOURCE_FISCAL_YEAR,
        "source_file": SOURCE_PATH_IN_REPO,
        "source_sha256": source_sha256,
        "source_sheet": (
            f"{NUM_SHEETS}シート(構想区域ごとに1シート。シート名は「101北海道南渡島」のように"
            "区域コード+都道府県名+区域名)"
        ),
        "acquired_date": SOURCE_ACQUIRED_DATE,
        "license": LICENSE_NOTE,
        "original_title": result.title,
        "original_notes": result.notes,
    }

    basic_header = [
        "published_fy",
        "pref_code",
        "pref_name",
        "area_code",
        "area_name",
        "record_id",
        "source_sheet",
        "source_row",
        "facility_seq",
        "facility_name",
        "municipality",
    ]
    basic_csv, _ = write_csv_with_meta(
        out_dir / "facility_basic.csv",
        basic_header,
        _rows_to_tuples(result.basic_rows, basic_header),
        title="構想区域別 医療機関一覧(識別情報)",
        source=base_source,
        processing={
            "script": "tools/parse_facility_beds.py",
            "date": today,
            "steps": STEPS_COMMON,
            "caveat": CAVEAT_BASIC,
        },
        fields=FIELDS_BASIC,
        known_issues=KNOWN_ISSUES,
    )
    print(f"[ok] 出力: {basic_csv} ({len(result.basic_rows)}行)")

    observations_header = [
        "published_fy",
        "pref_code",
        "pref_name",
        "area_code",
        "area_name",
        "record_id",
        "metric",
        "bed_function",
        "value",
        "value_status",
        "source_literal",
    ]
    observations_csv, _ = write_csv_with_meta(
        out_dir / "facility_observations.csv",
        observations_header,
        _rows_to_tuples(result.observation_rows, observations_header),
        title="構想区域別 医療機関の病床数・医師数・診療実績(数値指標)",
        source=base_source,
        processing={
            "script": "tools/parse_facility_beds.py",
            "date": today,
            "steps": STEPS_COMMON,
            "caveat": CAVEAT_OBSERVATIONS,
        },
        fields=FIELDS_OBSERVATIONS,
    )
    print(f"[ok] 出力: {observations_csv} ({len(result.observation_rows)}行)")

    functions_header = [
        "published_fy",
        "pref_code",
        "pref_name",
        "area_code",
        "area_name",
        "record_id",
        "function_name",
        "source_literal",
    ]
    functions_csv, _ = write_csv_with_meta(
        out_dir / "facility_functions.csv",
        functions_header,
        _rows_to_tuples(result.function_rows, functions_header),
        title="構想区域別 医療機関機能(特定機能・地域支援・三次救急・二次救急・在宅療養支援)",
        source=base_source,
        processing={
            "script": "tools/parse_facility_beds.py",
            "date": today,
            "steps": STEPS_COMMON,
            "caveat": CAVEAT_FUNCTIONS,
        },
        fields=FIELDS_FUNCTIONS,
    )
    print(f"[ok] 出力: {functions_csv} ({len(result.function_rows)}行)")

    return {"basic": basic_csv, "observations": observations_csv, "functions": functions_csv}


def main():
    out_dir = REPO_ROOT / "data" / "processed"
    build_and_write(out_dir)


if __name__ == "__main__":
    main()
