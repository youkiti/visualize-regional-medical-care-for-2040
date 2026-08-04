# -*- coding: utf-8 -*-
"""厚労省「①都道府県の病床数等」(R7: 001722915.xlsx / R6: 別添４②)の帳票Excel
を tidy CSV へ変換する。

処理内容:
  1. `verify_source()` で元データのSHA-256を `SHA256SUMS` と照合(改変検知)
  2. openpyxl(`data_only=True`)でシートを開き、3行目から15行ずつ・48ブロック
     (先頭=全国、続く47ブロック=都道府県コード1〜47順)を走査
  3. サブヘッダー行(各ブロック先頭+8行目)の文字列("2015実績"等)から
     実績/見込量/必要数の列を解決する。列位置は公表年度により異なる
     (下記「R6との列ずれ」参照)ためハードコードしない
  4. 全48ブロックのサブヘッダー行が先頭ブロックと完全一致することを検証し、
     不一致ならレイアウト変更とみなして例外で中断する(取りこぼし防止)
  5. 3つの tidy CSV + 由来メタデータ(`<csv名>.meta.json`)を `data/processed/`
     に出力する
       - prefecture_beds.csv: 病床数(実績/見込量/必要数 × 5機能 × 年)
       - prefecture_bed_report_rate.csv: 病床機能報告の報告率
       - prefecture_basic.csv: 基礎情報(2020人口・面積)

⚠ R6との列ずれ: R6(別添４②)は実績年が1年少なく(2015, 2018〜2024)、その
分だけ見込量/必要数の列がR7よりも1列前にずれる。見込量の対象年も異なる
(R6=2025年見込量 / R7=2026年見込量)。そのため列は位置ではなく、サブヘッダー
行の文字列から都度解決する(`_read_header_row` / `_column_map_from_header`)。

必要環境: Python 3.11+, openpyxl

使い方:
    python tools/parse_prefecture_beds.py [--source R7]
"""
import argparse
import datetime
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

if __package__ in (None, ""):
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

from tools.lib.codes import normalize_pref_code
from tools.lib.provenance import REPO_ROOT, verify_source, write_csv_with_meta

# 各公表年度の元データ設定。将来 R6 の出力にも対応できるよう両方定義して
# おくが、CLI(--source)では現時点で R7 のみ受け付ける(R6 出力は未対応、
# パース自体はテストから `load_sheet("R6")` で直接利用できる)。
SOURCES = {
    "R7": {
        # R7のシートX1セルの表記(「別添４」)に合わせる。R6の「別添４②」を
        # R7の由来情報へ書き込むと出典が不正確になるため、年度ごとに個別に持つ。
        "name": "①都道府県の病床数等（別添４）",
        "path_in_repo": "R7/001722915.xlsx",
        "sheet_name": "都道府県別必要量との比較",
        "download_url": "https://www.mhlw.go.jp/content/10800000/001722915.xlsx",
        "fiscal_year": "令和7年度（2025年度）",
        "acquired_date": "2026-08-04",
    },
    "R6": {
        "name": "別添４②（都道府県の病床数等の状況）",
        "path_in_repo": "R6/別添４②（都道府県の病床数等の状況）.xlsx",
        "sheet_name": "都道府県別必要量との比較",
        # R6は令和6年度版一括DL(001723128.zip)からの取得と推定(doc/DATA_SOURCES.md参照)。
        # R6出力は今回未対応のためmeta.jsonには使わない。
        "download_url": None,
        "fiscal_year": "令和6年度",
        "acquired_date": None,
    },
}

SOURCE_PAGE_URL = "https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/0000080850_00014.html"
LICENSE_NOTE = (
    "厚生労働省ホームページ利用規約（政府標準利用規約準拠） "
    "https://www.mhlw.go.jp/chosakuken/index.html"
)
CAVEAT = (
    "病床機能報告の集計結果(実績)と将来の病床数の必要量(必要数)は計算方法が"
    "異なることから、単純に比較するのではなく、詳細な分析や検討を行った上で"
    "地域医療構想調整会議で協議を行うことが重要(原典C2注記より)。可視化で"
    "これらを併記する際は、この点を注記として必ず表示すること。また年度ごと"
    "に報告率が異なることにも留意する。"
)

BLOCK_TOP0 = 3   # 最初のブロック(全国, block=0)の先頭行
BLOCK_SIZE = 15  # 1ブロックあたりの行数
NUM_BLOCKS = 48  # 全国(0) + 都道府県(1〜47)

BED_FUNCTIONS = ["合計", "高度急性期", "急性期", "回復期", "慢性期"]

_ACTUAL_RE = re.compile(r"^(\d{4})実績$")
_PLAN_RE = re.compile(r"^(\d{4})見込量$")
_REQUIRED_RE = re.compile(r"^(\d{4})必要数$")

PREF_CODE_DESC = (
    "都道府県コード(ゼロ埋め2桁の文字列)。'00'=全国、'01'〜'47'=都道府県"
    "(01=北海道…47=沖縄県、原典の都道府県コード順)"
)
PREF_NAME_DESC = "都道府県名(全国の行は'全国')"
PUBLISHED_FY_DESC = (
    "公表年度を表す識別子。'R7'=令和7年度公表分。将来R6等の行を追加する際のキー"
)

FIELDS_BEDS = {
    "published_fy": PUBLISHED_FY_DESC,
    "pref_code": PREF_CODE_DESC,
    "pref_name": PREF_NAME_DESC,
    "bed_function": "病床機能区分(合計/高度急性期/急性期/回復期/慢性期)。合計は他4区分の和",
    "series": "系列。実績=病床機能報告の報告値、見込量=直近年からの見込み、必要数=2025年の必要病床数",
    "year": "対象年(西暦)。実績・見込量・必要数それぞれの対象年は公表年度により異なる(下記caveat参照)",
    "beds": "病床数(床)",
}
FIELDS_REPORT_RATE = {
    "published_fy": PUBLISHED_FY_DESC,
    "pref_code": PREF_CODE_DESC,
    "pref_name": PREF_NAME_DESC,
    "year": "報告率の対象年(実績年のみ)",
    "report_rate": "病床機能報告の報告率(原典値をそのまま。丸めていない0〜1の割合)",
}
FIELDS_BASIC = {
    "published_fy": PUBLISHED_FY_DESC,
    "pref_code": PREF_CODE_DESC,
    "pref_name": PREF_NAME_DESC,
    "population_2020": "2020年国勢調査人口(人単位の整数)。原典の万人値を10000倍し四捨五入",
    "population_2020_source_value": "原典の人口値(万人単位、丸めなし)",
    "population_2020_source_unit": "population_2020_source_value の単位(万人)",
    "area_2020_km2": "2020年面積(km2)。原典の浮動小数点誤差を除くため小数2桁に丸め",
}

STEPS_COMMON = [
    "openpyxl(data_only=True)でシートを開き、3行目から15行ずつの48ブロック(全国+47都道府県)を走査",
    "サブヘッダー行(実績/見込量/必要数の列見出し)の文字列から列を解決(公表年度により列位置が異なるためハードコードしない)",
    "全48ブロックのサブヘッダー行が先頭ブロックと完全一致することを検証(不一致ならレイアウト変更とみなし中断)",
]


class LayoutMismatchError(Exception):
    """帳票のレイアウトが想定(サブヘッダー・ラベル位置)と異なる場合に送出する。"""


@dataclass
class ParseResult:
    published_fy: str
    title: str
    notes: list
    beds_rows: list = field(default_factory=list)
    report_rate_rows: list = field(default_factory=list)
    basic_rows: list = field(default_factory=list)


def _normalize_header_text(v) -> str:
    if v is None:
        return ""
    return str(v).replace("\n", "").strip()


def _read_header_row(ws, header_row: int, col_start: int, col_end: int):
    """サブヘッダー行を正規化した文字列のタプルとして読む(ブロック間比較用)。

    `col_start`〜`col_end`(両端含む)の範囲のみを対象とする。A列(ブロック
    番号)と最終列(「0　全国」等のブロックごとの通し番号ラベル)はブロックに
    よって値が変わるのが仕様上正しい(レイアウト崩れではない)ため、比較対象
    から除外する。
    """
    return tuple(
        _normalize_header_text(ws.cell(row=header_row, column=c).value)
        for c in range(col_start, col_end + 1)
    )


def _column_map_from_header(raw_header, col_start: int):
    """正規化済みサブヘッダーから {列番号: (series, year)} を作る。

    実績/見込量/必要数の列のみを対象とする。派生比率列
    (…に対する比／…との差／見込み／必要数)はこの関数の正規表現に
    マッチしないため自然に除外される(値はパーサ出力から再計算可能)。
    """
    col_map = {}
    for idx, text in enumerate(raw_header, start=col_start):
        if not text:
            continue
        m = _ACTUAL_RE.match(text)
        if m:
            col_map[idx] = ("実績", int(m.group(1)))
            continue
        m = _PLAN_RE.match(text)
        if m:
            col_map[idx] = ("見込量", int(m.group(1)))
            continue
        m = _REQUIRED_RE.match(text)
        if m:
            col_map[idx] = ("必要数", int(m.group(1)))
            continue
    return col_map


def _expect(actual, expected, message):
    if actual != expected:
        raise LayoutMismatchError(f"{message}: 期待={expected!r} 実際={actual!r}")


def _expect_int(value, *, block, row, col):
    """病床数セルの値が整数(またはinteger値のfloat)であることを検証する。

    素朴な `int(value)` は非整数の float を黙って切り捨て(`1234.7` ->
    `1234`)、`None` なら意味の分からない `TypeError` になる。真正性を
    担保するパイプラインで「静かに値が変わる」のは最悪の失敗モードのため、
    int、または `float.is_integer()` が真の float 以外はブロック番号・行・
    列・実際の値を添えて `LayoutMismatchError` で中断する。
    """
    if isinstance(value, int) and not isinstance(value, bool):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    raise LayoutMismatchError(
        f"ブロック{block} 行{row} 列{col}: 病床数セルの値が整数ではありません: {value!r}"
    )


def _convert_man_to_person(value_man) -> int:
    """人口(万人)を人単位の整数へ変換する(万人 -> 人、四捨五入)。

    検証の意図は「原典の万人値が本当に(小数4桁までの)人単位の整数か」の
    確認であり、`round()` の性質上ほぼ必ず真になってしまう `>= 0.5` という
    閾値では実質的に検証にならない。許容誤差は `1e-6` とする。
    """
    scaled = float(value_man) * 10000
    person = round(scaled)
    if abs(scaled - person) >= 1e-6:
        raise ValueError(f"人口換算の丸め誤差が大きすぎます: {value_man} -> {scaled}")
    return person


def parse_sheet(ws, published_fy: str) -> ParseResult:
    """帳票シートを48ブロック走査し、tidy行を組み立てる。

    `ws` は openpyxl の Worksheet(`data_only=True` で開いたもの)。
    """
    title = ws["C1"].value
    c2 = ws["C2"].value
    notes = c2.split("\n") if c2 else []

    max_col = ws.max_column
    # A列(ブロック番号)と最終列(ブロックごとの通し番号ラベル)はブロックにより
    # 値が変わって当然のため、サブヘッダー比較の対象範囲から除外する。
    header_col_start = 2
    header_col_end = max_col - 1
    result = ParseResult(published_fy=published_fy, title=title, notes=notes)
    reference_header = None
    reference_col_map = None

    for block in range(NUM_BLOCKS):
        top = BLOCK_TOP0 + BLOCK_SIZE * block

        block_code = ws.cell(row=top, column=1).value
        _expect(block_code, block, f"ブロック{block}: A列のブロック番号が不一致")

        label_row = top + 2
        _expect(
            ws.cell(row=label_row, column=4).value,
            "都道府県",
            f"ブロック{block}: 基礎情報ラベル行(D列)",
        )
        pref_code_num = ws.cell(row=label_row, column=6).value
        pref_name = ws.cell(row=label_row, column=8).value
        pref_code = normalize_pref_code(pref_code_num)
        # A列のブロック番号(block_code)とF列の都道府県コード(pref_code_num)は
        # 本来常に一致するはずの結合キー。ここで検証しないと、食い違いが
        # あった場合に静かにF列側が採用されてしまい、下流の全結合がこの
        # キーに乗るため気づけない。
        _expect(
            pref_code,
            f"{block:02d}",
            f"ブロック{block}: A列のブロック番号とF列の都道府県コード(正規化後)が不一致",
        )

        pop_row = top + 4
        _expect(
            ws.cell(row=pop_row, column=4).value,
            "2020国勢調査人口",
            f"ブロック{block}: 人口ラベル行(D列)",
        )
        population_source_value = ws.cell(row=pop_row, column=6).value

        area_row = top + 5
        _expect(
            ws.cell(row=area_row, column=4).value,
            "2020面積",
            f"ブロック{block}: 面積ラベル行(D列)",
        )
        area_source_value = ws.cell(row=area_row, column=6).value

        section_row = top + 6
        _expect(
            ws.cell(row=section_row, column=3).value,
            "○病床数の状況",
            f"ブロック{block}: 病床数セクション見出し行(C列)",
        )

        header_row = top + 8
        raw_header = _read_header_row(ws, header_row, header_col_start, header_col_end)
        if reference_header is None:
            reference_header = raw_header
            reference_col_map = _column_map_from_header(raw_header, header_col_start)
        elif raw_header != reference_header:
            raise LayoutMismatchError(
                f"ブロック{block}のサブヘッダー行が先頭ブロックと異なります\n"
                f"  先頭ブロック: {reference_header}\n"
                f"  ブロック{block}: {raw_header}"
            )
        col_map = reference_col_map

        for i, bed_function in enumerate(BED_FUNCTIONS):
            row = top + 9 + i
            _expect(
                ws.cell(row=row, column=5).value,
                bed_function,
                f"ブロック{block}: 病床機能ラベル行(E列, {bed_function})",
            )
            for col, (series, year) in col_map.items():
                value = ws.cell(row=row, column=col).value
                beds = _expect_int(value, block=block, row=row, col=col)
                result.beds_rows.append(
                    {
                        "published_fy": published_fy,
                        "pref_code": pref_code,
                        "pref_name": pref_name,
                        "bed_function": bed_function,
                        "series": series,
                        "year": year,
                        "beds": beds,
                    }
                )

        rate_row = top + 9 + len(BED_FUNCTIONS)
        _expect(
            ws.cell(row=rate_row, column=5).value,
            "（報告率）",
            f"ブロック{block}: 報告率ラベル行(E列)",
        )
        for col, (series, year) in col_map.items():
            if series != "実績":
                continue
            value = ws.cell(row=rate_row, column=col).value
            if value is None:
                continue
            result.report_rate_rows.append(
                {
                    "published_fy": published_fy,
                    "pref_code": pref_code,
                    "pref_name": pref_name,
                    "year": year,
                    "report_rate": value,
                }
            )

        population_2020 = _convert_man_to_person(population_source_value)
        area_2020_km2 = round(float(area_source_value), 2)
        result.basic_rows.append(
            {
                "published_fy": published_fy,
                "pref_code": pref_code,
                "pref_name": pref_name,
                "population_2020": population_2020,
                "population_2020_source_value": population_source_value,
                "population_2020_source_unit": "万人",
                "area_2020_km2": area_2020_km2,
            }
        )

    return result


def load_sheet(source_key: str):
    """`SOURCES[source_key]` の設定に従い対象xlsxのシートを開く。

    元データのSHA-256を `SHA256SUMS` と照合したうえで開く。
    戻り値: (worksheet, source_config, source_sha256)
    """
    cfg = SOURCES[source_key]
    source_sha256 = verify_source(cfg["path_in_repo"])
    print(f"[ok] 生データ検証: {cfg['path_in_repo']} = {source_sha256[:16]}...")
    xlsx_path = REPO_ROOT / cfg["path_in_repo"]
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[cfg["sheet_name"]]
    return ws, cfg, source_sha256


def _rows_to_tuples(rows, header):
    return [tuple(r[h] for h in header) for r in rows]


def build_and_write(source_key: str, out_dir: Path) -> dict:
    """指定ソースをパースし、3つのCSV+meta.jsonを `out_dir` へ出力する。

    書き出したCSVパスの辞書({"beds": ..., "report_rate": ..., "basic": ...})
    を返す(再現性テスト等での再利用を想定)。
    """
    out_dir = Path(out_dir)
    ws, cfg, source_sha256 = load_sheet(source_key)
    result = parse_sheet(ws, published_fy=source_key)
    print(
        f"[ok] パース完了: beds={len(result.beds_rows)}行 "
        f"report_rate={len(result.report_rate_rows)}行 basic={len(result.basic_rows)}行"
    )

    today = datetime.date.today().isoformat()
    base_source = {
        "name": cfg["name"],
        "publisher": "厚生労働省",
        "url": cfg["download_url"],
        "page_url": SOURCE_PAGE_URL,
        "fiscal_year": cfg["fiscal_year"],
        "source_file": cfg["path_in_repo"],
        "source_sha256": source_sha256,
        "source_sheet": cfg["sheet_name"],
        "acquired_date": cfg["acquired_date"],
        "license": LICENSE_NOTE,
        "original_title": result.title,
        "original_notes": result.notes,
    }

    beds_header = [
        "published_fy",
        "pref_code",
        "pref_name",
        "bed_function",
        "series",
        "year",
        "beds",
    ]
    beds_csv, _ = write_csv_with_meta(
        out_dir / "prefecture_beds.csv",
        beds_header,
        _rows_to_tuples(result.beds_rows, beds_header),
        title="都道府県別 病床数(実績/見込量/必要数)",
        source=base_source,
        processing={
            "script": "tools/parse_prefecture_beds.py",
            "date": today,
            "steps": STEPS_COMMON
            + ["派生比率列(2025年必要数に対する比等)は再計算可能なため出力対象から除外"],
            "caveat": CAVEAT,
        },
        fields=FIELDS_BEDS,
    )
    print(f"[ok] 出力: {beds_csv} ({len(result.beds_rows)}行)")

    rate_header = ["published_fy", "pref_code", "pref_name", "year", "report_rate"]
    rate_csv, _ = write_csv_with_meta(
        out_dir / "prefecture_bed_report_rate.csv",
        rate_header,
        _rows_to_tuples(result.report_rate_rows, rate_header),
        title="都道府県別 病床機能報告の報告率",
        source=base_source,
        processing={
            "script": "tools/parse_prefecture_beds.py",
            "date": today,
            "steps": STEPS_COMMON
            + ["各ブロックの（報告率）行から実績年の値のみを抽出(見込量・必要数列は報告率を持たない)"],
            "caveat": CAVEAT,
        },
        fields=FIELDS_REPORT_RATE,
    )
    print(f"[ok] 出力: {rate_csv} ({len(result.report_rate_rows)}行)")

    basic_header = [
        "published_fy",
        "pref_code",
        "pref_name",
        "population_2020",
        "population_2020_source_value",
        "population_2020_source_unit",
        "area_2020_km2",
    ]
    basic_csv, _ = write_csv_with_meta(
        out_dir / "prefecture_basic.csv",
        basic_header,
        _rows_to_tuples(result.basic_rows, basic_header),
        title="都道府県別 基礎情報(2020年人口・面積)",
        source=base_source,
        processing={
            "script": "tools/parse_prefecture_beds.py",
            "date": today,
            "steps": STEPS_COMMON
            + [
                "人口(万人)を10000倍し四捨五入して人単位に変換(population_2020)。原典値は population_2020_source_value に保持",
                "面積の浮動小数点誤差を除くため小数2桁に丸め(area_2020_km2)",
            ],
            "caveat": CAVEAT,
        },
        fields=FIELDS_BASIC,
    )
    print(f"[ok] 出力: {basic_csv} ({len(result.basic_rows)}行)")

    return {"beds": beds_csv, "report_rate": rate_csv, "basic": basic_csv}


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--source",
        choices=["R7"],
        default="R7",
        help="対象の公表年度データ(現時点はR7のみ。R6は将来対応)",
    )
    args = ap.parse_args()

    out_dir = REPO_ROOT / "data" / "processed"
    build_and_write(args.source, out_dir)


if __name__ == "__main__":
    main()
