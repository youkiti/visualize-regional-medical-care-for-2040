# -*- coding: utf-8 -*-
"""固定サイズで同じレイアウトが繰り返される帳票(区域ごとのブロックが
一定行数ずつ並ぶExcelシート)を走査するための共通基盤。

この抽象化が担当するのは「算術とブロック構造」だけ(ブロック番号の連番
検証・サブヘッダー行の一致検証・ヘッダー文字列から列を解決すること)で、
ブロック内の行オフセットの意味づけ(何行目に何のラベルがあるか)や
スカラー項目(人口・面積等)の抽出は各パーサ側の責務として残す。
"""
import re
from dataclasses import dataclass
from typing import Callable, Iterator, Optional

from openpyxl.utils import get_column_letter

from tools.lib.layout import LayoutMismatchError, expect, read_header_row


@dataclass(frozen=True)
class Block:
    """繰り返しブロック1件分の位置情報。"""

    index: int  # 0始まりの反復インデックス
    number: int  # ブロック番号列(既定ではA列)から読んだブロック番号
    top_row: int  # ブロック先頭行(1始まりの行番号)


def iter_fixed_blocks(
    ws,
    *,
    first_row: int,
    block_size: int,
    count: int,
    first_number: int,
    number_col: int = 1,
) -> Iterator[Block]:
    """固定サイズの繰り返しブロックを先頭から順に走査し `Block` を yield する。

    ブロック番号列(`number_col`、既定はA列)の値が `first_number` からの
    連番(`first_number, first_number+1, ...`)であることを検証しながら
    走査する。不一致ならレイアウト変更とみなし `LayoutMismatchError` で
    中断する。

    ジェネレータなので1回しか走査できない。`assert_repeated_header` など
    複数回の走査が必要な処理に渡す場合は、呼び出し側で
    `blocks = list(iter_fixed_blocks(...))` のように一度リスト化すること。
    """
    col_letter = get_column_letter(number_col)
    for index in range(count):
        top_row = first_row + block_size * index
        expected_number = first_number + index
        actual_number = ws.cell(row=top_row, column=number_col).value
        expect(
            actual_number,
            expected_number,
            f"ブロック{index}(行{top_row}): {col_letter}列のブロック番号が不一致",
        )
        yield Block(index=index, number=actual_number, top_row=top_row)


def assert_repeated_header(ws, blocks, *, row_offset: int, col_start: int, col_end: int):
    """全ブロックのサブヘッダー行が先頭ブロックと一致することを検証する。

    `blocks` は `Block` の反復可能オブジェクト(複数回走査できる必要が
    あるため、ジェネレータではなく `list` 等を渡すこと)。各ブロックの
    `top_row + row_offset` 行目を `read_header_row(ws, ..., col_start,
    col_end)` で読み、先頭ブロックのヘッダーと比較する。不一致なら
    `LayoutMismatchError` で中断する。

    戻り値: 先頭ブロックの正規化済みヘッダー(文字列のタプル)。
    """
    reference_header = None
    for block in blocks:
        header_row = block.top_row + row_offset
        raw_header = read_header_row(ws, header_row, col_start, col_end)
        if reference_header is None:
            reference_header = raw_header
        elif raw_header != reference_header:
            raise LayoutMismatchError(
                f"ブロック{block.index}(行{header_row})のサブヘッダー行が先頭ブロックと異なります\n"
                f"  先頭ブロック: {reference_header}\n"
                f"  ブロック{block.index}: {raw_header}"
            )
    return reference_header


ColumnKey = tuple[str, int]  # (series, year)。dataclassにしない軽量な型エイリアス


def resolve_columns(
    raw_header, *, col_start: int, classify: Callable[[str], Optional[ColumnKey]]
) -> dict[int, ColumnKey]:
    """正規化済みヘッダー列から `{列番号: (series, year)}` を作る。

    `raw_header` は `read_header_row` が返す文字列タプル(先頭が `col_start`
    列に対応)。各列の文字列を `classify` に渡し、`None` が返った列
    (派生比率列など)は結果から除外する。
    """
    col_map = {}
    for idx, text in enumerate(raw_header, start=col_start):
        if not text:
            continue
        key = classify(text)
        if key is not None:
            col_map[idx] = key
    return col_map


_ACTUAL_RE = re.compile(r"^(\d{4})実績$")
_PLAN_RE = re.compile(r"^(\d{4})見込量$")
_REQUIRED_RE = re.compile(r"^(\d{4})必要数$")


def classify_bed_column(text: str) -> Optional[ColumnKey]:
    """病床帳票のサブヘッダー文字列を `(series, year)` に分類する。

    `'2015実績'` -> `('実績', 2015)` / `'2026見込量'` -> `('見込量', 2026)` /
    `'2025必要数'` -> `('必要数', 2025)`。「2025年必要数に対する比」「2015年
    に対する比」「2015年との差」「見込み／必要数」等の派生比率列はこの
    正規表現にマッチしないため `None` を返す(値は出力済み実績・見込量・
    必要数から再計算可能なため、出力対象から自然に除外される)。
    """
    m = _ACTUAL_RE.match(text)
    if m:
        return ("実績", int(m.group(1)))
    m = _PLAN_RE.match(text)
    if m:
        return ("見込量", int(m.group(1)))
    m = _REQUIRED_RE.match(text)
    if m:
        return ("必要数", int(m.group(1)))
    return None
