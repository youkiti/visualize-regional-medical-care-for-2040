# -*- coding: utf-8 -*-
"""厚労省「構想区域別の医療需要推計」(R7: 001728462.xlsx)を tidy CSV へ変換する。

`tools/parse_area_beds.py`(帳票レイアウト・ブロック走査)とは異なり、この
xlsxは339構想区域が単純な表形式(1区域=1行)で並んでいるため、
`tools/lib/block_report.py` は使わない。年度・見出しの検証だけ
`tools/lib/layout.py` を利用する。

原典の構造(2シート、レイアウトは同一):
  - シート「将来の在宅（訪問診療）需要推計」「将来の外来需要推計」
  - 1行目=表題、2行目=注記(出典等)、3行目=表全体の見出し
  - 4行目=年度ラベル行(F〜K列: '2024年度','2030年度（現状投影）',...)。
    列位置ではなく、この行の文字列から正規表現で年を抽出する(CLAUDE.md
    「R6の列ずれの罠」と同じ理由による規律。この帳票にR6版は無いが、将来
    レイアウトが変わっても列位置のハードコードに頼らないため)
  - 5行目=見出し行(A都道府県 B構想区域コード C構想区域名 D人口(2024年度)
    E人口(2040年) F〜Kレセプト件数/月)
  - 6行目から339行、構想区域コードの昇順

処理内容:
  1. `verify_source()` で元データのSHA-256を `SHA256SUMS` と照合(改変検知)
  2. openpyxl(`data_only=True`)で2シートを開き、シート名・見出し行(5行目)・
     年度ラベル行(4行目)を検証する
  3. 6行目から339行を走査し、tidy行を組み立てる
  4. area_code集合・area_name・pref_nameが両シート間で完全一致し、かつ
     `data/processed/area_basic.csv` とも完全一致することを検証する
     (区域の対応が崩れた場合に必ず気付くため)
  5. 2つの tidy CSV + 由来メタデータ(`<csv名>.meta.json`)を `data/processed/`
     に出力する
       - demand_forecast.csv: 在宅(訪問診療)・外来のレセプト件数/月推計
       - demand_population.csv: 人口(2024年度・2040年)

⚠ センチネル値の罠(CLAUDE.md): この帳票には現時点で非数値センチネルは
確認されていないが、`001723349.xlsx`(構想区域の病床数等)の推計流出入
患者割合のように将来 'XXX' 等が混入しても静かに壊れないよう、需要値・
人口は数値かつ0より大きいことを明示的に検証してから出力する。

必要環境: Python 3.11+, openpyxl

使い方:
    python tools/parse_demand_forecast.py
"""
import csv
import datetime
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

if __package__ in (None, ""):
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

from tools.lib.codes import normalize_area_code, normalize_pref_code
from tools.lib.layout import LayoutMismatchError, expect, normalize_header_text
from tools.lib.provenance import REPO_ROOT, verify_source, write_csv_with_meta

SOURCE_NAME = "構想区域別の医療需要推計"
SOURCE_PATH_IN_REPO = "R7/001728462.xlsx"
SOURCE_DOWNLOAD_URL = "https://www.mhlw.go.jp/content/10800000/001728462.xlsx"
SOURCE_PAGE_URL = "https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/0000080850_00014.html"
SOURCE_FISCAL_YEAR = "令和7年度（2025年度）"
SOURCE_ACQUIRED_DATE = "2026-08-04"

LICENSE_NOTE = (
    "厚生労働省ホームページ利用規約（政府標準利用規約準拠） "
    "https://www.mhlw.go.jp/chosakuken/index.html"
)

# シート名と、原典の日本語ラベルそのままの需要区分。出力順(demand_category
# の並び順)にもこの順序(在宅→外来)を使う。
SHEET_HOME_CARE = "将来の在宅（訪問診療）需要推計"
SHEET_OUTPATIENT = "将来の外来需要推計"
DEMAND_CATEGORY_HOME_CARE = "在宅（訪問診療）"
DEMAND_CATEGORY_OUTPATIENT = "外来"
SHEETS = [
    (SHEET_HOME_CARE, DEMAND_CATEGORY_HOME_CARE),
    (SHEET_OUTPATIENT, DEMAND_CATEGORY_OUTPATIENT),
]

YEAR_LABEL_ROW = 4  # 年度ラベル行(F〜K列)
HEADER_ROW = 5  # 見出し行(A〜K列)
DATA_START_ROW = 6  # データ開始行
NUM_AREAS = 339  # 構想区域数
DATA_END_ROW = DATA_START_ROW + NUM_AREAS - 1  # 344

VALUE_COL_START = 6  # F列(レセプト件数/月の最初の年度列)
VALUE_COL_END = 11  # K列(レセプト件数/月の最後の年度列)

EXPECTED_HEADER_AE = [
    "都道府県",
    "構想区域コード",
    "構想区域名",
    "人口(2024年度)",
    "人口(2040年)",
]
RECEIPTS_HEADER_LABEL = "レセプト件数/月"

EXPECTED_YEARS = [2024, 2030, 2035, 2040, 2045, 2050]

POPULATION_2024_HEADER = "人口(2024年度)"
POPULATION_2040_HEADER = "人口(2040年)"

PREF_CODE_DESC = "都道府県コード(ゼロ埋め2桁の文字列、01=北海道…47=沖縄県、原典の都道府県コード順)"
PREF_NAME_DESC = "都道府県名"
AREA_CODE_DESC = (
    "構想区域(二次医療圏)コード(ゼロ埋め4桁の文字列)。上2桁が都道府県コードと一致する"
)
AREA_NAME_DESC = "構想区域名"
PUBLISHED_FY_DESC = (
    "公表年度を表す識別子。'R7'=令和7年度公表分(この帳票にR6版は存在しない)"
)

FIELDS_FORECAST = {
    "published_fy": PUBLISHED_FY_DESC,
    "pref_code": PREF_CODE_DESC,
    "pref_name": PREF_NAME_DESC,
    "area_code": AREA_CODE_DESC,
    "area_name": AREA_NAME_DESC,
    "demand_category": (
        "需要区分。原典シート名に対応する日本語ラベルをそのまま使う"
        "('在宅（訪問診療）'=将来の在宅（訪問診療）需要推計シート、"
        "'外来'=将来の外来需要推計シート)。英字キーへの変換は表示用データ生成側の責務"
    ),
    "year": (
        "対象年(西暦4桁の整数)。4行目の年度ラベル文字列(例:'2030年度（現状投影）')から"
        "正規表現で抽出した値であり、列位置はハードコードしていない"
    ),
    "year_label": (
        "4行目の年度ラベルの原文そのまま(例:'2030年度（現状投影）')。2024年度のみ"
        "「（現状投影）」が付かない原典の区別を、こちらで解釈せず保持するための列。"
        "yearはこの文字列から抽出した数値"
    ),
    "receipts_per_month": (
        "レセプト件数/月(原典セルの値をそのまま。丸め・整形はしていない)。"
        "患者数・人数ではない点に留意(caveat参照)"
    ),
}

FIELDS_POPULATION = {
    "published_fy": PUBLISHED_FY_DESC,
    "pref_code": PREF_CODE_DESC,
    "pref_name": PREF_NAME_DESC,
    "area_code": AREA_CODE_DESC,
    "area_name": AREA_NAME_DESC,
    "population_2024": (
        "原典の見出し『人口(2024年度)』の値(人単位の整数)。原典の見出し文字列自体が"
        "「年度」表記であり、population_2040(「年」表記)と基準が混在している点に留意"
    ),
    "population_2040": (
        "原典の見出し『人口(2040年)』の値(人単位の整数)。原典の見出し文字列自体が"
        "「年」表記であり、population_2024(「年度」表記)と基準が混在している点に留意"
    ),
}

CAVEAT_FORECAST = (
    "receipts_per_monthは「レセプト件数/月」であり、患者数・人数そのものではない"
    "(原典の値をそのまま出力)。年度は2024年度のみ実績相当で、2030〜2050年度は"
    "いずれも「現状投影」(year_label参照)。"
)
CAVEAT_POPULATION = (
    "population_2024/population_2040は需要推計の参考情報であり、原典の見出しが"
    "『人口(2024年度)』『人口(2040年)』で年度/年の表記が混在している"
    "(fields参照。単位変換は不要、原典から実数のまま)。"
)

STEPS_COMMON = [
    "verify_source()でR7/001728462.xlsxのSHA-256をSHA256SUMSと照合",
    "openpyxl(data_only=True)でシートがちょうど2枚(将来の在宅（訪問診療）需要推計/"
    "将来の外来需要推計)であること、シート名を検証",
    "5行目(見出し行)のA〜E列の見出し文字列、F〜K列が全て『レセプト件数/月』であることを検証",
    "4行目(年度ラベル行)のF〜K列の文字列から正規表現で年を抽出し、両シートで一致し"
    "[2024, 2030, 2035, 2040, 2045, 2050]の6個・昇順・重複なしであることを検証",
    "6行目から339行(構想区域ごと)を走査し、area_codeがゼロ埋め4桁の文字列で"
    "重複がなく、ちょうど339件であることを検証",
    "area_code集合・area_name・pref_nameが両シート(在宅（訪問診療）/外来)間で"
    "完全一致し、かつdata/processed/area_basic.csvとも完全一致することを検証"
    "(不一致は区域の対応が崩れた合図として中断)",
    "area_code[:2]がpref_codeと一致することを検証",
    "需要値(レセプト件数/月)・人口が全て数値かつ0より大きいことを検証",
    "両シートで人口2列(人口(2024年度)/人口(2040年))が完全一致することを検証",
]


@dataclass
class SheetParseResult:
    demand_category: str
    title: str
    notes: list
    years: list  # [(col, year, year_label), ...] 4行目から抽出
    forecast_rows: list = field(default_factory=list)
    # area_code -> {"pref_code":..., "pref_name":..., "area_name":...,
    #                "population_2024":..., "population_2040":...}
    population_by_area: dict = field(default_factory=dict)


def _extract_year(year_label) -> int:
    """年度ラベル文字列(例 '2030年度（現状投影）')から西暦4桁の年を抽出する。

    列位置ではなく文字列から抽出することで、将来レイアウトが変わっても
    静かに誤った年を割り当てることを防ぐ(CLAUDE.md「R6の列ずれの罠」と
    同じ規律)。
    """
    if not isinstance(year_label, str):
        raise LayoutMismatchError(f"年度ラベルが文字列ではありません: {year_label!r}")
    text = normalize_header_text(year_label)
    m = re.match(r"^(\d{4})年度", text)
    if not m:
        raise LayoutMismatchError(f"年度ラベルから年を抽出できません: {year_label!r}")
    return int(m.group(1))


def _expect_positive_number(value, *, row, col, label):
    """セル値が数値(bool除く)かつ0より大きいことを検証し、そのまま返す。

    将来 'XXX' のような非数値センチネルが混入しても(CLAUDE.md「センチネル
    値の罠」参照)、int()/float() で静かに壊れず必ず検知するための共通処理。
    値は丸め・型変換せず原典のまま返す(receipts_per_monthの整形禁止の要件)。
    """
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise LayoutMismatchError(
            f"行{row} 列{col}: {label}が数値ではありません: {value!r}"
        )
    if not (value > 0):
        raise LayoutMismatchError(
            f"行{row} 列{col}: {label}が0以下です: {value!r}"
        )
    return value


def _validate_header_row(ws):
    """5行目(見出し行)のA〜E列・F〜K列を検証する。"""
    for col, expected in enumerate(EXPECTED_HEADER_AE, start=1):
        expect(
            ws.cell(row=HEADER_ROW, column=col).value,
            expected,
            f"5行目 列{col}: 見出し",
        )
    for col in range(VALUE_COL_START, VALUE_COL_END + 1):
        expect(
            ws.cell(row=HEADER_ROW, column=col).value,
            RECEIPTS_HEADER_LABEL,
            f"5行目 列{col}: 見出し",
        )


def _read_year_columns(ws):
    """4行目(年度ラベル行)のF〜K列から (col, year, year_label) のリストを返す。"""
    year_cols = []
    for col in range(VALUE_COL_START, VALUE_COL_END + 1):
        year_label = ws.cell(row=YEAR_LABEL_ROW, column=col).value
        year = _extract_year(year_label)
        year_cols.append((col, year, year_label))
    return year_cols


def parse_sheet(ws, demand_category: str) -> SheetParseResult:
    """1シート分(339構想区域 × 6年度)を tidy 行に変換する。

    `ws` は openpyxl の Worksheet(`data_only=True` で開いたもの)。
    """
    title = ws.cell(row=1, column=1).value
    c2 = ws.cell(row=2, column=1).value
    notes = c2.split("\n") if c2 else []

    expect(ws.max_column, VALUE_COL_END, "シートの最終列(K列であること)")
    expect(ws.max_row, DATA_END_ROW, "シートの最終行(データが339行であること)")

    _validate_header_row(ws)
    year_cols = _read_year_columns(ws)
    years = [y for _, y, _ in year_cols]
    expect(years, EXPECTED_YEARS, "4行目から抽出した年度の並び")

    result = SheetParseResult(demand_category=demand_category, title=title, notes=notes, years=year_cols)

    seen_area_codes = set()
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        pref_name = ws.cell(row=row, column=1).value
        area_code_raw = ws.cell(row=row, column=2).value
        area_name = ws.cell(row=row, column=3).value

        # 原典の構想区域コードはすでにゼロ埋め4桁の文字列("0101")のはず
        # (病床系ファイルの数値コードとは違う。CLAUDE.md「結合キーの罠」参照)。
        # 型が違えば(数値化されている等)レイアウト前提が崩れているとみなし中断する。
        if not (isinstance(area_code_raw, str) and re.fullmatch(r"\d{4}", area_code_raw)):
            raise LayoutMismatchError(
                f"行{row}: 構想区域コードがゼロ埋め4桁の文字列ではありません: {area_code_raw!r}"
            )
        # normalize_area_code() を通し、値の範囲(0〜9999)もあわせて検証する
        # (他パーサと同じく常にcodes.pyの正規化関数を経由する流儀)。
        area_code = normalize_area_code(area_code_raw)
        if area_code in seen_area_codes:
            raise LayoutMismatchError(f"行{row}: 構想区域コード{area_code}が重複しています")
        seen_area_codes.add(area_code)

        pref_code_part = area_code[:2]
        pref_code = normalize_pref_code(int(pref_code_part))
        expect(
            pref_code_part,
            pref_code,
            f"行{row}: 構想区域コード{area_code}の上2桁が都道府県コードとして不正です",
        )

        population_2024 = _expect_positive_number(
            ws.cell(row=row, column=4).value, row=row, col=4, label=POPULATION_2024_HEADER
        )
        population_2040 = _expect_positive_number(
            ws.cell(row=row, column=5).value, row=row, col=5, label=POPULATION_2040_HEADER
        )
        result.population_by_area[area_code] = {
            "pref_code": pref_code,
            "pref_name": pref_name,
            "area_name": area_name,
            "population_2024": population_2024,
            "population_2040": population_2040,
        }

        for col, year, year_label in year_cols:
            value = _expect_positive_number(
                ws.cell(row=row, column=col).value,
                row=row,
                col=col,
                label=RECEIPTS_HEADER_LABEL,
            )
            result.forecast_rows.append(
                {
                    "pref_code": pref_code,
                    "pref_name": pref_name,
                    "area_code": area_code,
                    "area_name": area_name,
                    "demand_category": demand_category,
                    "year": year,
                    "year_label": year_label,
                    "receipts_per_month": value,
                }
            )

    return result


def load_workbook():
    """R7/001728462.xlsx を開く(SHA-256照合を済ませたうえで)。

    戻り値: (workbook, source_sha256)
    """
    source_sha256 = verify_source(SOURCE_PATH_IN_REPO)
    print(f"[ok] 生データ検証: {SOURCE_PATH_IN_REPO} = {source_sha256[:16]}...")
    xlsx_path = REPO_ROOT / SOURCE_PATH_IN_REPO
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    return wb, source_sha256


def _load_area_basic_reference():
    """`data/processed/area_basic.csv` から area_code -> (pref_name, area_name)
    の参照テーブルを読み込む(構想区域の対応が崩れていないことの検証用)。
    """
    path = REPO_ROOT / "data" / "processed" / "area_basic.csv"
    reference = {}
    with open(path, encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            reference[row["area_code"]] = (row["pref_name"], row["area_name"])
    return reference


def _names_by_area(population_by_area: dict) -> dict:
    """`population_by_area` から area_code -> (pref_name, area_name) を取り出す。"""
    return {
        area_code: (info["pref_name"], info["area_name"])
        for area_code, info in population_by_area.items()
    }


def _validate_against_area_basic(sheet_results):
    """area_code集合・area_name・pref_nameが「両シート間」かつ「area_basic.csv」の
    双方で完全一致することを検証する。

    在宅（訪問診療）シートだけをarea_basic.csvと突合しても、外来シート側の
    区域名・都道府県名がずれていた場合に気付けない(demand_forecast.csvの
    外来行は外来シートから読んだpref_name/area_nameをそのまま書き出すため)。
    ここで全シートの名称をまず突き合わせ、代表(先頭シート)の名称のみを
    area_basic.csvと照合することで、「シート間のずれ」と「原典 vs
    area_basic.csvのずれ」の両方を1か所で検知する。
    """
    reference = _load_area_basic_reference()

    per_sheet_names = [
        (result.demand_category, _names_by_area(result.population_by_area))
        for result in sheet_results
    ]
    first_category, first_names = per_sheet_names[0]

    expect(
        set(first_names.keys()),
        set(reference.keys()),
        "area_code集合がdata/processed/area_basic.csvと不一致",
    )
    for area_code, (pref_name, area_name) in first_names.items():
        ref_pref_name, ref_area_name = reference[area_code]
        expect(
            pref_name,
            ref_pref_name,
            f"area_code={area_code}: pref_nameがarea_basic.csvと不一致({first_category}シート)",
        )
        expect(
            area_name,
            ref_area_name,
            f"area_code={area_code}: area_nameがarea_basic.csvと不一致({first_category}シート)",
        )

    for category, names in per_sheet_names[1:]:
        expect(
            set(names.keys()),
            set(first_names.keys()),
            f"{category}シートのarea_code集合が{first_category}シートと不一致",
        )
        for area_code, (pref_name, area_name) in names.items():
            ref_pref_name, ref_area_name = first_names[area_code]
            expect(
                pref_name,
                ref_pref_name,
                f"area_code={area_code}: pref_nameが{first_category}シートと{category}シートで不一致",
            )
            expect(
                area_name,
                ref_area_name,
                f"area_code={area_code}: area_nameが{first_category}シートと{category}シートで不一致",
            )


def build_and_write(out_dir: Path) -> dict:
    """R7/001728462.xlsxをパースし、2つのCSV+meta.jsonを `out_dir` へ出力する。

    書き出したCSVパスの辞書({"forecast": ..., "population": ...})を返す
    (再現性テスト等での再利用を想定)。
    """
    out_dir = Path(out_dir)
    wb, source_sha256 = load_workbook()

    # シートがちょうど2枚で、シート名が期待どおりであることを検証する
    # (`wb[sheet_name]` はキー不在時にKeyErrorになるだけで意図が伝わらないため、
    # 先に明示的に検証してから使う)。
    expect(wb.sheetnames, [SHEET_HOME_CARE, SHEET_OUTPATIENT], "ワークブックのシート構成")

    sheet_results = []
    for sheet_name, demand_category in SHEETS:
        ws = wb[sheet_name]
        result = parse_sheet(ws, demand_category=demand_category)
        sheet_results.append(result)
        print(
            f"[ok] パース完了: {sheet_name} forecast={len(result.forecast_rows)}行 "
            f"population={len(result.population_by_area)}区域"
        )

    home_care, outpatient = sheet_results

    # 両シートで年度ラベルが一致することを検証する(片方だけレイアウトが
    # 崩れているケースを見逃さないため)。
    expect(
        [(y, label) for _, y, label in home_care.years],
        [(y, label) for _, y, label in outpatient.years],
        "在宅（訪問診療）シートと外来シートで年度ラベルが不一致",
    )

    # 両シートで人口2列が完全一致することを検証する(原典の構造上の前提)。
    for area_code, home_info in home_care.population_by_area.items():
        out_info = outpatient.population_by_area.get(area_code)
        if out_info is None:
            raise LayoutMismatchError(
                f"area_code={area_code}: 外来シートに存在しません"
            )
        expect(
            (home_info["population_2024"], home_info["population_2040"]),
            (out_info["population_2024"], out_info["population_2040"]),
            f"area_code={area_code}: 在宅（訪問診療）シートと外来シートで人口が不一致",
        )
    expect(
        set(home_care.population_by_area.keys()),
        set(outpatient.population_by_area.keys()),
        "在宅（訪問診療）シートと外来シートでarea_code集合が不一致",
    )

    _validate_against_area_basic(sheet_results)

    today = datetime.date.today().isoformat()
    base_source = {
        "name": SOURCE_NAME,
        "publisher": "厚生労働省",
        "url": SOURCE_DOWNLOAD_URL,
        "page_url": SOURCE_PAGE_URL,
        "fiscal_year": SOURCE_FISCAL_YEAR,
        "source_file": SOURCE_PATH_IN_REPO,
        "source_sha256": source_sha256,
        "source_sheet": [SHEET_HOME_CARE, SHEET_OUTPATIENT],
        "acquired_date": SOURCE_ACQUIRED_DATE,
        "license": LICENSE_NOTE,
        "original_title": [home_care.title, outpatient.title],
        "original_notes": home_care.notes
        + [
            "出典説明書「構想区域別の医療需要推計について」(R7/001728467.pdf)が"
            "本データの一次資料(集計方法等の詳細を記載)"
        ],
    }

    # 出力順: area_code昇順 -> カテゴリ(在宅→外来のシート順) -> 年昇順。
    # 両シートは area_code が同じ昇順で並んでいることを確認済み(パース時に
    # 元の行順のまま forecast_rows に積んでいる)ため、シート単位でまとめた
    # 後段の並べ替えではなく、area_code -> シート(在宅→外来)の順で束ね直す。
    forecast_by_area = {}
    for result in sheet_results:
        for row in result.forecast_rows:
            forecast_by_area.setdefault(row["area_code"], []).append(row)

    forecast_rows = []
    for area_code in sorted(forecast_by_area.keys()):
        forecast_rows.extend(forecast_by_area[area_code])

    forecast_header = [
        "published_fy",
        "pref_code",
        "pref_name",
        "area_code",
        "area_name",
        "demand_category",
        "year",
        "year_label",
        "receipts_per_month",
    ]
    forecast_tuples = [
        (
            "R7",
            row["pref_code"],
            row["pref_name"],
            row["area_code"],
            row["area_name"],
            row["demand_category"],
            row["year"],
            row["year_label"],
            row["receipts_per_month"],
        )
        for row in forecast_rows
    ]
    forecast_csv, _ = write_csv_with_meta(
        out_dir / "demand_forecast.csv",
        forecast_header,
        forecast_tuples,
        title="構想区域別 医療需要推計(在宅（訪問診療）・外来のレセプト件数/月)",
        source=base_source,
        processing={
            "script": "tools/parse_demand_forecast.py",
            "date": today,
            "steps": STEPS_COMMON
            + [
                "年度ごとの値をtidy化(published_fy,pref_code,pref_name,area_code,"
                "area_name,demand_category,year,year_label,receipts_per_month)。"
                "並び順はarea_code昇順→カテゴリ(在宅→外来のシート順)→年昇順"
            ],
            "caveat": CAVEAT_FORECAST,
        },
        fields=FIELDS_FORECAST,
    )
    print(f"[ok] 出力: {forecast_csv} ({len(forecast_rows)}行)")

    population_header = [
        "published_fy",
        "pref_code",
        "pref_name",
        "area_code",
        "area_name",
        "population_2024",
        "population_2040",
    ]
    population_tuples = [
        (
            "R7",
            info["pref_code"],
            info["pref_name"],
            area_code,
            info["area_name"],
            info["population_2024"],
            info["population_2040"],
        )
        for area_code, info in sorted(home_care.population_by_area.items())
    ]
    population_csv, _ = write_csv_with_meta(
        out_dir / "demand_population.csv",
        population_header,
        population_tuples,
        title="構想区域別 人口(2024年度・2040年、医療需要推計の参考情報)",
        source=base_source,
        processing={
            "script": "tools/parse_demand_forecast.py",
            "date": today,
            "steps": STEPS_COMMON
            + [
                "人口2列(population_2024,population_2040)をarea_code単位でtidy化"
                "(2シートで値が共通のため一方(在宅（訪問診療）シート)の値を採用。"
                "事前に両シートで完全一致することを検証済み)"
            ],
            "caveat": CAVEAT_POPULATION,
        },
        fields=FIELDS_POPULATION,
    )
    print(f"[ok] 出力: {population_csv} ({len(population_tuples)}行)")

    return {"forecast": forecast_csv, "population": population_csv}


def main():
    out_dir = REPO_ROOT / "data" / "processed"
    build_and_write(out_dir)


if __name__ == "__main__":
    main()
