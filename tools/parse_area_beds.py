# -*- coding: utf-8 -*-
"""厚労省「②構想区域の病床数等」(R7: 001723349.xlsx / R6: 別添４③)の帳票Excel
を tidy CSV へ変換する。R7・R6の両方を出力対象とし、1本のCSVに
`published_fy` で年度を並存させる。

帳票のブロック構造(3行目開始・1ブロック15行)は都道府県版
(`tools/parse_prefecture_beds.py`、001722915.xlsx)と完全に同一だが、
以下の点が異なる:

  - ブロック数は339(構想区域ごと)。A列のブロック番号は1始まりの連番
    1..339(都道府県版は全国=0始まり)
  - 各ブロックの内訳(top=ブロック先頭行)にD/F/H列で「都道府県」と
    「構想区域」の2階層のコード・名称を持つ
  - 流出入関連の値を持つ(都道府県版にはこれに相当する項目がない)。
    年度により列位置・項目数・値域が異なる(下記「流出入項目の年度差」参照)

処理内容:
  1. `verify_source()` で元データのSHA-256を `SHA256SUMS` と照合(改変検知)
  2. openpyxl(`data_only=True`)でシートを開き、3行目から15行ずつ・339ブロック
     を走査
  3. サブヘッダー行(各ブロック先頭+8行目)の文字列("2015実績"等)から
     実績/見込量/必要数の列を解決する(都道府県版と同様、公表年度により
     列位置が異なるためハードコードしない。下記「R6との列ずれ」参照)
  4. 全339ブロックのサブヘッダー行が先頭ブロックと完全一致することを検証し、
     不一致ならレイアウト変更とみなして例外で中断する(取りこぼし防止)
  5. 流出入項目のラベルを、値を読む前に検証する(列は位置でハードコード
     するため、ラベルが動いたら必ず失敗させる)
  6. `--source`(既定'all'=R7+R6)で指定した各年度をパースし、行はR7を先に・
     R6を後に固定した順で連結する(R7部分がバイト単位で不変になり、
     `git diff` で「R6の行が末尾に増えた」ことだけが読めるようにするため)
  7. 3つの tidy CSV + 由来メタデータ(`<csv名>.meta.json`)を `data/processed/`
     に出力する
       - area_beds.csv: 病床数(実績/見込量/必要数 × 5機能 × 年)
       - area_bed_report_rate.csv: 病床機能報告の報告率
       - area_basic.csv: 基礎情報(2020人口・面積・流出入関連)

⚠ R6との列ずれ: R6(別添４③)は実績年が1年少なく(2015, 2018〜2024)、その
分だけ見込量/必要数の列がR7よりも1列前にずれる。見込量の対象年も異なる
(R6=2025年見込量 / R7=2026年見込量)。そのため列は位置ではなく、サブヘッダー
行の文字列から都度解決する(`tools/lib/block_report.py` の
`resolve_columns` / `classify_bed_column`)。

⚠ 流出入項目の年度差: R7はR列(18)に「推計流出患者割合」(ラベルtop+2/値
top+3)と「推計流入患者割合」(ラベルtop+4/値top+5)を別々に持ち、値域は
0〜1。R6はQ列(17)に「（一般病床患者流出入）」という単一の値をラベルtop+4/
値top+5に持ち、値域は-1〜1で負値を取りうる(実測 -0.893〜0.434)。**R7の
outflow_rate/inflow_rateとR6のnet_flow_rateは別概念であり、並べて比較・
可視化してはならない**(area_basic_r6_net_flow_rate_different_concept参照)。
この年度差は `SOURCES[<年度>]["flow_items"]` に構造化してあり、
`parse_sheet()` は `published_fy` からその定義を引いて列・ラベル・値域を
解決する(値を読む前にラベルを検証する規律は維持したまま、列位置の
ハードコード自体は年度別設定に閉じ込める)。

⚠ 実績セルの欠測(R6のみ): R6原典に実績セルの欠測が1件ある(区域コード
0102「南檜山」の高度急性期・2015実績、行28・列6が空)。この欠測は
`EXPECTED_MISSING_BEDS` に列挙されたものだけを許容し(空セルを検出する
たびにこの集合と照合し、未知の欠測なら`LayoutMismatchError`で中断)、
出力では該当セルの `beds` を空欄のまま出力する(合計から逆算して埋める
ことはしない。`doc/REQUIREMENTS.md` §4.3「位置の推測はしない」と同じ規律)。
「合計」==4機能の和の検証は、欠測を含む(series, year)についてのみ
「合計 >= 欠測を除く機能の和」の不等式に緩める(area_beds_r6_2015_actual_missing_minamihiyama参照)。

⚠ 既知のデータ品質問題(値は勝手に直さず原典どおり出力し、meta.jsonの
`known_issues` に記録する。詳細は `KNOWN_ISSUES` 参照):
  1. area_beds.csv の「2024実績」列が「2025実績」列と全1695セル完全同一
     (R7のみの問題。R6の2024実績は健全)
  2. area_beds.csv のR6区域コード0102「南檜山」高度急性期・2015実績が空欄
  3. area_bed_report_rate.csv の2024年報告率が339区域中105区域でR6とR7で
     異なる(2015〜2023年は全区域で一致)
  4. area_basic.csv の推計流出/流入患者割合(R7)・一般病床患者流出入(R6)が、
     三重県の8区域(2405〜2412)でいずれも文字列 'XXX'(未算出)
  5. area_basic.csv のR6のnet_flow_rateはR7のoutflow_rate/inflow_rateとは
     別概念で値域も異なる(上記「流出入項目の年度差」参照)

必要環境: Python 3.11+, openpyxl

使い方:
    python tools/parse_area_beds.py [--source all|R7|R6]
"""
import argparse
import datetime
import sys
from dataclasses import dataclass, field
from pathlib import Path

if __package__ in (None, ""):
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

from tools.lib.block_report import (
    assert_repeated_header,
    classify_bed_column,
    iter_fixed_blocks,
    resolve_columns,
)
from tools.lib.codes import normalize_area_code, normalize_pref_code
from tools.lib.layout import LayoutMismatchError, expect, expect_int
from tools.lib.provenance import REPO_ROOT, verify_source, write_csv_with_meta

OUTFLOW_LABEL = "（推計流出患者割合）"
INFLOW_LABEL = "（推計流入患者割合）"
NET_FLOW_LABEL = "（一般病床患者流出入）"

# 各公表年度の元データ設定。R7・R6の両方をCLI(--source)から出力できる。
# 都道府県パーサ(parse_prefecture_beds.py)と同じ流儀。
#
# `flow_items` は流出入関連の値の年度別レイアウトを構造化したもの。各要素は
# 「出力列名(field)・列番号(col)・ラベル文字列(label)・ラベル行オフセット
# (label_row_offset)・値行オフセット(value_row_offset)・許容値域
# (value_range)」を持つ(いずれもブロック先頭行(top)からの相対位置)。
# 列位置をハードコードする以上、値を読む前に必ずラベルを検証する
# (`parse_sheet()` 参照)。
SOURCES = {
    "R7": {
        "name": "②構想区域の病床数等（別添４）",
        "path_in_repo": "R7/001723349.xlsx",
        "sheet_name": "構想区域別必要量との比較",
        "download_url": "https://www.mhlw.go.jp/content/10800000/001723349.xlsx",
        "fiscal_year": "令和7年度（2025年度）",
        "acquired_date": "2026-08-04",
        "flow_items": [
            {
                "field": "outflow_rate",
                "col": 18,
                "label": OUTFLOW_LABEL,
                "label_row_offset": 2,
                "value_row_offset": 3,
                "value_range": (0, 1),
            },
            {
                "field": "inflow_rate",
                "col": 18,
                "label": INFLOW_LABEL,
                "label_row_offset": 4,
                "value_row_offset": 5,
                "value_range": (0, 1),
            },
        ],
    },
    "R6": {
        "name": "別添４③（構想区域の病床数等の状況）",
        "path_in_repo": "R6/別添４③（構想区域の病床数等の状況）.xlsx",
        "sheet_name": "構想区域別必要量との比較",
        # xlsx単体の直リンクではなく、令和6年度版一括DL zip に同梱されている
        # (doc/DATA_SOURCES.md「R6/」節参照。zip自身のSHA-256は
        # `tools/verify_r6_bundle.py` で検証済み)。
        "download_url": "https://www.mhlw.go.jp/content/10800000/001723128.zip",
        "source_note": "令和6年度版一括DL zip に同梱(xlsx単体の直リンクではない)",
        "fiscal_year": "令和6年度",
        "acquired_date": "2026-08-05",
        "flow_items": [
            {
                "field": "net_flow_rate",
                "col": 17,
                "label": NET_FLOW_LABEL,
                "label_row_offset": 4,
                "value_row_offset": 5,
                "value_range": (-1, 1),
            },
        ],
    },
}

# 出力行の順序(build_and_write()で常にこの順に固定する。--sourceの指定順に
# 依らない。理由はモジュールdocstring参照)。
SOURCE_ORDER = ["R7", "R6"]

# area_basic.csv が持つ流出入関連フィールド全種類。年度により算出される
# フィールドが異なる(R7=outflow_rate/inflow_rate、R6=net_flow_rate)ため、
# 出力行は常に全フィールドを持たせ、算出されないものはNone(CSV上は空欄)にする。
ALL_FLOW_FIELDS = ["outflow_rate", "inflow_rate", "net_flow_rate"]

# R6原典の実績セル欠測(モジュールdocstring「⚠ 実績セルの欠測」参照)。
# (published_fy, area_code, bed_function, series, year) のタプルで、
# パース中に検出した空セルの集合とここが完全一致することを検証する
# (新しい欠測が増えたらここで気付ける設計。合計から逆算して埋めることはしない)。
EXPECTED_MISSING_BEDS = {
    ("R6", "0102", "高度急性期", "実績", 2015),
}

SOURCE_PAGE_URL = "https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/0000080850_00014.html"
LICENSE_NOTE = (
    "厚生労働省ホームページ利用規約（政府標準利用規約準拠） "
    "https://www.mhlw.go.jp/chosakuken/index.html"
)
CAVEAT = (
    "病床機能報告の集計結果(実績)と将来の病床数の必要量(必要数)は計算方法が"
    "異なることから、単純に比較するのではなく、詳細な分析や検討を行った上で"
    "地域医療構想調整会議で協議を行うことが重要(原典C2注記より)。可視化で"
    "これらを併記する際は、この点を注記として必ず表示すること。また年度ごと"
    "に報告率が異なることにも留意する。"
)

BLOCK_TOP0 = 3   # 最初のブロック(block.index=0, 構想区域コード101)の先頭行
BLOCK_SIZE = 15  # 1ブロックあたりの行数
NUM_BLOCKS = 339  # 構想区域数

BED_FUNCTIONS = ["合計", "高度急性期", "急性期", "回復期", "慢性期"]

PREF_CODE_DESC = "都道府県コード(ゼロ埋め2桁の文字列、01=北海道…47=沖縄県、原典の都道府県コード順)"
PREF_NAME_DESC = "都道府県名"
AREA_CODE_DESC = (
    "構想区域(二次医療圏)コード(ゼロ埋め4桁の文字列)。上2桁が都道府県コードと一致する"
)
AREA_NAME_DESC = "構想区域名"
PUBLISHED_FY_DESC = (
    "公表年度を表す識別子。'R7'=令和7年度公表分、'R6'=令和6年度公表分。"
    "行の出典はmeta.jsonのsource配列をpublished_fyで引くこと"
)

FIELDS_BEDS = {
    "published_fy": PUBLISHED_FY_DESC,
    "pref_code": PREF_CODE_DESC,
    "pref_name": PREF_NAME_DESC,
    "area_code": AREA_CODE_DESC,
    "area_name": AREA_NAME_DESC,
    "bed_function": "病床機能区分(合計/高度急性期/急性期/回復期/慢性期)。合計は他4区分の和",
    "series": "系列。実績=病床機能報告の報告値、見込量=直近年からの見込み、必要数=2025年の必要病床数",
    "year": "対象年(西暦)。実績・見込量・必要数それぞれの対象年は公表年度により異なる(下記caveat参照)",
    "beds": "病床数(床)",
}
FIELDS_REPORT_RATE = {
    "published_fy": PUBLISHED_FY_DESC,
    "pref_code": PREF_CODE_DESC,
    "pref_name": PREF_NAME_DESC,
    "area_code": AREA_CODE_DESC,
    "area_name": AREA_NAME_DESC,
    "year": "報告率の対象年(実績年のみ)",
    "report_rate": "病床機能報告の報告率(原典値をそのまま。丸めていない0〜1の割合)",
}
FIELDS_BASIC = {
    "published_fy": PUBLISHED_FY_DESC,
    "pref_code": PREF_CODE_DESC,
    "pref_name": PREF_NAME_DESC,
    "area_code": AREA_CODE_DESC,
    "area_name": AREA_NAME_DESC,
    "population_2020": "2020年国勢調査人口(人単位の整数)。原典の万人値を10000倍し四捨五入",
    "population_2020_source_value": "原典の人口値(万人単位、丸めなし)",
    "population_2020_source_unit": "population_2020_source_value の単位(万人)",
    "area_2020_km2": "2020年面積(km2)。原典の浮動小数点誤差を除くため小数2桁に丸め",
    "outflow_rate": (
        "推計流出患者割合(0〜1)。R7のみ算出(R6行は常に空)。原典が数値の場合のみ値を持つ。"
        "三重県の8区域では原典が'XXX'のため空(outflow_rate_source_value参照)"
    ),
    "outflow_rate_source_value": "推計流出患者割合の原典値(数値または'XXX'をそのまま。R6行は常に空)",
    "inflow_rate": (
        "推計流入患者割合(0〜1)。R7のみ算出(R6行は常に空)。原典が数値の場合のみ値を持つ。"
        "三重県の8区域では原典が'XXX'のため空(inflow_rate_source_value参照)"
    ),
    "inflow_rate_source_value": "推計流入患者割合の原典値(数値または'XXX'をそのまま。R6行は常に空)",
    "net_flow_rate": (
        "一般病床患者流出入(-1〜1、負値を取りうる)。R6のみ算出(R7行は常に空)。"
        "原典が数値の場合のみ値を持つ。三重県の8区域では原典が'XXX'のため空"
        "(net_flow_rate_source_value参照)。**outflow_rate/inflow_rate(R7)とは別概念であり、"
        "並べて比較・可視化してはならない**(area_basic_r6_net_flow_rate_different_concept参照)"
    ),
    "net_flow_rate_source_value": "一般病床患者流出入の原典値(数値または'XXX'をそのまま。R7行は常に空)",
}

STEPS_COMMON = [
    "openpyxl(data_only=True)でシートを開き、3行目から15行ずつの339ブロック(構想区域ごと)を走査",
    "サブヘッダー行(実績/見込量/必要数の列見出し)の文字列から列を解決(公表年度により列位置が異なるためハードコードしない)",
    "全339ブロックのサブヘッダー行が先頭ブロックと完全一致することを検証(不一致ならレイアウト変更とみなし中断)",
]

KNOWN_ISSUE_BEDS_2024_DUP = {
    "id": "area_beds_2024_actual_duplicated_as_2025",
    "scope": {"csv": "area_beds.csv", "published_fy": "R7", "series": "実績", "year": 2024},
    "summary": (
        "R7公表分の構想区域別「2024実績」列が「2025実績」列と全セルで完全に同一な値に"
        "なっている(原典のR7/001723349.xlsxそのものの問題)"
    ),
    "evidence": [
        "R7: 2024実績列と2025実績列が339区域×5機能=1695セル全てで一致",
        "R7: 構想区域の実績を都道府県コードで集計し都道府県版(001722915.xlsx)の"
        "2024年実績と突合すると、2585キー(47都道府県×5機能×11系列相当)中"
        "230キーが2024年に集中して不一致になる",
        "R6公表分(別添４③)の同一列と突合すると339区域×5機能=1695セル中1281セルで"
        "不一致。一方2015〜2023実績と2025必要数は、area_beds_r6_2015_actual_missing_minamihiyama"
        "(南檜山・高度急性期の2015実績1件がR6原典で空欄)を除き1695セル全てで一致する",
        "R6の区域別2024実績を都道府県へ集計すると①都道府県版(別添４②)と235/235キーで"
        "完全一致するが、R7の同じ集計は230/235キーで不一致になる。したがってR6側の"
        "2024実績が健全な値である",
    ],
    "action": (
        "値は原典どおり出力している(勝手に補正しない)。可視化では構想区域"
        "レベルの2024年実績を用いないこと(2025年実績または都道府県レベルの"
        "2024年実績を使うこと)。R6公表分(published_fy=='R6')の2024実績は健全なので、"
        "区域レベルの2024年実績が必要な場合はそちらを使うこと"
    ),
}

KNOWN_ISSUE_BEDS_R6_MISSING_MINAMIHIYAMA = {
    "id": "area_beds_r6_2015_actual_missing_minamihiyama",
    "scope": {
        "csv": "area_beds.csv",
        "published_fy": "R6",
        "area_code": "0102",
        "bed_function": "高度急性期",
        "series": "実績",
        "year": 2015,
    },
    "summary": (
        "R6公表分(別添４③)の区域コード0102「南檜山」の2015実績・高度急性期が"
        "原典で空欄になっている"
    ),
    "evidence": [
        "同ブロックの2015実績は 合計=399 / 高度急性期=空 / 急性期=202 / 回復期=0 / "
        "慢性期=197(空欄以外の4値は原典どおり)",
        "原典の派生比率列(2015年に対する比等)もこのセルに対応する箇所は数値ではなく"
        "'-'表記になっている(レイアウト崩れではなく原典データそのものの欠測)",
    ],
    "action": (
        "beds列は空欄のまま出力し、合計から逆算して埋めることはしない"
        "(位置の推測はしない、doc/REQUIREMENTS.md §4.3と同じ規律)。空欄は"
        "モジュール定数EXPECTED_MISSING_BEDSと完全一致することを検証しているため、"
        "新しい欠測が増えた場合はLayoutMismatchErrorで検知される"
    ),
}

KNOWN_ISSUE_REPORT_RATE_R6_R7_DIFF_2024 = {
    "id": "area_bed_report_rate_2024_differs_between_r6_r7",
    "scope": {"csv": "area_bed_report_rate.csv", "year": 2024},
    "summary": (
        "構想区域別の病床機能報告の報告率(2024年)が、339区域中105区域でR6公表分と"
        "R7公表分とで異なる値になっている(2015〜2023年は全区域で一致)"
    ),
    "evidence": [
        "R6(published_fy=='R6')とR7(published_fy=='R7')の同一area_code・year=2024の"
        "報告率を突合すると339区域中105区域で不一致",
        "2015〜2023年は339区域全てでR6・R7の値が一致",
    ],
    "action": (
        "値はいずれも原典どおり出力している(勝手に補正しない)。年度をまたいだ"
        "報告率の比較・可視化ではpublished_fyを明示すること"
    ),
}

KNOWN_ISSUE_BASIC_XXX_MIE = {
    "id": "area_basic_outflow_inflow_rate_xxx_mie",
    "scope": {
        "csv": "area_basic.csv",
        "area_codes": ["2405", "2406", "2407", "2408", "2409", "2410", "2411", "2412"],
    },
    "summary": (
        "三重県の8構想区域(2405〜2412)のみ、R7の推計流出患者割合・推計流入患者割合と"
        "R6の一般病床患者流出入が、いずれも原典で文字列'XXX'(未算出)になっている"
    ),
    "evidence": [
        "R7: 他331区域はoutflow_rate/inflow_rateが0〜1の数値だが、三重県の8区域のみ'XXX'",
        "R6: 同じ8区域でnet_flow_rateも'XXX'",
        "三重県は令和2年度時点の二次医療圏4圏域から令和7年度の構想区域では"
        "8区域へ細分化されており、細分化後の区域単位での流出入率が算出されていない",
    ],
    "action": (
        "outflow_rate/inflow_rate/net_flow_rateはいずれも空(欠測)とし、原典値は"
        "各*_source_valueに'XXX'のまま保持している"
    ),
}

KNOWN_ISSUE_BASIC_R6_NET_FLOW_DIFFERENT_CONCEPT = {
    "id": "area_basic_r6_net_flow_rate_different_concept",
    "scope": {
        "csv": "area_basic.csv",
        "published_fy": "R6",
        "columns": ["net_flow_rate", "net_flow_rate_source_value"],
    },
    "summary": (
        "R6公表分(別添４③)のQ列(17)「（一般病床患者流出入）」は、R7の推計流出患者割合・"
        "推計流入患者割合(2つの別項目)とは別概念であり、値域も-0.893〜0.434と負値を"
        "含む(R7は0〜1)"
    ),
    "evidence": [
        "R7は推計流出患者割合・推計流入患者割合の2項目を別々の値として持つが、R6は"
        "「（一般病床患者流出入）」という単一の値のみを持つ",
        "R6のnet_flow_rateの実測値域は-0.893〜0.434で負の値を取る。"
        "R7のoutflow_rate/inflow_rateは0〜1の値域",
    ],
    "action": "outflow_rate/inflow_rate(R7)とnet_flow_rate(R6)は別概念のため、並べて比較・可視化してはならない",
}

KNOWN_ISSUES = [
    KNOWN_ISSUE_BEDS_2024_DUP,
    KNOWN_ISSUE_BEDS_R6_MISSING_MINAMIHIYAMA,
    KNOWN_ISSUE_REPORT_RATE_R6_R7_DIFF_2024,
    KNOWN_ISSUE_BASIC_XXX_MIE,
    KNOWN_ISSUE_BASIC_R6_NET_FLOW_DIFFERENT_CONCEPT,
]


@dataclass
class ParseResult:
    published_fy: str
    title: str
    notes: list
    beds_rows: list = field(default_factory=list)
    report_rate_rows: list = field(default_factory=list)
    basic_rows: list = field(default_factory=list)


def _convert_man_to_person(value_man) -> int:
    """人口(万人)を人単位の整数へ変換する(万人 -> 人、四捨五入)。

    検証の意図は「原典の万人値が本当に(小数4桁までの)人単位の整数か」の
    確認であり、`round()` の性質上ほぼ必ず真になってしまう `>= 0.5` という
    閾値では実質的に検証にならない。許容誤差は `1e-6` とする。
    """
    scaled = float(value_man) * 10000
    person = round(scaled)
    if abs(scaled - person) >= 1e-6:
        raise ValueError(f"人口換算の丸め誤差が大きすぎます: {value_man} -> {scaled}")
    return person


def _split_rate(value, *, value_range):
    """流出入関連セルの値を (数値 or None, 原典値) に分解する。

    原典が数値なら `value_range`(min, max)の範囲内であることを確認して
    そのまま採用する(年度により許容値域が異なる。R7の推計流出/流入患者
    割合は0〜1、R6の一般病床患者流出入は-1〜1で負値を取りうる)。三重県の
    8区域のように非数値(文字列'XXX')の場合は数値側をNoneにし、原典値は
    そのまま保持する(CLAUDE.mdの流儀: 値を勝手に補正・欠落させず、原典値を
    併記する)。
    """
    if isinstance(value, bool):
        raise LayoutMismatchError(f"流出入関連セルの値が不正です: {value!r}")
    if isinstance(value, (int, float)):
        lo, hi = value_range
        if not (lo <= value <= hi):
            raise LayoutMismatchError(
                f"流出入関連セルの値が範囲外です({lo}〜{hi}): {value!r}"
            )
        return value, value
    return None, value


def parse_sheet(ws, published_fy: str) -> ParseResult:
    """帳票シートを339ブロック走査し、tidy行を組み立てる。

    `ws` は openpyxl の Worksheet(`data_only=True` で開いたもの)。
    """
    title = ws["C1"].value
    c2 = ws["C2"].value
    notes = c2.split("\n") if c2 else []

    max_col = ws.max_column
    # A列(ブロック番号)と最終列(ブロックごとの通し番号ラベル)はブロックにより
    # 値が変わって当然のため、サブヘッダー比較の対象範囲から除外する。
    header_col_start = 2
    header_col_end = max_col - 1
    result = ParseResult(published_fy=published_fy, title=title, notes=notes)

    # 構想区域コードは1始まりの連番1..339。`assert_repeated_header` で
    # 複数回走査するため、ジェネレータのままではなくリスト化しておく。
    blocks = list(
        iter_fixed_blocks(
            ws,
            first_row=BLOCK_TOP0,
            block_size=BLOCK_SIZE,
            count=NUM_BLOCKS,
            first_number=1,
        )
    )

    reference_header = assert_repeated_header(
        ws, blocks, row_offset=8, col_start=header_col_start, col_end=header_col_end
    )
    col_map = resolve_columns(
        reference_header, col_start=header_col_start, classify=classify_bed_column
    )

    seen_area_codes = set()
    pref_name_by_code = {}
    seen_missing_beds = set()

    for block in blocks:
        top = block.top_row

        pref_label_row = top + 2
        expect(
            ws.cell(row=pref_label_row, column=4).value,
            "都道府県",
            f"ブロック{block.index}: 都道府県ラベル行(D列)",
        )
        pref_code_num = ws.cell(row=pref_label_row, column=6).value
        pref_name = ws.cell(row=pref_label_row, column=8).value
        pref_code = normalize_pref_code(pref_code_num)

        area_label_row = top + 3
        expect(
            ws.cell(row=area_label_row, column=4).value,
            "構想区域",
            f"ブロック{block.index}: 構想区域ラベル行(D列)",
        )
        area_code_num = ws.cell(row=area_label_row, column=6).value
        area_name = ws.cell(row=area_label_row, column=8).value
        area_code = normalize_area_code(area_code_num)

        # 構想区域コードの上2桁は都道府県コードと一致する規則になっている
        # (AREA_CODE_DESC参照)ため、その規則が原典でも保たれていることを検証する。
        expect(
            area_code[:2],
            pref_code,
            f"ブロック{block.index}: 構想区域コード{area_code}の上2桁が都道府県コード{pref_code}と不一致",
        )
        if area_code in seen_area_codes:
            raise LayoutMismatchError(
                f"ブロック{block.index}: 構想区域コード{area_code}が重複しています"
            )
        seen_area_codes.add(area_code)

        if pref_code in pref_name_by_code and pref_name_by_code[pref_code] != pref_name:
            raise LayoutMismatchError(
                f"ブロック{block.index}: 都道府県コード{pref_code}の都道府県名が"
                f"ブロック間で一致しません({pref_name_by_code[pref_code]!r} != {pref_name!r})"
            )
        pref_name_by_code[pref_code] = pref_name

        pop_row = top + 4
        expect(
            ws.cell(row=pop_row, column=4).value,
            "2020国勢調査人口",
            f"ブロック{block.index}: 人口ラベル行(D列)",
        )
        population_source_value = ws.cell(row=pop_row, column=6).value

        area_row = top + 5
        expect(
            ws.cell(row=area_row, column=4).value,
            "2020面積",
            f"ブロック{block.index}: 面積ラベル行(D列)",
        )
        area_source_value = ws.cell(row=area_row, column=6).value

        # 流出入関連の列は位置でハードコードしているため、値を読む前に
        # ラベルが想定どおりの行位置にあることを検証する(ラベルが動いて
        # いたらここで必ず失敗させ、誤った列を静かに読むことを防ぐ)。
        # 年度により項目数・列位置・値域が異なるため `SOURCES[published_fy]
        # ["flow_items"]` から解決する(モジュールdocstring「⚠ 流出入項目の
        # 年度差」参照)。
        flow_values = {f: None for f in ALL_FLOW_FIELDS}
        for key in ALL_FLOW_FIELDS:
            flow_values[f"{key}_source_value"] = None
        for item in SOURCES[published_fy]["flow_items"]:
            label_row = top + item["label_row_offset"]
            value_row = top + item["value_row_offset"]
            expect(
                ws.cell(row=label_row, column=item["col"]).value,
                item["label"],
                f"ブロック{block.index}: {item['field']}ラベル行(列{item['col']})",
            )
            rate, source_value = _split_rate(
                ws.cell(row=value_row, column=item["col"]).value,
                value_range=item["value_range"],
            )
            flow_values[item["field"]] = rate
            flow_values[f"{item['field']}_source_value"] = source_value

        section_row = top + 6
        expect(
            ws.cell(row=section_row, column=3).value,
            "○病床数の状況",
            f"ブロック{block.index}: 病床数セクション見出し行(C列)",
        )

        # (series, year) ごとに5機能分の病床数を集め、後段で「合計」==4機能の和を検証する。
        block_beds_by_key = {}
        for i, bed_function in enumerate(BED_FUNCTIONS):
            row = top + 9 + i
            expect(
                ws.cell(row=row, column=5).value,
                bed_function,
                f"ブロック{block.index}: 病床機能ラベル行(E列, {bed_function})",
            )
            # A列のブロック番号は反復インデックス(1始まり)であり構想区域コードそのもの
            # ではないため、代わりにB列(病床機能行)とF列(構想区域ラベル行、area_code)の
            # 構想区域コードが一致することを検証する(本来常に一致するはずの結合キー)。
            expect(
                normalize_area_code(ws.cell(row=row, column=2).value),
                area_code,
                f"ブロック{block.index}: 病床機能行(B列)の構想区域コードが不一致",
            )
            for col, (series, year) in col_map.items():
                value = ws.cell(row=row, column=col).value
                if value is None:
                    # 病床数セルの空欄は原典側の欠測(モジュールdocstring
                    # 「⚠ 実績セルの欠測」参照)。EXPECTED_MISSING_BEDSに
                    # 列挙されたものだけを許容し、未知の空欄は即座に中断する
                    # (取りこぼし防止。合計から逆算して埋めることはしない)。
                    missing_key = (published_fy, area_code, bed_function, series, year)
                    if missing_key not in EXPECTED_MISSING_BEDS:
                        raise LayoutMismatchError(
                            f"ブロック{block.index} 行{row} 列{col}: "
                            f"病床数セルが空です(未知の欠測): {missing_key}"
                        )
                    seen_missing_beds.add(missing_key)
                    beds = None
                else:
                    beds = expect_int(value, block=block.index, row=row, col=col)
                result.beds_rows.append(
                    {
                        "published_fy": published_fy,
                        "pref_code": pref_code,
                        "pref_name": pref_name,
                        "area_code": area_code,
                        "area_name": area_name,
                        "bed_function": bed_function,
                        "series": series,
                        "year": year,
                        "beds": beds,
                    }
                )
                block_beds_by_key.setdefault((series, year), {})[bed_function] = beds

        # 「合計」が他4機能の和と一致することを検証する(静かに値がずれるのを防ぐ)。
        # 欠測(None)を含む(series, year)は等式検証できないため、代わりに
        # 「合計 >= 欠測を除く機能の和」であることのみ検証する(欠測値を
        # 合計から逆算して埋めることは絶対にしない)。
        for (series, year), funcs in block_beds_by_key.items():
            parts = [funcs[f] for f in ("高度急性期", "急性期", "回復期", "慢性期")]
            if funcs["合計"] is None or any(p is None for p in parts):
                known_sum = sum(p for p in parts if p is not None)
                if funcs["合計"] is not None and funcs["合計"] < known_sum:
                    raise LayoutMismatchError(
                        f"ブロック{block.index} {series}{year}年: 欠測を含むが「合計」が"
                        f"欠測を除く機能の和({known_sum})を下回っています(不整合)"
                    )
            else:
                parts_sum = sum(parts)
                expect(
                    funcs["合計"],
                    parts_sum,
                    f"ブロック{block.index} {series}{year}年: 「合計」が4機能の和と不一致",
                )

        rate_row = top + 9 + len(BED_FUNCTIONS)
        expect(
            ws.cell(row=rate_row, column=5).value,
            "（報告率）",
            f"ブロック{block.index}: 報告率ラベル行(E列)",
        )
        # 病床機能行と同様、B列(報告率行)とF列(area_code)の構想区域コードが
        # 一致することを検証する(本来常に一致するはずの結合キー)。
        expect(
            normalize_area_code(ws.cell(row=rate_row, column=2).value),
            area_code,
            f"ブロック{block.index}: 報告率行(B列)の構想区域コードが不一致",
        )
        for col, (series, year) in col_map.items():
            if series != "実績":
                continue
            value = ws.cell(row=rate_row, column=col).value
            if value is None:
                continue
            if isinstance(value, bool) or not isinstance(value, (int, float)):
                raise LayoutMismatchError(
                    f"ブロック{block.index} 行{rate_row} 列{col}: "
                    f"報告率セルの値が数値ではありません: {value!r}"
                )
            if not (0 <= value <= 1):
                raise LayoutMismatchError(
                    f"ブロック{block.index} 行{rate_row} 列{col}: "
                    f"報告率セルの値が0〜1の範囲外です: {value!r}"
                )
            result.report_rate_rows.append(
                {
                    "published_fy": published_fy,
                    "pref_code": pref_code,
                    "pref_name": pref_name,
                    "area_code": area_code,
                    "area_name": area_name,
                    "year": year,
                    "report_rate": value,
                }
            )

        population_2020 = _convert_man_to_person(population_source_value)
        area_2020_km2 = round(float(area_source_value), 2)
        result.basic_rows.append(
            {
                "published_fy": published_fy,
                "pref_code": pref_code,
                "pref_name": pref_name,
                "area_code": area_code,
                "area_name": area_name,
                "population_2020": population_2020,
                "population_2020_source_value": population_source_value,
                "population_2020_source_unit": "万人",
                "area_2020_km2": area_2020_km2,
                "outflow_rate": flow_values["outflow_rate"],
                "outflow_rate_source_value": flow_values["outflow_rate_source_value"],
                "inflow_rate": flow_values["inflow_rate"],
                "inflow_rate_source_value": flow_values["inflow_rate_source_value"],
                "net_flow_rate": flow_values["net_flow_rate"],
                "net_flow_rate_source_value": flow_values["net_flow_rate_source_value"],
            }
        )

    # 検出した欠測セルの集合がEXPECTED_MISSING_BEDSの当該年度分と完全一致
    # することを検証する。不一致(想定していた欠測が実は埋まっている、または
    # 想定外の欠測が別途あった)ならレイアウト変更とみなして中断する。
    expected_for_this_fy = {k for k in EXPECTED_MISSING_BEDS if k[0] == published_fy}
    if seen_missing_beds != expected_for_this_fy:
        raise LayoutMismatchError(
            f"病床数セルの欠測がEXPECTED_MISSING_BEDSと一致しません(published_fy={published_fy!r})\n"
            f"  期待: {sorted(expected_for_this_fy)}\n"
            f"  実際: {sorted(seen_missing_beds)}"
        )

    return result


def load_sheet(source_key: str):
    """`SOURCES[source_key]` の設定に従い対象xlsxのシートを開く。

    元データのSHA-256を `SHA256SUMS` と照合したうえで開く。
    戻り値: (worksheet, source_config, source_sha256)
    """
    cfg = SOURCES[source_key]
    source_sha256 = verify_source(cfg["path_in_repo"])
    print(f"[ok] 生データ検証: {cfg['path_in_repo']} = {source_sha256[:16]}...")
    xlsx_path = REPO_ROOT / cfg["path_in_repo"]
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[cfg["sheet_name"]]
    return ws, cfg, source_sha256


def _rows_to_tuples(rows, header):
    return [tuple(r[h] for h in header) for r in rows]


ROW_ORDER_STEP = (
    "複数の公表年度(--source)を対象にする場合、行はR7を先に・R6を後に固定した"
    "順で連結する(R7部分をバイト単位で不変に保ち、差分でR6の追加だけが読めるようにするため)"
)


def build_and_write(source_keys, out_dir: Path) -> dict:
    """指定ソース(複数可)をパースし、3つのCSV+meta.jsonを `out_dir` へ出力する。

    `source_keys` は `SOURCES` のキー("R7"/"R6")の列挙(順不同で渡してよい)。
    出力行は指定順に関わらず常に `SOURCE_ORDER`(R7→R6)の順に固定する。

    書き出したCSVパスの辞書({"beds": ..., "report_rate": ..., "basic": ...})
    を返す(再現性テスト等での再利用を想定)。
    """
    out_dir = Path(out_dir)
    ordered_keys = [k for k in SOURCE_ORDER if k in source_keys]
    if not ordered_keys:
        raise ValueError(f"source_keys が空、または未知のキーを含みます: {source_keys!r}")

    results = []
    sources = []
    for key in ordered_keys:
        ws, cfg, source_sha256 = load_sheet(key)
        result = parse_sheet(ws, published_fy=key)
        print(
            f"[ok] パース完了({key}): beds={len(result.beds_rows)}行 "
            f"report_rate={len(result.report_rate_rows)}行 basic={len(result.basic_rows)}行"
        )
        results.append(result)

        source_entry = {
            "published_fy": key,
            "name": cfg["name"],
            "publisher": "厚生労働省",
            "url": cfg["download_url"],
        }
        if cfg.get("source_note"):
            source_entry["source_note"] = cfg["source_note"]
        source_entry.update(
            {
                "page_url": SOURCE_PAGE_URL,
                "fiscal_year": cfg["fiscal_year"],
                "source_file": cfg["path_in_repo"],
                "source_sha256": source_sha256,
                "source_sheet": cfg["sheet_name"],
                "acquired_date": cfg["acquired_date"],
                "license": LICENSE_NOTE,
                "original_title": result.title,
                "original_notes": result.notes,
            }
        )
        sources.append(source_entry)

    beds_rows = [row for result in results for row in result.beds_rows]
    report_rate_rows = [row for result in results for row in result.report_rate_rows]
    basic_rows = [row for result in results for row in result.basic_rows]

    today = datetime.date.today().isoformat()

    beds_header = [
        "published_fy",
        "pref_code",
        "pref_name",
        "area_code",
        "area_name",
        "bed_function",
        "series",
        "year",
        "beds",
    ]
    beds_csv, _ = write_csv_with_meta(
        out_dir / "area_beds.csv",
        beds_header,
        _rows_to_tuples(beds_rows, beds_header),
        title="構想区域別 病床数(実績/見込量/必要数)",
        source=sources,
        processing={
            "script": "tools/parse_area_beds.py",
            "date": today,
            "steps": STEPS_COMMON
            + [
                "派生比率列(2025年必要数に対する比等)は再計算可能なため出力対象から除外",
                "病床数セルの空欄(R6のみ)はEXPECTED_MISSING_BEDSと突合したうえでbedsを空欄のまま出力(合計から逆算して埋めない)",
                ROW_ORDER_STEP,
            ],
            "caveat": CAVEAT,
        },
        fields=FIELDS_BEDS,
        known_issues=[KNOWN_ISSUE_BEDS_2024_DUP, KNOWN_ISSUE_BEDS_R6_MISSING_MINAMIHIYAMA],
    )
    print(f"[ok] 出力: {beds_csv} ({len(beds_rows)}行)")

    rate_header = [
        "published_fy",
        "pref_code",
        "pref_name",
        "area_code",
        "area_name",
        "year",
        "report_rate",
    ]
    rate_csv, _ = write_csv_with_meta(
        out_dir / "area_bed_report_rate.csv",
        rate_header,
        _rows_to_tuples(report_rate_rows, rate_header),
        title="構想区域別 病床機能報告の報告率",
        source=sources,
        processing={
            "script": "tools/parse_area_beds.py",
            "date": today,
            "steps": STEPS_COMMON
            + [
                "各ブロックの（報告率）行から実績年の値のみを抽出(見込量・必要数列は報告率を持たない)",
                ROW_ORDER_STEP,
            ],
            "caveat": CAVEAT,
        },
        fields=FIELDS_REPORT_RATE,
        known_issues=[KNOWN_ISSUE_REPORT_RATE_R6_R7_DIFF_2024],
    )
    print(f"[ok] 出力: {rate_csv} ({len(report_rate_rows)}行)")

    basic_header = [
        "published_fy",
        "pref_code",
        "pref_name",
        "area_code",
        "area_name",
        "population_2020",
        "population_2020_source_value",
        "population_2020_source_unit",
        "area_2020_km2",
        "outflow_rate",
        "outflow_rate_source_value",
        "inflow_rate",
        "inflow_rate_source_value",
        "net_flow_rate",
        "net_flow_rate_source_value",
    ]
    basic_csv, _ = write_csv_with_meta(
        out_dir / "area_basic.csv",
        basic_header,
        _rows_to_tuples(basic_rows, basic_header),
        title="構想区域別 基礎情報(2020年人口・面積・流出入関連)",
        source=sources,
        processing={
            "script": "tools/parse_area_beds.py",
            "date": today,
            "steps": STEPS_COMMON
            + [
                "人口(万人)を10000倍し四捨五入して人単位に変換(population_2020)。原典値は population_2020_source_value に保持",
                "面積の浮動小数点誤差を除くため小数2桁に丸め(area_2020_km2)",
                "流出入関連(R7=推計流出/流入患者割合、R6=一般病床患者流出入)は原典が数値の場合のみ採用し、'XXX'等の非数値は空にして原典値を*_source_valueに保持。年度により算出されない列は常に空",
                ROW_ORDER_STEP,
            ],
            "caveat": CAVEAT,
        },
        fields=FIELDS_BASIC,
        known_issues=[KNOWN_ISSUE_BASIC_XXX_MIE, KNOWN_ISSUE_BASIC_R6_NET_FLOW_DIFFERENT_CONCEPT],
    )
    print(f"[ok] 出力: {basic_csv} ({len(basic_rows)}行)")

    return {"beds": beds_csv, "report_rate": rate_csv, "basic": basic_csv}


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--source",
        choices=["R7", "R6", "all"],
        default="all",
        help="対象の公表年度データ('all'=R7+R6の両方(既定)、'R7'/'R6'=単独)",
    )
    args = ap.parse_args()

    source_keys = ["R7", "R6"] if args.source == "all" else [args.source]
    out_dir = REPO_ROOT / "data" / "processed"
    build_and_write(source_keys, out_dir)


if __name__ == "__main__":
    main()
