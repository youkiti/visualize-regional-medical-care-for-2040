# -*- coding: utf-8 -*-
"""厚労省「④構想区域の流出率及び流入率」(R7: 001723366.xlsx、別添５)の帳票Excel
をtidy CSVへ変換する。

`tools/parse_area_beds.py`(帳票のブロック走査)と同じ`tools/lib/block_report.py`
を使うが、この帳票は1ブロック=65行(他の病床帳票は15行)・339ブロックが並ぶ点が
異なる。ブロック内の区域サマリは「(1)構想区域の状況」の再掲であり、原典自身が
`data/processed/area_basic.csv`の元データ(001723349.xlsx)と独立に同じ値を
埋め込んでいるため、このパーサでは区域サマリの値をarea_basic.csvと突合する
ことでレイアウト・結合キーの取り違えを検出する。

原典の構造(2シート「流入率」「流出率」、339ブロック、実測で確認済み):
  - シートはちょうど2枚、シート名は「流入率」「流出率」(この順)。
    どちらも max_row=22035(=65行×339ブロック)・max_column=26
  - ブロック先頭行(表題行)は `1 + 65*i`。区域サマリ行(+6)を基準にして、
    それ以降は `block.top_row`(=区域サマリ行)からの相対オフセットで各行を
    解決する(表題行だけ -6)
  - ⚠ ブロック番号列は存在しない: 事前調査メモは「区域サマリ行のA列=1始まり
    のブロック番号」としていたが、339ブロック全件の実測でA列は**都道府県
    コード**(1〜47、同一都道府県の複数ブロックで同じ値が続く)であり、
    ブロック番号ではないことを確認済み(例: ブロック0・1はいずれも北海道の
    区域のためA列=1、最終ブロック338は沖縄県のためA列=47)。339ブロックの
    どの列を探しても1始まりの連番は存在しないため、`tools/lib/block_report.py`
    の `iter_fixed_blocks()`(ブロック番号列の連番検証が前提)は使わず、
    `_iter_blocks()` でブロック位置を算術的に生成する(下記「原典の構造」の
    グリッドが `ws.max_row` にちょうど一致することと、339区域の構想区域
    コードが重複なく`area_basic.csv`と完全一致することの2点で、ブロック
    位置がずれていないことを別ルートから担保する)
  - ブロック先頭行(表題行)からのオフセット:
      +0  表題行。25列目に「（北海道・南渡島）」形式で「（都道府県名・
          構想区域名）」が入る(⚠ 実測では25列目。事前調査メモは26列目と
          していたが、339ブロック全件の実測で25列目が正しいことを確認済み)
      +6  区域サマリ行。A列=都道府県コード(1〜47の数値。ブロック番号では
          ない)、B列=都道府県名、C列=構想区域名、D列=人口(万人)、
          E列=面積(km2)。A列は原典の値としてそのまま出力に採用し、
          area_basic.csvと突合する(ブロック位置がずれれば必ずどこかの
          都道府県境界で不一致になるため、_iter_blocks()が失ったブロック
          番号の連番検証を別ルートで代替する)
      +7  A列=構想区域コード(数値。ゼロ埋めされていない)
      +8  A列=「(2)流入率」/「(2)流出率」(シートに対応)
      +9  B列=「全体の流入率」/「全体の流出率」、D列=その値(0〜1の数値)
      +12 区分ヘッダー行。B列(2)=「高度急性期+急性期」、J列(10)=「包括期」、
          R列(18)=「慢性期」
      +13 サブヘッダー行。各区分の先頭列から順に「都道府県コード」「都道
          府県名」「構想区域コード」「構想区域名」(空)「流入率」(流出率
          シートでは「流出率」)
      +14〜+64 データ行(最大51行ぶんの枠。実測最大50行)
  - データ行は率の降順(違反0件を実測)。空行が現れたらそれ以降そのグループは
    全て空(gap後の再出現0件を実測)
  - 相手区域コードは非ゼロ埋め(例 `101`)のため `normalize_area_code()` を
    通す(CLAUDE.md「結合キーの罠」)

処理内容:
  1. `verify_source()` で元データのSHA-256を `SHA256SUMS` と照合(改変検知)
  2. openpyxl(`data_only=True`)で2シートを開き、シート名・各種ラベル・
     区域サマリの値(area_basic.csvとの突合)を検証しながら339ブロックを走査
  3. 相手区域(流入元/流出先)の実在・重複なし・降順であることを検証
  4. 「全体の流入率/流出率」が「高度急性期+急性期」表の自区域行の率の余事象
     (1-rate)と厳密一致することを検証(下記KNOWN_ISSUES #1の根拠そのもの)
  5. 流出率シートの慢性期で原典セルがExcelのエラー値`#VALUE!`になっている
     行(2区域)は捨てず、value_status='error'として出力する(下記KNOWN_ISSUES
     #2参照)
  6. 2つの tidy CSV + 由来メタデータ(`<csv名>.meta.json`)を `data/processed/`
     に出力する
       - patient_flow.csv: 区域×方向×区分×相手区域の流入率・流出率(long)
       - patient_flow_total.csv: 区域×方向の「全体の流入率・流出率」

原典側の欠陥は `KNOWN_ISSUES`(id/scope/summary/evidence/action)へ構造化して
記録する(`tools/parse_area_beds.py`と同じ形)。値は勝手に補正せず原典どおり
出力し、`known_issues_for()` が `scope.csv` で振り分けて各CSVのmeta.jsonへ載る。

⚠ 単位・出典の罠: この流入率・流出率はNDBから集計したものであり、
`area_basic.csv`のoutflow_rate/inflow_rate(患者調査に基づく③構想区域別の
推計流出患者割合等)とは出典・対象年が異なる別物(原典2行目注記・出典説明書
R7/001723348.pdfより)。可視化で併記する際は必ず区別すること。

⚠ 表示分の合計は1にならない: 原典の注記により「一定数以上の患者がいる
区域のみ表示」しているため、各グループ(area_code×direction×phase)の率の
合計は1にならない(実測: 流出率・包括期で最小0.5173)。

必要環境: Python 3.11+, openpyxl

使い方:
    python tools/parse_patient_flow.py
"""
import csv
import datetime
import sys
from dataclasses import dataclass, field
from pathlib import Path

if __package__ in (None, ""):
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

from tools.lib.block_report import Block, assert_repeated_header
from tools.lib.codes import normalize_area_code, normalize_pref_code
from tools.lib.layout import LayoutMismatchError, expect
from tools.lib.provenance import REPO_ROOT, verify_source, write_csv_with_meta

SOURCE_NAME = "④構想区域の流出率及び流入率（別添５）"
SOURCE_PATH_IN_REPO = "R7/001723366.xlsx"
SOURCE_DOWNLOAD_URL = "https://www.mhlw.go.jp/content/10800000/001723366.xlsx"
SOURCE_PAGE_URL = "https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/0000080850_00014.html"
SOURCE_FISCAL_YEAR = "令和7年度（2025年度）"
SOURCE_ACQUIRED_DATE = "2026-08-04"

LICENSE_NOTE = (
    "厚生労働省ホームページ利用規約（政府標準利用規約準拠） "
    "https://www.mhlw.go.jp/chosakuken/index.html"
)

PUBLISHED_FY = "R7"

SHEET_INFLOW = "流入率"
SHEET_OUTFLOW = "流出率"
SHEETS = [SHEET_INFLOW, SHEET_OUTFLOW]

BLOCK_SIZE = 65  # 1ブロックあたりの行数
NUM_BLOCKS = 339  # 構想区域数
# 最初のブロックの区域サマリ行(表題行+6)。以降の全ての行は
# block.top_row(=区域サマリ行)からの相対オフセットで解決する(表題行だけ -6)。
FIRST_ROW = 7
MAX_ROW = BLOCK_SIZE * NUM_BLOCKS  # 22035
MAX_COL = 26

# block.top_row(区域サマリ行、原典オフセット+6)からの相対オフセット。
OFFSET_TITLE = -6            # 表題行(原典オフセット+0)
OFFSET_AREA_CODE = 1         # 構想区域コード行(+7)
OFFSET_LABEL = 2             # 「(2)流入率/流出率」ラベル行(+8)
OFFSET_OVERALL = 3           # 「全体の流入率/流出率」行(+9)
OFFSET_CATEGORY_HEADER = 6   # 区分ヘッダー行(+12)
OFFSET_SUBHEADER = 7         # サブヘッダー行(+13)
OFFSET_DATA_START = 8        # データ開始行(+14)
OFFSET_DATA_END = 58         # データ終了行(+64)
DATA_ROW_COUNT = OFFSET_DATA_END - OFFSET_DATA_START + 1  # 51

# ⚠ 表題行の「（都道府県名・構想区域名）」marker の列。事前調査メモは26列目
# としていたが、339ブロック×2シート全件の実測で25列目が正しいことを確認済み
# (26列目は常にNone)。
TITLE_COL = 25

# 区分(表の3ブロック)。原典の列位置・並び順(高度急性期+急性期→包括期→慢性期)。
# 病床機能報告の4区分(高度急性期/急性期/回復期/慢性期)とは別の区切りであり、
# 「包括期」は病床側に存在しない名称であることに留意(FIELDS_FLOW参照)。
PHASES = [
    ("高度急性期+急性期", 2),
    ("包括期", 10),
    ("慢性期", 18),
]
SUBHEADER_COL_START = 2
SUBHEADER_COL_END = 23  # 慢性期グループの末尾列(開始列18+5)
SUBHEADER_ITEMS = ["都道府県コード", "都道府県名", "構想区域コード", "構想区域名", ""]

VALUE_ERROR_LITERAL = "#VALUE!"

VALUE_STATUS_OBSERVED = "observed"
VALUE_STATUS_ERROR = "error"

PREF_CODE_DESC = "都道府県コード(ゼロ埋め2桁の文字列、01=北海道…47=沖縄県、原典の都道府県コード順)"
PREF_NAME_DESC = "都道府県名"
AREA_CODE_DESC = (
    "構想区域(二次医療圏)コード(ゼロ埋め4桁の文字列)。上2桁が都道府県コードと一致する"
)
AREA_NAME_DESC = "構想区域名"
PUBLISHED_FY_DESC = "公表年度を表す識別子。'R7'=令和7年度公表分(この帳票にR6版は存在しない)"
DIRECTION_DESC = (
    "原典シート名をそのまま使う('流入率'=自区域の医療機関に入院した患者の住所地別の"
    "構成比、'流出率'=自区域に住む患者が入院した医療機関所在地別の構成比)。"
    "英字キーへの変換は表示用データ生成側の責務"
)

FIELDS_FLOW = {
    "published_fy": PUBLISHED_FY_DESC,
    "direction": DIRECTION_DESC,
    "pref_code": PREF_CODE_DESC,
    "pref_name": PREF_NAME_DESC,
    "area_code": AREA_CODE_DESC,
    "area_name": AREA_NAME_DESC,
    "phase": (
        "原典の区分ヘッダーそのまま('高度急性期+急性期'/'包括期'/'慢性期')。"
        "病床機能報告の4区分(高度急性期/急性期/回復期/慢性期)とは別の区切りであり、"
        "'包括期'は病床側に存在しない名称である点に留意"
    ),
    "rank": (
        "そのグループ(area_code×direction×phase)内の原典の行位置(1始まり)。"
        "原典は率の降順で並ぶが、同値時の順序は導出できないため原典の並びを"
        "保持する列として出す"
    ),
    "partner_pref_code": (
        "相手方(流入元/流出先)の都道府県コード(ゼロ埋め2桁)。"
        "value_status='error'の行では空"
    ),
    "partner_pref_name": "相手方の都道府県名。value_status='error'の行では空",
    "partner_area_code": (
        "相手方の構想区域コード(ゼロ埋め4桁)。自区域自身の行(partner_area_code"
        "==area_code)も原典どおりそのまま出す。value_status='error'の行では空"
    ),
    "partner_area_name": "相手方の構想区域名。value_status='error'の行では空",
    "rate": "率(0〜1)。value_status='observed'のときのみ値を持つ(それ以外は空、欠測ではない)",
    "rate_source_value": (
        "原典セルの値そのまま(area_basic.csvのoutflow_rate_source_valueと同じ流儀)。"
        "value_status='error'では文字列'#VALUE!'"
    ),
    "value_status": (
        "'observed'=数値、'error'=原典セルがExcelのエラー値'#VALUE!'"
        "(known_issuesのflow_outflow_chronic_value_error_cells参照)。"
        "欠測を真偽値1本で持たせない(CLAUDE.md)"
    ),
}

FIELDS_TOTAL = {
    "published_fy": PUBLISHED_FY_DESC,
    "pref_code": PREF_CODE_DESC,
    "pref_name": PREF_NAME_DESC,
    "area_code": AREA_CODE_DESC,
    "area_name": AREA_NAME_DESC,
    "direction": DIRECTION_DESC,
    "overall_rate": (
        "原典の「全体の流入率」/「全体の流出率」の値(0〜1)。3区分の合計では"
        "ない点に留意(known_issuesのflow_overall_rate_equals_acute_phase_"
        "complement参照)"
    ),
}

CAVEAT_FLOW = (
    "原典の注記「医療機関所在地別患者住所地及び患者住所地別医療機関所在地の割合は"
    "一定数以上の患者がいる区域のみ表示」により、各グループ(area_code×direction×"
    "phase)の率の合計は1にならない(実測: 流出率・包括期で最小0.5173)。表示分の"
    "合計と1との差は打ち切られた区域の合計であって0ではない。"
    "この流入率・流出率はNDBから集計したものであり、area_basic.csvの"
    "outflow_rate/inflow_rate(患者調査に基づく③構想区域別の推計流出患者割合・"
    "推計流入患者割合)とは出典・対象年が異なる別物(出典説明書R7/001723348.pdfに"
    "『流入率及び流出率(2024年度)※NDBにより算出』とある)。可視化で併記する際は"
    "必ず区別すること。"
)
CAVEAT_TOTAL = (
    "overall_rateは「高度急性期+急性期」表の自区域シェアの余事象であり、3区分の"
    "合計ではない(known_issuesのflow_overall_rate_equals_acute_phase_complement"
    "参照)。この流入率・流出率はNDBから集計したものであり、area_basic.csvの"
    "outflow_rate/inflow_rateとは出典・対象年が異なる別物(caveatは"
    "patient_flow.csvも参照)。"
)

# 原典データ自体が抱える既知の品質問題(値は勝手に補正せず、機械可読な形で
# 記録する)。`tools/parse_area_beds.py` の KNOWN_ISSUES と同じ形
# (id/scope/summary/evidence/action)。
KNOWN_ISSUES = [
    {
        "id": "flow_overall_rate_equals_acute_phase_complement",
        "scope": {"csv": "patient_flow_total.csv", "columns": ["overall_rate"]},
        "summary": (
            "「全体の流入率／流出率」が3区分(高度急性期+急性期/包括期/慢性期)の"
            "合計ではなく、「高度急性期+急性期」の表の自区域シェアの余事象と完全に"
            "一致しており、「全体」という語が3区分計を意味していない"
        ),
        "evidence": [
            "339区域×2シート=678件すべてで「全体の流入率/流出率」が"
            "1-(高度急性期+急性期の自区域率)と差0で一致",
            "「包括期」「慢性期」の自区域率は同じ区域でも別の値をとる"
            "(例: 0101南渡島の流入率は 高度急性期+急性期=0.9062844376965826, "
            "包括期=0.9351719625916906, 慢性期=0.9488514140062714 で、"
            "全体の流入率0.09371556230341738は 1-0.9062844376965826 と一致)",
            "D列は数式ではなく値として保存されている",
        ],
        "action": (
            "値は原典どおり出力し、可視化では「全体」の語のまま出さず、どの区分に"
            "対応する値かを注記する。パーサはこの関係が崩れたら中断する"
            "(検証14参照)"
        ),
    },
    {
        "id": "flow_outflow_chronic_value_error_cells",
        "scope": {
            "csv": "patient_flow.csv",
            "area_codes": ["1313", "4207"],
            "direction": "流出率",
            "phase": "慢性期",
        },
        "summary": (
            "流出率シートの慢性期で、2区域のデータ行が原典でExcelのエラー値"
            "'#VALUE!'になっている"
        ),
        "evidence": [
            "該当は'1313'(島しょ)・'4207'(上五島)の2件のみ",
            "どちらもそのグループの唯一の行で、相手区域コード・名称も空",
            "この2区域を含む6区域('0502'・'0508'・'1313'・'1704'・'4207'・'4209')は"
            "流入率シートの慢性期の表が0行(自区域に慢性期の入院がない)",
        ],
        "action": (
            "行は捨てず value_status='error' として出力し、率は空にする"
            "(rate_source_valueに'#VALUE!'をそのまま保持)"
        ),
    },
]


def known_issues_for(csv_name: str):
    """`KNOWN_ISSUES` のうち `scope.csv` が `csv_name` のものを返す。

    該当が無ければ `None`(空リストではなく)を返す。`write_csv_with_meta` は
    `None` のとき meta.json に `known_issues` キー自体を出力しないため、問題が
    記録されていないCSVの出力はキーごと現れない。
    """
    issues = [issue for issue in KNOWN_ISSUES if issue["scope"]["csv"] == csv_name]
    return issues or None


STEPS_COMMON = [
    "verify_source()でR7/001723366.xlsxのSHA-256をSHA256SUMSと照合",
    "openpyxl(data_only=True)でシートがちょうど2枚(流入率/流出率、この順)であることを検証",
    "1ブロック65行×339ブロック(区域サマリ行を起点)の固定グリッドで走査。この帳票には"
    "1始まりの連番を持つブロック番号列が存在しない(区域サマリ行A列は都道府県コード"
    "(1〜47)であり、ブロック番号ではない)ため、グリッドがws.max_rowにちょうど一致する"
    "ことと、339区域の構想区域コードが重複なくarea_basic.csvと完全一致することの2点で"
    "ブロック位置のずれがないことを担保する",
    "サブヘッダー行が全339ブロックで先頭ブロックと完全一致することを検証"
    "(assert_repeated_header)し、都道府県コード〜流入率/流出率の6項目"
    "(6項目目はシート名と一致)であることを検証",
    "表題行の「（都道府県名・構想区域名）」表記、区分ヘッダー(高度急性期+急性期/"
    "包括期/慢性期)、「(2)流入率/流出率」ラベル、「全体の流入率/流出率」ラベルと"
    "値(0〜1)をブロックごとに検証",
    "区域サマリ行(都道府県名・構想区域名・人口(万人)・面積(km2))がdata/processed/"
    "area_basic.csvと一致することを検証((1)構想区域の状況の再掲であることを"
    "利用した独立検証)",
    "相手区域(都道府県コード・都道府県名・構想区域コード・構想区域名)が"
    "area_basic.csvの339件に含まれ、名称が一致することを検証。同一グループ内の"
    "重複・非降順も検証(observed行のみ対象)",
    "「全体の流入率/流出率」が「高度急性期+急性期」表の自区域行の率の余事象"
    "(1-rate)と厳密一致することを検証(不一致はKNOWN_ISSUESの"
    "flow_overall_rate_equals_acute_phase_complementの前提が崩れた合図として中断)",
    "原典セルがExcelのエラー値'#VALUE!'になっている行は、先頭列が'#VALUE!'かつ"
    "相手区域コード・名称がNoneという形であることを検証したうえでvalue_status="
    "'error'として出力(KNOWN_ISSUESのflow_outflow_chronic_value_error_cells参照)",
]


@dataclass
class SheetParseResult:
    direction: str
    title: str
    notes: list
    flow_rows: list = field(default_factory=list)
    total_rows: list = field(default_factory=list)
    # (area_code, pref_code, pref_name, area_name, population_source_value,
    #  area_source_value) のブロック走査順のリスト。2シート間の突合(検証13)に使う。
    block_summaries: list = field(default_factory=list)


def _expect_rate(value, *, context: str):
    """セル値が0〜1の数値(bool除く)であることを検証し、そのまま返す。"""
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise LayoutMismatchError(f"{context}: 率が数値ではありません: {value!r}")
    if not (0 <= value <= 1):
        raise LayoutMismatchError(f"{context}: 率が0〜1の範囲外です: {value!r}")
    return value


def _parse_phase_rows(ws, *, block, start_col: int, direction: str, phase: str, area_basic_ref: dict):
    """1ブロック×1区分ぶんのデータ行(offset+14〜+64)をtidy行の素材へ変換する。

    戻り値は7要素タプルのリスト:
        (partner_pref_code, partner_pref_name, partner_area_code,
         partner_area_name, rate, rate_source_value, value_status)
    観測行(value_status='observed')に対しては、相手区域の実在(area_basic.csv
    突合)・率の降順・同一グループ内の重複なしを検証する。
    Excelのエラー値'#VALUE!'の行(value_status='error')はこれらの検証対象外
    (相手区域自体が特定できないため。KNOWN_ISSUESの
    flow_outflow_chronic_value_error_cells参照)。
    空行が現れた後にデータ行が再出現しないことも検証する。
    """
    col_pref_code = start_col
    col_pref_name = start_col + 1
    col_area_code = start_col + 2
    col_area_name = start_col + 3
    col_rate = start_col + 5

    data_start_row = block.top_row + OFFSET_DATA_START
    rows = []
    prev_rate = None
    gap_started = False
    seen_partner_codes = set()

    for i in range(DATA_ROW_COUNT):
        row = data_start_row + i
        context = f"ブロック{block.index}({direction} {phase}) 行{row}"
        raw_pref_code = ws.cell(row=row, column=col_pref_code).value

        if raw_pref_code is None:
            gap_started = True
            continue

        if gap_started:
            raise LayoutMismatchError(f"{context}: 空行以降にデータ行が再出現しました")

        if raw_pref_code == VALUE_ERROR_LITERAL:
            raw_pref_name = ws.cell(row=row, column=col_pref_name).value
            raw_area_code = ws.cell(row=row, column=col_area_code).value
            raw_area_name = ws.cell(row=row, column=col_area_name).value
            raw_rate = ws.cell(row=row, column=col_rate).value
            if raw_pref_name is not None or raw_area_code is not None or raw_area_name is not None:
                raise LayoutMismatchError(
                    f"{context}: '#VALUE!'行の形が想定外です"
                    f"(相手都道府県名={raw_pref_name!r} 相手区域コード={raw_area_code!r} "
                    f"相手区域名={raw_area_name!r})"
                )
            if raw_rate != VALUE_ERROR_LITERAL:
                raise LayoutMismatchError(
                    f"{context}: '#VALUE!'行の率セルが想定外です: {raw_rate!r}"
                )
            if rows:
                raise LayoutMismatchError(f"{context}: '#VALUE!'行がグループの先頭以外に出現しました")
            rows.append((None, None, None, None, None, VALUE_ERROR_LITERAL, VALUE_STATUS_ERROR))
            # 実測では常にそのグループ唯一のデータ行であるため、以降に行が
            # 再出現したら検証11(空行以降の再出現なし)と同じ扱いで中断する。
            gap_started = True
            continue

        if isinstance(raw_pref_code, bool) or not isinstance(raw_pref_code, (int, float)):
            raise LayoutMismatchError(
                f"{context}: 相手都道府県コードが数値ではありません: {raw_pref_code!r}"
            )
        partner_pref_code = normalize_pref_code(raw_pref_code)
        partner_pref_name = ws.cell(row=row, column=col_pref_name).value
        partner_area_code = normalize_area_code(ws.cell(row=row, column=col_area_code).value)
        partner_area_name = ws.cell(row=row, column=col_area_name).value
        rate = _expect_rate(ws.cell(row=row, column=col_rate).value, context=context)

        # 相手区域コードの上2桁も相手都道府県コードと一致する規則になっている
        # (自区域と同じ規律。他パーサ全てが持つ検証)。
        expect(
            partner_area_code[:2],
            partner_pref_code,
            f"{context}: 相手区域コード{partner_area_code}の上2桁が相手都道府県コード{partner_pref_code}と不一致",
        )

        if prev_rate is not None and rate > prev_rate + 1e-12:
            raise LayoutMismatchError(
                f"{context}: 率が降順ではありません(前行={prev_rate!r} 今行={rate!r})"
            )
        prev_rate = rate

        if partner_area_code in seen_partner_codes:
            raise LayoutMismatchError(
                f"{context}: 相手区域コード{partner_area_code}が同一グループ内で重複しています"
            )
        seen_partner_codes.add(partner_area_code)

        ref = area_basic_ref.get(partner_area_code)
        if ref is None:
            raise LayoutMismatchError(
                f"{context}: 相手区域コード{partner_area_code}がarea_basic.csvに存在しません"
            )
        expect(
            partner_pref_name, ref["pref_name"], f"{context}: 相手都道府県名がarea_basic.csvと不一致"
        )
        expect(
            partner_area_name, ref["area_name"], f"{context}: 相手区域名がarea_basic.csvと不一致"
        )

        rows.append(
            (partner_pref_code, partner_pref_name, partner_area_code, partner_area_name, rate, rate, VALUE_STATUS_OBSERVED)
        )

    return rows


def _validate_subheader(reference_header, *, direction: str):
    """`assert_repeated_header()` が返す先頭ブロックのサブヘッダーの中身を検証する。

    `reference_header` は SUBHEADER_COL_START(2)〜SUBHEADER_COL_END(23)の
    正規化済み文字列タプル。3区分それぞれの先頭6列が「都道府県コード」「都道
    府県名」「構想区域コード」「構想区域名」(空)「流入率/流出率」であることを
    検証する(6項目目がシート名と一致することの検証を兼ねる)。
    """
    expected = SUBHEADER_ITEMS + [direction]
    for phase, start_col in PHASES:
        idx = start_col - SUBHEADER_COL_START
        actual = list(reference_header[idx : idx + 6])
        expect(actual, expected, f"{direction} {phase}: サブヘッダー6項目(都道府県コード〜{direction})")


def _iter_blocks():
    """339ブロックの位置(`Block`)を算術的に生成する。

    `tools/lib/block_report.py` の `iter_fixed_blocks()` はブロック番号列の
    連番検証を前提とするが、この帳票にはそれに相当する列が存在しない(区域
    サマリ行A列は都道府県コード(1〜47)であり、ブロック番号ではないことを
    実測で確認済み。モジュールdocstring「ブロック番号列は存在しない」参照)。
    そのため `Block` を直接組み立てる。ブロック位置のずれは、グリッドが
    `ws.max_row` にちょうど一致すること(`parse_sheet()` 冒頭で検証済み)と、
    339区域の構想区域コードが重複なく `area_basic.csv` と完全一致すること
    (`parse_sheet()` 末尾・`build_and_write()` で検証)の2点で担保する。
    """
    for index in range(NUM_BLOCKS):
        yield Block(index=index, number=index + 1, top_row=FIRST_ROW + BLOCK_SIZE * index)


def parse_sheet(ws, *, direction: str, area_basic_ref: dict) -> SheetParseResult:
    """1シート分(339ブロック)をtidy行へ変換する。

    `ws` は openpyxl の Worksheet(`data_only=True` で開いたもの)。
    `area_basic_ref` は area_code -> {"pref_code","pref_name","area_name",
    "population_2020_source_value","area_2020_km2"} の参照テーブル
    (`_load_area_basic_reference()` が作る。区域サマリ行の値・相手区域の
    実在検証に使う)。
    """
    expect(ws.max_row, MAX_ROW, f"{direction}: シートの最終行")
    expect(ws.max_column, MAX_COL, f"{direction}: シートの最終列")

    title = ws.cell(row=1, column=1).value
    notes_raw = ws.cell(row=2, column=1).value
    notes = notes_raw.split("\n") if notes_raw else []

    blocks = list(_iter_blocks())

    reference_header = assert_repeated_header(
        ws, blocks, row_offset=OFFSET_SUBHEADER, col_start=SUBHEADER_COL_START, col_end=SUBHEADER_COL_END
    )
    _validate_subheader(reference_header, direction=direction)

    label_text = f"(2){direction}"
    overall_label_text = f"全体の{direction}"

    result = SheetParseResult(direction=direction, title=title, notes=notes)
    seen_area_codes = set()

    for block in blocks:
        summary_row = block.top_row
        # A列=都道府県コード(1〜47の数値)。原典の値を正とし、area_basic.csvは
        # 突合相手として使う(値の供給元にしない)。この検証はブロック位置が
        # ずれた場合に都道府県境界で必ず不一致になるため、_iter_blocks()が
        # 失ったブロック番号の連番検証を別ルートで代替する意味も持つ。
        pref_code = normalize_pref_code(ws.cell(row=summary_row, column=1).value)
        pref_name = ws.cell(row=summary_row, column=2).value
        area_name = ws.cell(row=summary_row, column=3).value
        population_source_value = ws.cell(row=summary_row, column=4).value
        area_source_value = ws.cell(row=summary_row, column=5).value

        title_row = summary_row + OFFSET_TITLE
        expected_title = f"（{pref_name}・{area_name}）"
        expect(
            ws.cell(row=title_row, column=TITLE_COL).value,
            expected_title,
            f"ブロック{block.index}: 表題行({TITLE_COL}列目)",
        )

        area_code_row = summary_row + OFFSET_AREA_CODE
        area_code = normalize_area_code(ws.cell(row=area_code_row, column=1).value)
        if area_code in seen_area_codes:
            raise LayoutMismatchError(f"ブロック{block.index}: 構想区域コード{area_code}が重複しています")
        seen_area_codes.add(area_code)

        # 構想区域コードの上2桁は都道府県コードと一致する規則になっている
        # (他パーサ全て(parse_area_beds.py・parse_demand_forecast.py)と同じ規律)。
        expect(
            area_code[:2],
            pref_code,
            f"ブロック{block.index}: 構想区域コード{area_code}の上2桁が都道府県コード{pref_code}と不一致",
        )

        ref = area_basic_ref.get(area_code)
        if ref is None:
            raise LayoutMismatchError(
                f"ブロック{block.index}: 構想区域コード{area_code}がarea_basic.csvに存在しません"
            )
        expect(pref_code, ref["pref_code"], f"ブロック{block.index}: 都道府県コードがarea_basic.csvと不一致")
        expect(pref_name, ref["pref_name"], f"ブロック{block.index}: 都道府県名がarea_basic.csvと不一致")
        expect(area_name, ref["area_name"], f"ブロック{block.index}: 構想区域名がarea_basic.csvと不一致")
        expect(
            population_source_value,
            ref["population_2020_source_value"],
            f"ブロック{block.index}: 人口(万人)がarea_basic.csvと不一致",
        )
        expect(
            round(float(area_source_value), 2),
            ref["area_2020_km2"],
            f"ブロック{block.index}: 面積(km2)がarea_basic.csvと不一致",
        )

        label_row = summary_row + OFFSET_LABEL
        expect(
            ws.cell(row=label_row, column=1).value,
            label_text,
            f"ブロック{block.index}: 「(2){direction}」ラベル行(A列)",
        )

        overall_row = summary_row + OFFSET_OVERALL
        expect(
            ws.cell(row=overall_row, column=2).value,
            overall_label_text,
            f"ブロック{block.index}: 「全体の{direction}」ラベル行(B列)",
        )
        overall_value = _expect_rate(
            ws.cell(row=overall_row, column=4).value,
            context=f"ブロック{block.index}: 「全体の{direction}」(D列)",
        )

        category_header_row = summary_row + OFFSET_CATEGORY_HEADER
        for phase, start_col in PHASES:
            expect(
                ws.cell(row=category_header_row, column=start_col).value,
                phase,
                f"ブロック{block.index}: 区分ヘッダー行(列{start_col})",
            )

        rows_by_phase = {
            phase: _parse_phase_rows(
                ws, block=block, start_col=start_col, direction=direction, phase=phase, area_basic_ref=area_basic_ref
            )
            for phase, start_col in PHASES
        }

        # 検証14: 「全体のX率」が「高度急性期+急性期」表の自区域行の率の余事象
        # (1-rate)と厳密一致すること。この関係はKNOWN_ISSUESの
        # flow_overall_rate_equals_acute_phase_complementの根拠そのものなので、
        # 崩れたら記述が古くなった合図としてここで中断する(known_issueのidを
        # メッセージに含める)。
        acute_rows = rows_by_phase["高度急性期+急性期"]
        self_row = next((r for r in acute_rows if r[2] == area_code), None)
        if self_row is None:
            raise LayoutMismatchError(
                f"ブロック{block.index}: 「高度急性期+急性期」表に自区域({area_code})の行がありません"
                "(known_issue: flow_overall_rate_equals_acute_phase_complement の前提が崩れています)"
            )
        self_share = self_row[4]
        expected_overall = 1 - self_share
        if overall_value != expected_overall:
            raise LayoutMismatchError(
                f"ブロック{block.index}: 「全体の{direction}」({overall_value!r})が"
                f"「高度急性期+急性期」自区域率の余事象(1-{self_share!r}={expected_overall!r})と不一致です"
                "(known_issue: flow_overall_rate_equals_acute_phase_complement が崩れた可能性があります)"
            )

        for phase, _start_col in PHASES:
            for rank, row_data in enumerate(rows_by_phase[phase], start=1):
                (
                    partner_pref_code,
                    partner_pref_name,
                    partner_area_code,
                    partner_area_name,
                    rate,
                    rate_source_value,
                    value_status,
                ) = row_data
                result.flow_rows.append(
                    {
                        "published_fy": PUBLISHED_FY,
                        "direction": direction,
                        "pref_code": pref_code,
                        "pref_name": pref_name,
                        "area_code": area_code,
                        "area_name": area_name,
                        "phase": phase,
                        "rank": rank,
                        "partner_pref_code": partner_pref_code,
                        "partner_pref_name": partner_pref_name,
                        "partner_area_code": partner_area_code,
                        "partner_area_name": partner_area_name,
                        "rate": rate,
                        "rate_source_value": rate_source_value,
                        "value_status": value_status,
                    }
                )

        result.total_rows.append(
            {
                "published_fy": PUBLISHED_FY,
                "pref_code": pref_code,
                "pref_name": pref_name,
                "area_code": area_code,
                "area_name": area_name,
                "direction": direction,
                "overall_rate": overall_value,
            }
        )

        result.block_summaries.append(
            (area_code, pref_code, pref_name, area_name, population_source_value, area_source_value)
        )

    return result


def load_workbook():
    """R7/001723366.xlsx を開く(SHA-256照合を済ませたうえで)。

    戻り値: (workbook, source_sha256)
    """
    source_sha256 = verify_source(SOURCE_PATH_IN_REPO)
    print(f"[ok] 生データ検証: {SOURCE_PATH_IN_REPO} = {source_sha256[:16]}...")
    xlsx_path = REPO_ROOT / SOURCE_PATH_IN_REPO
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    return wb, source_sha256


def _load_area_basic_reference():
    """`data/processed/area_basic.csv` から area_code -> 参照情報 の辞書を作る。

    区域サマリ行(都道府県名・構想区域名・人口・面積)・相手区域の実在検証に使う。

    001723366.xlsx（構想区域間の患者流入率・流出率）はR7のみで公表されている
    ファイル(「データ構成」参照。R6版なし)なので、参照先も published_fy=='R7'
    の行に絞り込む。area_basic.csvはR6/R7がpublished_fyで並存するようになった
    (M9)ため678行あり、絞り込まずに読むと同じarea_codeがR7行・R6行の順で2回
    出現し、辞書代入で後勝ち(=R6の値)になってしまう(値そのものはR6/R7で同じ
    はずだが、この帳票の突合が本来意図しないR6行に依存する状態は避ける)。
    """
    path = REPO_ROOT / "data" / "processed" / "area_basic.csv"
    reference = {}
    with open(path, encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            if row["published_fy"] != "R7":
                continue
            reference[row["area_code"]] = {
                "pref_code": row["pref_code"],
                "pref_name": row["pref_name"],
                "area_name": row["area_name"],
                "population_2020_source_value": float(row["population_2020_source_value"]),
                "area_2020_km2": float(row["area_2020_km2"]),
            }
    return reference


def _rows_to_tuples(rows, header):
    return [tuple(r[h] for h in header) for r in rows]


def build_and_write(out_dir: Path) -> dict:
    """R7/001723366.xlsxをパースし、2つのCSV+meta.jsonを `out_dir` へ出力する。

    書き出したCSVパスの辞書({"flow": ..., "total": ...})を返す
    (再現性テスト等での再利用を想定)。
    """
    out_dir = Path(out_dir)
    wb, source_sha256 = load_workbook()

    expect(wb.sheetnames, SHEETS, "ワークブックのシート構成")

    area_basic_ref = _load_area_basic_reference()

    sheet_results = []
    for sheet_name in SHEETS:
        ws = wb[sheet_name]
        result = parse_sheet(ws, direction=sheet_name, area_basic_ref=area_basic_ref)
        sheet_results.append(result)
        print(
            f"[ok] パース完了: {sheet_name} flow={len(result.flow_rows)}行 "
            f"total={len(result.total_rows)}行"
        )

    inflow_result, outflow_result = sheet_results

    # 検証13: 両シートでブロック順・区域コード・サマリ行の値が完全一致すること。
    expect(
        inflow_result.block_summaries,
        outflow_result.block_summaries,
        "流入率シートと流出率シートでブロック順・区域コード・サマリ行の値が不一致",
    )
    expect(len(inflow_result.block_summaries), NUM_BLOCKS, "区域数(339であること)")
    seen_codes = {row[0] for row in inflow_result.block_summaries}
    expect(len(seen_codes), NUM_BLOCKS, "構想区域コードの重複なし")
    expect(seen_codes, set(area_basic_ref.keys()), "構想区域コードの集合がarea_basic.csvと不一致")

    today = datetime.date.today().isoformat()
    base_source = {
        "name": SOURCE_NAME,
        "publisher": "厚生労働省",
        "url": SOURCE_DOWNLOAD_URL,
        "page_url": SOURCE_PAGE_URL,
        "fiscal_year": SOURCE_FISCAL_YEAR,
        "source_file": SOURCE_PATH_IN_REPO,
        "source_sha256": source_sha256,
        "source_sheet": SHEETS,
        "acquired_date": SOURCE_ACQUIRED_DATE,
        "license": LICENSE_NOTE,
        "original_title": inflow_result.title,
        "original_notes": inflow_result.notes
        + [
            "出典説明書「構想区域別の医療機関の病床数、診療実績、医師数等／構想区域別の"
            "流入率及び流出率」(R7/001723348.pdf)が本データの一次資料"
        ],
    }

    # 出力順: area_code昇順 -> direction(流入率→流出率のシート順) -> phase(原典の
    # 並び順) -> rank昇順。各シート内では既にブロック順(area_code)→phase(原典の
    # 並び順)→rank昇順でflow_rows/total_rowsを積んでいるため、area_codeをキーに
    # 束ね直すだけでよい(`parse_demand_forecast.py` の forecast_by_area と同じ手法)。
    flow_by_area = {}
    for result in sheet_results:
        for row in result.flow_rows:
            flow_by_area.setdefault(row["area_code"], []).append(row)
    flow_rows = []
    for area_code in sorted(flow_by_area.keys()):
        flow_rows.extend(flow_by_area[area_code])

    total_by_area = {}
    for result in sheet_results:
        for row in result.total_rows:
            total_by_area.setdefault(row["area_code"], []).append(row)
    total_rows = []
    for area_code in sorted(total_by_area.keys()):
        total_rows.extend(total_by_area[area_code])

    flow_header = [
        "published_fy",
        "direction",
        "pref_code",
        "pref_name",
        "area_code",
        "area_name",
        "phase",
        "rank",
        "partner_pref_code",
        "partner_pref_name",
        "partner_area_code",
        "partner_area_name",
        "rate",
        "rate_source_value",
        "value_status",
    ]
    flow_csv, _ = write_csv_with_meta(
        out_dir / "patient_flow.csv",
        flow_header,
        _rows_to_tuples(flow_rows, flow_header),
        title="構想区域別 流入率・流出率(相手区域別、高度急性期+急性期/包括期/慢性期)",
        source=base_source,
        processing={
            "script": "tools/parse_patient_flow.py",
            "date": today,
            "steps": STEPS_COMMON
            + [
                "相手区域別の行をtidy化して出力(並び順はarea_code昇順→direction"
                "(流入率→流出率)→phase(原典の並び順)→rank昇順)。自区域自身の行"
                "(partner_area_code==area_code)も原典どおりそのまま出力する"
            ],
            "caveat": CAVEAT_FLOW,
        },
        fields=FIELDS_FLOW,
        known_issues=known_issues_for("patient_flow.csv"),
    )
    print(f"[ok] 出力: {flow_csv} ({len(flow_rows)}行)")

    total_header = [
        "published_fy",
        "pref_code",
        "pref_name",
        "area_code",
        "area_name",
        "direction",
        "overall_rate",
    ]
    total_csv, _ = write_csv_with_meta(
        out_dir / "patient_flow_total.csv",
        total_header,
        _rows_to_tuples(total_rows, total_header),
        title="構想区域別 全体の流入率・流出率",
        source=base_source,
        processing={
            "script": "tools/parse_patient_flow.py",
            "date": today,
            "steps": STEPS_COMMON
            + [
                "各ブロックの「全体の流入率/流出率」をtidy化して出力"
                "(並び順はarea_code昇順→direction(流入率→流出率))"
            ],
            "caveat": CAVEAT_TOTAL,
        },
        fields=FIELDS_TOTAL,
        known_issues=known_issues_for("patient_flow_total.csv"),
    )
    print(f"[ok] 出力: {total_csv} ({len(total_rows)}行)")

    return {"flow": flow_csv, "total": total_csv}


def main():
    out_dir = REPO_ROOT / "data" / "processed"
    build_and_write(out_dir)


if __name__ == "__main__":
    main()
