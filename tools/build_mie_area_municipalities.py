# -*- coding: utf-8 -*-
"""三重県: 構想区域(8区域) × 二次医療圏(令和2年度、4圏域) × 構成市町(29市町)の
対応表を、一次資料(三重県公式PDF)から手転記した定数をもとに構築し、機械的な
整合性検証を行った上で `data/reference/mie_area_municipalities.csv` へ出力する。

M2チャンクC(境界合成)の前提として、`tools/verify_area_join.py`(M2チャンクB)が
「未検証の推定」としていた三重県の旧4圏域→新8区域の対応を、一次資料に基づく
検証済みの事実へ格上げするためのもの。

一次資料:
    三重県医療政策課「資料４ 第８次三重県医療計画における二次医療圏の設定に
    ついて」(`mie/001092203.pdf`)の9ページ目「現行の二次医療圏・構想区域」。
    このページには4二次医療圏(北勢・中勢伊賀・南勢志摩・東紀州)と8構想区域
    (桑員・三泗・鈴亀・津・伊賀・松阪・伊勢志摩・東紀州)、および各構想区域の
    構成市町(合計29市町)の対応表がある。三重県はこの4二次医療圏と8構想区域を
    併存させており、構想区域は二次医療圏の細分(入れ子構造)である。

対応表そのもの(`MAPPING`)は、上記9ページの表からの**手転記**である。PDFの
テキスト抽出によるパースは、あの表のレイアウト(2段組・地図と表が混在)では
誤読のリスクが高いため行わない。手転記である以上、転記ミスを機械的に検出
できることが決定的に重要であり、以下の検証をすべて実装し、1件でも違反が
あれば例外で中断する:

  1. 網羅性: CSVの29市町コードの集合が、A38(令和2年度)の三重県4圏域
     (2401〜2404)の構成市区町村(`A38b_001`)の和集合と完全一致すること
     (過不足なし)。市町コードは `data/processed/iryoken2_A38-20.geojson`
     から市町名で解決する(名称で引けなかった市町があれば中断)。
  2. 一意性: 各市町コードがCSV全体でちょうど1回だけ出現すること。
  3. 入れ子の整合: 各行の `parent_iryoken2_code` が、A38がその市町を実際に
     割り当てている二次医療圏コードと一致すること(A38の `A38b_001` から
     市町→二次医療圏の逆引きを作って照合。旧圏域をまたぐ誤転記を検出できる)。
  4. 名称の一致: CSVの `muni_name` がA38の `A38b_002` の表記と完全一致すること。
  5. 医療機関所在地との突合(独立した裏付け): `R7/001723127.xlsx`(339シート)の
     三重県8シートの「(2)区域内の医療機関」表の「②所在地」列から市町名を集め、
     その市町がCSVで割り当てられている構想区域と一致することを検証する。
  6. 残存リスクの明示: 上記5で裏付けが取れない市町を列挙する。ただし「旧圏域が
     分割されない(1対1で新区域に対応する)」ケースは、医療機関裏付けが無くても
     割り当てが一意に定まるため、「旧圏域が分割されるケースで、かつ医療機関に
     よる裏付けがない市町」だけを真の残存リスクとして計算する。

検証1〜5は失敗(不一致)があれば `ValueError`/`RuntimeError` で中断する(値を
黙って通さない)。検証結果の実測値(件数・不一致リスト)は
`data/reference/mie_area_municipalities.csv.meta.json` の `verification` に
構造化して記録し、`tools/verify_area_join.py` がレポート生成時に参照できる
ようにする(同スクリプトは生Excelを読まない設計のため、生Excel(`R7/001723127.xlsx`)
を読む検証5の結果はここで確定させ、meta.jsonを経由して受け渡す)。

必要環境: Python 3.11+, openpyxl

使い方:
    PYTHONIOENCODING=utf-8 python tools/build_mie_area_municipalities.py
"""
import datetime
import sys
from collections import Counter, defaultdict
from pathlib import Path

if __package__ in (None, ""):
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import json

import openpyxl

from tools.lib.codes import normalize_area_code
from tools.lib.provenance import REPO_ROOT, verify_source, write_csv_with_meta

GEOJSON_PATH = REPO_ROOT / "data" / "processed" / "iryoken2_A38-20.geojson"
MIE_PDF_PATH_IN_REPO = "mie/001092203.pdf"
MIE_XLSX_PATH_IN_REPO = "R7/001723127.xlsx"

OUT_CSV = REPO_ROOT / "data" / "reference" / "mie_area_municipalities.csv"

PARENT_CODES = ["2401", "2402", "2403", "2404"]

MIE_PDF_SOURCE = {
    "title": "資料４ 第８次三重県医療計画における二次医療圏の設定について",
    "publisher": "三重県医療政策課",
    "url": "https://www.pref.mie.lg.jp/common/content/001092203.pdf",
    "acquired_date": "2026-08-04",
    "reference_page": "9ページ「現行の二次医療圏・構想区域」",
}

# 三重県: 構想区域(新8区域) -> 構成市町 -> 二次医療圏(令和2年度、旧4圏域)の
# 対応。`mie/001092203.pdf` 9ページ目「現行の二次医療圏・構想区域」からの
# 手転記(合計29市町)。市町コード(`muni_code`)はここには持たず、
# `data/processed/iryoken2_A38-20.geojson` の構成市区町村リストから市町名で
# 解決する(下記 `build_rows()`)。
MAPPING = [
    {
        "area_code": "2405",
        "area_name": "桑員",
        "muni_names": ["桑名市", "いなべ市", "木曽岬町", "東員町"],
        "parent_code": "2401",
        "parent_name": "北勢",
    },
    {
        "area_code": "2406",
        "area_name": "三泗",
        "muni_names": ["四日市市", "菰野町", "朝日町", "川越町"],
        "parent_code": "2401",
        "parent_name": "北勢",
    },
    {
        "area_code": "2407",
        "area_name": "鈴亀",
        "muni_names": ["鈴鹿市", "亀山市"],
        "parent_code": "2401",
        "parent_name": "北勢",
    },
    {
        "area_code": "2408",
        "area_name": "津",
        "muni_names": ["津市"],
        "parent_code": "2402",
        "parent_name": "中勢伊賀",
    },
    {
        "area_code": "2409",
        "area_name": "伊賀",
        "muni_names": ["名張市", "伊賀市"],
        "parent_code": "2402",
        "parent_name": "中勢伊賀",
    },
    {
        "area_code": "2410",
        "area_name": "松阪",
        "muni_names": ["松阪市", "多気町", "明和町", "大台町", "大紀町"],
        "parent_code": "2403",
        "parent_name": "南勢志摩",
    },
    {
        "area_code": "2411",
        "area_name": "伊勢志摩",
        "muni_names": ["伊勢市", "鳥羽市", "志摩市", "玉城町", "度会町", "南伊勢町"],
        "parent_code": "2403",
        "parent_name": "南勢志摩",
    },
    {
        "area_code": "2412",
        "area_name": "東紀州",
        "muni_names": ["尾鷲市", "熊野市", "紀北町", "御浜町", "紀宝町"],
        "parent_code": "2404",
        "parent_name": "東紀州",
    },
]

FIELDS = {
    "area_code": "構想区域コード(ゼロ埋め4桁の文字列)。三重県の8構想区域(2405桑員〜2412東紀州)",
    "area_name": "構想区域名",
    "muni_code": (
        "市区町村コード(iryoken2_A38-20.geojson の A38b_001 より、市町名で解決)。"
        "行政区域コードの文字列表記"
    ),
    "muni_name": "市区町村名(iryoken2_A38-20.geojson の A38b_002 の表記と完全一致)",
    "parent_iryoken2_code": "当該市町が属する二次医療圏(令和2年度、A38-20)のコード。三重県の旧4圏域(2401北勢〜2404東紀州)",
    "parent_iryoken2_name": "上記二次医療圏名",
}


def _read_json(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_a38_mie_groups(geojson_path: Path = GEOJSON_PATH):
    """`iryoken2_A38-20.geojson` から三重県4圏域(2401〜2404)の構成市区町村を読む。

    戻り値: (groups, geo_metadata)
      groups: {二次医療圏コード: [(市町コード, 市町名), ...]}(A38b_001/A38b_002の並び順)
      geo_metadata: geojsonファイル冒頭の `metadata` メンバー
    """
    data = _read_json(geojson_path)
    groups = {}
    for feature in data["features"]:
        props = feature["properties"]
        code = normalize_area_code(props["A38b_003"])
        if code not in PARENT_CODES:
            continue
        muni_codes = [c.strip() for c in props["A38b_001"].split(",")]
        muni_names = [n.strip() for n in props["A38b_002"].split(",")]
        if len(muni_codes) != len(muni_names):
            raise RuntimeError(
                f"iryoken2_A38-20.geojson: {code} の A38b_001/A38b_002 の要素数が一致しません "
                f"({len(muni_codes)} vs {len(muni_names)})"
            )
        groups[code] = list(zip(muni_codes, muni_names))
    missing = set(PARENT_CODES) - set(groups)
    if missing:
        raise RuntimeError(f"iryoken2_A38-20.geojson に三重県の二次医療圏コードが見つかりません: {sorted(missing)}")
    return groups, data["metadata"]


def build_global_lookup(groups: dict):
    """A38の4圏域を横断した市町名/市町コードの相互引きを作る。

    戻り値: (name_to_code, code_to_owner, code_to_name)
      name_to_code: {市町名: 市町コード}(4圏域全体で一意であることを検証)
      code_to_owner: {市町コード: 実際に属する二次医療圏コード}(検証3の逆引き元)
      code_to_name: {市町コード: A38b_002の市町名表記}(検証4に使う)
    """
    name_to_code = {}
    code_to_owner = {}
    code_to_name = {}
    for iryoken_code, munis in groups.items():
        for muni_code, muni_name in munis:
            if muni_name in name_to_code and name_to_code[muni_name] != muni_code:
                raise RuntimeError(
                    f"市町名'{muni_name}'がA38の複数の市町コードに対応しており曖昧です "
                    f"({name_to_code[muni_name]} と {muni_code})"
                )
            name_to_code[muni_name] = muni_code
            if muni_code in code_to_owner and code_to_owner[muni_code] != iryoken_code:
                raise RuntimeError(
                    f"市町コード'{muni_code}'がA38の複数の二次医療圏に重複して属しています "
                    f"({code_to_owner[muni_code]} と {iryoken_code})"
                )
            code_to_owner[muni_code] = iryoken_code
            code_to_name[muni_code] = muni_name
    return name_to_code, code_to_owner, code_to_name


def build_rows(mapping=MAPPING, name_to_code: dict = None):
    """`MAPPING`(手転記の対応表)の市町名を市町コードで解決し、CSV行(dict)のリストを返す。

    名称でA38の構成市区町村リストから解決できない市町があれば `RuntimeError` で
    中断する(スコープ指示: 「名前で引けなかった市町があれば必ず中断すること」)。
    """
    if name_to_code is None:
        raise ValueError("name_to_code は省略できません")
    rows = []
    for entry in mapping:
        for muni_name in entry["muni_names"]:
            if muni_name not in name_to_code:
                raise RuntimeError(
                    f"{entry['area_code']}{entry['area_name']}: 市町名'{muni_name}'を"
                    "iryoken2_A38-20.geojson の構成市区町村リストから解決できません"
                    "(名称の不一致、または対応表の転記ミスの可能性)"
                )
            rows.append(
                {
                    "area_code": entry["area_code"],
                    "area_name": entry["area_name"],
                    "muni_code": name_to_code[muni_name],
                    "muni_name": muni_name,
                    "parent_iryoken2_code": entry["parent_code"],
                    "parent_iryoken2_name": entry["parent_name"],
                }
            )
    return rows


# --- 検証1〜4: CSV行とA38の整合性 -------------------------------------------


def validate_coverage(rows, groups) -> dict:
    """検証1(網羅性): CSVの市町コード集合がA38の三重県4圏域の構成市区町村の和集合と完全一致すること。"""
    csv_codes = {r["muni_code"] for r in rows}
    union_codes = {muni_code for munis in groups.values() for muni_code, _ in munis}
    if csv_codes != union_codes:
        raise ValueError(
            "網羅性検証に失敗: CSVの市町コード集合とA38の三重県4圏域の構成市区町村が一致しません。"
            f"CSVのみ={sorted(csv_codes - union_codes)} A38のみ={sorted(union_codes - csv_codes)}"
        )
    return {"csv_muni_count": len(csv_codes), "a38_union_muni_count": len(union_codes)}


def validate_uniqueness(rows) -> dict:
    """検証2(一意性): 各市町コードがCSV全体でちょうど1回だけ出現すること。"""
    counts = Counter(r["muni_code"] for r in rows)
    duplicates = {code: n for code, n in counts.items() if n != 1}
    if duplicates:
        raise ValueError(f"一意性検証に失敗: 重複した市町コードがあります: {duplicates}")
    return {"muni_count": len(rows)}


def validate_nesting(rows, code_to_owner: dict) -> dict:
    """検証3(入れ子の整合): 各行の parent_iryoken2_code が、A38の実際の市町→二次医療圏の割当と一致すること。"""
    mismatches = []
    for r in rows:
        actual = code_to_owner.get(r["muni_code"])
        if actual != r["parent_iryoken2_code"]:
            mismatches.append(
                {
                    "area_code": r["area_code"],
                    "muni_code": r["muni_code"],
                    "muni_name": r["muni_name"],
                    "csv_parent_iryoken2_code": r["parent_iryoken2_code"],
                    "actual_parent_iryoken2_code": actual,
                }
            )
    if mismatches:
        raise ValueError(f"入れ子の整合検証に失敗: {mismatches}")
    return {"checked": len(rows), "mismatches": mismatches}


def validate_name_match(rows, code_to_name: dict) -> dict:
    """検証4(名称の一致): CSVのmuni_nameがA38のA38b_002の表記と完全一致すること。"""
    mismatches = []
    for r in rows:
        expected = code_to_name.get(r["muni_code"])
        if expected != r["muni_name"]:
            mismatches.append(
                {
                    "area_code": r["area_code"],
                    "muni_code": r["muni_code"],
                    "csv_muni_name": r["muni_name"],
                    "a38_muni_name": expected,
                }
            )
    if mismatches:
        raise ValueError(f"名称の一致検証に失敗: {mismatches}")
    return {"checked": len(rows), "mismatches": mismatches}


# --- 検証5: 医療機関所在地との突合(独立した裏付け、生Excelを読む) --------------


def _scan_institution_municipalities(ws) -> set:
    """1シートの「(2)区域内の医療機関」表(13行目以降、B列=医療機関名、H列=所在地)から
    所在地の市町名集合を返す。空白行が2行連続したら表の終端とみなす。
    """
    munis = set()
    row = 13
    blank_streak = 0
    while True:
        a = ws.cell(row=row, column=1).value
        b = ws.cell(row=row, column=2).value
        h = ws.cell(row=row, column=8).value
        if a is None and b is None:
            blank_streak += 1
            if blank_streak >= 2:
                break
        else:
            blank_streak = 0
            if b is not None and h is not None:
                munis.add(str(h).strip())
        row += 1
        if row > 5000:
            raise RuntimeError("医療機関テーブルの終端を検出できませんでした(帳票レイアウトが変わった可能性)")
    return munis


def verify_against_institutions(rows, xlsx_path_in_repo: str = MIE_XLSX_PATH_IN_REPO) -> dict:
    """検証5: `R7/001723127.xlsx` の三重県8シートの医療機関所在地から市町名を集め、
    その市町がCSVで割り当てられている構想区域と一致することを検証する
    (突合(コード)とは独立した経路で得られる裏付け)。

    シートは339あるため `read_only=True` で開く。不一致があれば `ValueError` で中断する。
    """
    source_sha256 = verify_source(xlsx_path_in_repo)

    assigned = defaultdict(set)
    for r in rows:
        assigned[r["area_code"]].add(r["muni_name"])

    area_codes = sorted(assigned)
    wb = openpyxl.load_workbook(REPO_ROOT / xlsx_path_in_repo, read_only=True, data_only=True)
    try:
        verified_pairs = []
        mismatches = []
        for area_code in area_codes:
            sheet_name = next((n for n in wb.sheetnames if n.startswith(area_code)), None)
            if sheet_name is None:
                raise RuntimeError(f"{xlsx_path_in_repo}: 構想区域{area_code}のシートが見つかりません")
            ws = wb[sheet_name]
            munis_in_sheet = _scan_institution_municipalities(ws)
            for muni_name in sorted(munis_in_sheet):
                if muni_name in assigned[area_code]:
                    verified_pairs.append((area_code, muni_name))
                else:
                    mismatches.append(
                        {
                            "area_code": area_code,
                            "sheet": sheet_name,
                            "institution_muni_name": muni_name,
                            "csv_assigned_munis": sorted(assigned[area_code]),
                        }
                    )
    finally:
        wb.close()

    if mismatches:
        raise ValueError(f"医療機関所在地との突合検証に失敗: {mismatches}")

    return {
        "source_file": xlsx_path_in_repo,
        "source_sha256": source_sha256,
        "verified_pairs": sorted(verified_pairs),
        "verified_muni_count": len(verified_pairs),
        "mismatches": mismatches,
    }


# --- 検証6: 残存リスクの明示(裏付けの取れない市町のうち、真にリスクがあるもの) --


def compute_residual_risk(rows, mapping=MAPPING, verified_pairs=None) -> dict:
    """検証6: 医療機関所在地で裏付けが取れない市町を列挙し、そのうち旧圏域が分割
    される(2区域以上に分かれる)ケースだけを「真の残存リスク」として抽出する。

    旧圏域が分割されない(1対1)ケースは、医療機関裏付けが無くても市町の割当が
    一意に定まるため、残存リスクに含めない。
    """
    if verified_pairs is None:
        raise ValueError("verified_pairs は省略できません")
    verified_set = set(verified_pairs)

    parent_area_codes = defaultdict(set)
    for entry in mapping:
        parent_area_codes[entry["parent_code"]].add(entry["area_code"])
    split_parents = {parent for parent, areas in parent_area_codes.items() if len(areas) > 1}

    unverified = [r for r in rows if (r["area_code"], r["muni_name"]) not in verified_set]
    true_residual_risk = [r for r in unverified if r["parent_iryoken2_code"] in split_parents]
    unambiguous_unverified = [r for r in unverified if r["parent_iryoken2_code"] not in split_parents]

    def _brief(r):
        return {
            "area_code": r["area_code"],
            "area_name": r["area_name"],
            "muni_name": r["muni_name"],
            "parent_iryoken2_code": r["parent_iryoken2_code"],
            "parent_iryoken2_name": r["parent_iryoken2_name"],
        }

    return {
        "split_parent_iryoken2_codes": sorted(split_parents),
        "unverified_count": len(unverified),
        "unverified": [_brief(r) for r in unverified],
        "true_residual_risk_count": len(true_residual_risk),
        "true_residual_risk": [_brief(r) for r in true_residual_risk],
        "unambiguous_unverified_count": len(unambiguous_unverified),
        "unambiguous_unverified": [_brief(r) for r in unambiguous_unverified],
    }


def build_and_write(out_dir: Path) -> dict:
    """全処理(読み込み→対応表構築→検証1〜6→出力)を実行する。

    `out_dir` に `mie_area_municipalities.csv`(+`.meta.json`)を出力する。
    テスト(再現性の検証)から一時ディレクトリを渡して呼べるよう、出力先を
    パラメータ化してある(`tools/verify_area_join.py` の `build_and_write()` と同じ流儀)。

    戻り値: {"csv": ..., "meta": ...} の Path 辞書。
    """
    out_dir = Path(out_dir)

    pdf_sha256 = verify_source(MIE_PDF_PATH_IN_REPO)
    print(f"[ok] 一次資料の完全性検証: {MIE_PDF_PATH_IN_REPO} ({pdf_sha256[:12]}…)")

    groups, geo_metadata = load_a38_mie_groups()
    print(f"[ok] iryoken2_A38-20.geojson から三重県4圏域を読み込み: " + ", ".join(f"{k}({len(v)}市町)" for k, v in sorted(groups.items())))

    name_to_code, code_to_owner, code_to_name = build_global_lookup(groups)

    rows = build_rows(MAPPING, name_to_code)
    print(f"[ok] 対応表(MAPPING)の市町名をA38の市町コードで解決: {len(rows)}行")

    coverage = validate_coverage(rows, groups)
    print(f"[ok] 検証1(網羅性): CSV={coverage['csv_muni_count']}市町 == A38={coverage['a38_union_muni_count']}市町")

    uniqueness = validate_uniqueness(rows)
    print(f"[ok] 検証2(一意性): {uniqueness['muni_count']}市町すべて重複なし")

    nesting = validate_nesting(rows, code_to_owner)
    print(f"[ok] 検証3(入れ子の整合): {nesting['checked']}行すべてA38の実際の割当と一致")

    name_match = validate_name_match(rows, code_to_name)
    print(f"[ok] 検証4(名称の一致): {name_match['checked']}件すべてA38b_002表記と一致")

    institution = verify_against_institutions(rows)
    print(
        f"[ok] 検証5(医療機関所在地との突合、{institution['source_file']}): "
        f"{institution['verified_muni_count']}市町を裏付け、不一致{len(institution['mismatches'])}件"
    )

    residual = compute_residual_risk(rows, MAPPING, institution["verified_pairs"])
    print(
        f"[ok] 検証6(残存リスク): 裏付けのない市町{residual['unverified_count']}件のうち、"
        f"旧圏域が分割されるケース(真の残存リスク){residual['true_residual_risk_count']}件、"
        f"旧圏域が分割されない(裏付け不要)ケース{residual['unambiguous_unverified_count']}件"
    )

    header = [
        "area_code",
        "area_name",
        "muni_code",
        "muni_name",
        "parent_iryoken2_code",
        "parent_iryoken2_name",
    ]
    tuples = [
        (
            r["area_code"],
            r["area_name"],
            r["muni_code"],
            r["muni_name"],
            r["parent_iryoken2_code"],
            r["parent_iryoken2_name"],
        )
        for r in rows
    ]

    today = datetime.date.today().isoformat()
    source = {
        "name": "三重県公式資料(対応表) × A38-20(市町コード解決)",
        "inputs": [
            {
                "file": MIE_PDF_PATH_IN_REPO,
                "title": MIE_PDF_SOURCE["title"],
                "publisher": MIE_PDF_SOURCE["publisher"],
                "url": MIE_PDF_SOURCE["url"],
                "acquired_date": MIE_PDF_SOURCE["acquired_date"],
                "source_sha256": pdf_sha256,
                "reference_page": MIE_PDF_SOURCE["reference_page"],
                "role": "対応表(構想区域8区域×構成市町29市町×二次医療圏4圏域)の手転記元",
            },
            {
                "file": "data/processed/iryoken2_A38-20.geojson",
                "feature_count": geo_metadata["feature_count"],
                "source_file": geo_metadata["source"]["source_file"],
                "source_sha256": geo_metadata["source"]["source_sha256"],
                "role": "市町コード(muni_code)の解決元(A38b_001/A38b_002を市町名で照合)",
            },
        ],
        "license": (
            "三重県ウェブサイト利用規約 / "
            "国土数値情報ダウンロードサービス利用約款（オープンデータ）"
        ),
    }
    processing = {
        "script": "tools/build_mie_area_municipalities.py",
        "date": today,
        "steps": [
            "対応表(構想区域8区域→構成市町29市町→二次医療圏4圏域)はmie/001092203.pdf "
            "9ページ「現行の二次医療圏・構想区域」の表からの手転記(MAPPING定数)",
            "市町コード(muni_code)はiryoken2_A38-20.geojsonのA38b_001/A38b_002(構成市区町村の"
            "コード/名称)から市町名で解決(名称で引けなかった市町があれば中断)",
            f"検証1(網羅性): CSVの{coverage['csv_muni_count']}市町コードが、A38の三重県4圏域"
            f"(2401〜2404)の構成市区町村{coverage['a38_union_muni_count']}市町と完全一致することを確認",
            f"検証2(一意性): {uniqueness['muni_count']}市町コードすべてがCSV内でちょうど1回だけ出現することを確認",
            f"検証3(入れ子の整合): {nesting['checked']}行すべてでparent_iryoken2_codeが、A38が実際に"
            "その市町を割り当てている二次医療圏コードと一致することを確認(不一致0件)",
            f"検証4(名称の一致): {name_match['checked']}件すべてでmuni_nameがA38のA38b_002表記と"
            "完全一致することを確認(不一致0件)",
            f"検証5(独立した裏付け): R7/001723127.xlsx(339シート中、三重県8シート)の"
            "「(2)区域内の医療機関」表の②所在地列から市町名を収集し、"
            f"{institution['verified_muni_count']}市町がCSVの構想区域割当と一致することを確認"
            "(不一致0件。病床報告機関のない市町は裏付けの対象外)",
            f"検証6(残存リスク): 検証5で裏付けが取れない市町{residual['unverified_count']}件のうち、"
            f"旧圏域が複数の新区域に分割されるケースに該当する{residual['true_residual_risk_count']}件を"
            "真の残存リスクとして明示(旧圏域が分割されない残り"
            f"{residual['unambiguous_unverified_count']}件は、対応する新区域が一意に定まるため"
            "残存リスクに含めない)",
        ],
        "caveat": (
            "対応表(MAPPING)自体は一次資料PDFからの手転記であり、上記検証1〜5はいずれも"
            "『転記結果とA38・医療機関所在地との整合』を確認するものであって、PDF自体の"
            "OCR的な自動検証ではない。真の残存リスク(手転記のみに依拠する市町)は"
            "`verification.residual_risk.true_residual_risk` を参照。監査できるよう"
            "一次資料PDF(mie/001092203.pdf)をリポジトリに収載している。"
        ),
    }

    csv_path, meta_path = write_csv_with_meta(
        out_dir / "mie_area_municipalities.csv",
        header,
        tuples,
        title="三重県 構想区域 × 二次医療圏(令和2年度) × 構成市町 対応表",
        source=source,
        processing=processing,
        fields=FIELDS,
    )
    print(f"[ok] 出力: {csv_path} ({len(tuples)}行)")

    # write_csv_with_meta() の固定スキーマ(title/source/processing/fields/row_count)に
    # 加えて、検証1〜6の実測値を `verification` として追記する(tools/verify_area_join.py
    # が生Excelを読まずにレポートを再生成できるようにするため)。
    meta = _read_json(meta_path)
    row_count = meta.pop("row_count")
    meta["verification"] = {
        "coverage": coverage,
        "uniqueness": uniqueness,
        "nesting": nesting,
        "name_match": name_match,
        "institution_corroboration": institution,
        "residual_risk": residual,
    }
    meta["row_count"] = row_count
    with open(meta_path, "w", encoding="utf-8", newline="\n") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
        f.write("\n")
    print(f"[ok] 出力: {meta_path} (verification を追記)")

    return {"csv": csv_path, "meta": meta_path}


def main():
    build_and_write(OUT_CSV.parent)


if __name__ == "__main__":
    main()
